import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from datetime import date, datetime, timedelta
from io import BytesIO
from sqlalchemy import func, extract, distinct, case as sql_case
import os
import sys
import asyncio
import requests as _requests

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))

TRACKED_VENDEDOR_IDS = ["28856", "28886", "28887", "28891", "28892"]

BILLING_DOC_TYPES_G = ['Factura Electr.', 'Factura Exenta', 'Boleta Electr.']
NC_DOC_TYPES_G = ['Nota Credito']
# Notas de Debito: cargos adicionales al cliente. Suman positivo en ventas y
# en cartera (igual que Facturas). Ver excel_generator.py para detalle.
ND_DOC_TYPES_G = ['Nota Debito']
VALID_DOC_TYPES_G = BILLING_DOC_TYPES_G + NC_DOC_TYPES_G + ND_DOC_TYPES_G

from src.database import SessionLocal, engine, Base
from src.models.models import (
    VentaHistorico, CompraHistorico, Producto, ContabilidadHistorico,
    ClienteFinal, SyncLog, ReporteGenerado, ObumaApiEndpoint,
    Proveedor, ClienteContacto, ClienteDireccion, Empleado, Remuneracion,
    VentaItem, VentaCotizacion, VentaCobro, VentaDte, CompraOC, CompraPago,
    CrmLead, ProductoCategoria,
    ProductoSubCategoria, ProductoFabricante, ProductoPrecio, CostoHistorico,
    VendedorMeta, VendedorCartera
)
from src.etl.sync_service import SyncService
from src.etl.obuma_client import ObumaClient
from src.reports.excel_generator import generate_vendedor_report, generate_all_vendedor_reports
from src.reports.email_service import send_report_email, build_report_email_html, check_email_config, test_email_delivery, check_admin_alert_config
from src.models.models import ReporteProgramado
from src.utils.date_filters import date_range_filters, year_month_range

@st.cache_resource
def _ensure_schema_once():
    """Crea las tablas solo UNA vez por proceso de Streamlit (no en cada
    interaccion). Antes esto corria en cada rerun (cada filtro, cada
    cambio de pestana) lo cual ejecutaba un MetaData reflection completo
    contra Postgres y agregaba latencia gratis a cada accion del usuario."""
    Base.metadata.create_all(bind=engine)
    return True


_ensure_schema_once()

st.set_page_config(
    page_title="BI Platform - Gabriel Hoyos",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ── LOGIN ──────────────────────────────────────────────────────────────────
import hashlib as _hashlib

_USERS = {}
_u1 = os.environ.get("DASHBOARD_USER", "")
_p1 = os.environ.get("DASHBOARD_PASSWORD", "")
if _u1 and _p1:
    _USERS[_u1] = {"password": _p1, "role": "user"}
_u2 = os.environ.get("ADMIN_USER", "")
_p2 = os.environ.get("ADMIN_PASSWORD", "")
if _u2 and _p2:
    _USERS[_u2] = {"password": _p2, "role": "admin"}

_AUTH_TOKENS = {}
for _u, _info in _USERS.items():
    _tok = _hashlib.sha256(f"{_u}:{_info['password']}:bi_platform_2026".encode()).hexdigest()[:24]
    _AUTH_TOKENS[_tok] = {"user": _u, "role": _info["role"]}

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False
if "user_role" not in st.session_state:
    st.session_state.user_role = None
if "user_name" not in st.session_state:
    st.session_state.user_name = None

if not st.session_state.authenticated:
    _param_token = st.query_params.get("auth")
    if _param_token and _param_token in _AUTH_TOKENS:
        st.session_state.authenticated = True
        st.session_state.user_role = _AUTH_TOKENS[_param_token]["role"]
        st.session_state.user_name = _AUTH_TOKENS[_param_token]["user"]
        st.rerun()

if not st.session_state.authenticated:
    st.markdown("""
    <style>
    div[data-testid="stSidebar"] {display: none;}
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)

    col_l, col_c, col_r = st.columns([1, 1.2, 1])
    with col_c:
        st.markdown("<br><br>", unsafe_allow_html=True)
        st.markdown("""
        <div style="background:#1a1f2e;border:1px solid #2d3548;border-radius:16px;padding:40px 36px;text-align:center;">
            <h1 style="color:#e2e8f0;font-size:1.6rem;margin-bottom:4px;">📊 BI Platform</h1>
            <p style="color:#94a3b8;font-size:0.9rem;margin-bottom:24px;">VLSur - Acceso Privado</p>
        </div>
        """, unsafe_allow_html=True)

        with st.form("login_form"):
            user_input = st.text_input("Usuario", placeholder="usuario")
            pass_input = st.text_input("Contraseña", type="password", placeholder="••••••••")
            submitted = st.form_submit_button("Ingresar", use_container_width=True, type="primary")

        if submitted:
            _matched = _USERS.get(user_input.strip())
            if _matched and pass_input == _matched["password"]:
                _tok = _hashlib.sha256(f"{user_input.strip()}:{pass_input}:bi_platform_2026".encode()).hexdigest()[:24]
                st.session_state.authenticated = True
                st.session_state.user_role = _matched["role"]
                st.session_state.user_name = user_input.strip()
                st.query_params["auth"] = _tok
                st.rerun()
            else:
                st.error("Usuario o contraseña incorrectos.")

    st.stop()
# ── FIN LOGIN ──────────────────────────────────────────────────────────────

DARK_BG = "#0e1117"
CARD_BG = "#1a1f2e"
CARD_BORDER = "#2d3548"
TEXT_PRIMARY = "#e2e8f0"
TEXT_SECONDARY = "#94a3b8"
ACCENT_BLUE = "#3b82f6"
ACCENT_GREEN = "#10b981"
ACCENT_RED = "#ef4444"
ACCENT_AMBER = "#f59e0b"
ACCENT_PURPLE = "#8b5cf6"

st.markdown(f"""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');

    .stApp {{
        font-family: 'Inter', sans-serif;
    }}

    div[data-testid="stSidebar"] {{
        background: linear-gradient(180deg, #0f172a 0%, #1e293b 100%);
    }}
    div[data-testid="stSidebar"] .stMarkdown p,
    div[data-testid="stSidebar"] .stMarkdown h1,
    div[data-testid="stSidebar"] .stMarkdown h2,
    div[data-testid="stSidebar"] .stMarkdown h3 {{
        color: #e2e8f0;
    }}
    div[data-testid="stSidebar"] .stRadio label {{
        color: #cbd5e1 !important;
    }}
    div[data-testid="stSidebar"] .stRadio label:hover {{
        color: #ffffff !important;
    }}
    div[data-testid="stSidebar"] hr {{
        border-color: rgba(255,255,255,0.08);
    }}

    div[data-testid="stMetric"] {{
        background: {CARD_BG};
        border: 1px solid {CARD_BORDER};
        border-radius: 12px;
        padding: 1rem 1.2rem;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.3);
    }}
    div[data-testid="stMetric"] label {{
        color: {TEXT_SECONDARY} !important;
    }}
    div[data-testid="stMetric"] [data-testid="stMetricValue"] {{
        color: {TEXT_PRIMARY} !important;
    }}

    .metric-card {{
        background: {CARD_BG};
        border: 1px solid {CARD_BORDER};
        border-radius: 12px;
        padding: 1.2rem 1.5rem;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.3);
        transition: all 0.2s ease;
    }}
    .metric-card:hover {{
        border-color: {ACCENT_BLUE};
        box-shadow: 0 8px 15px -3px rgba(59,130,246,0.15);
    }}
    .metric-value {{
        font-size: 1.8rem;
        font-weight: 700;
        color: {TEXT_PRIMARY};
        margin: 0;
        line-height: 1.2;
    }}
    .metric-label {{
        font-size: 0.78rem;
        font-weight: 500;
        color: {TEXT_SECONDARY};
        text-transform: uppercase;
        letter-spacing: 0.8px;
        margin: 0 0 0.4rem 0;
    }}
    .metric-icon {{
        font-size: 1.5rem;
        margin-bottom: 0.3rem;
    }}
    .metric-delta {{
        font-size: 0.8rem;
        font-weight: 600;
        margin-top: 0.3rem;
    }}

    .section-header {{
        font-size: 1.1rem;
        font-weight: 600;
        color: {TEXT_PRIMARY};
        margin: 1.5rem 0 0.8rem 0;
        padding-bottom: 0.5rem;
        border-bottom: 2px solid {CARD_BORDER};
    }}

    .page-title {{
        font-size: 1.8rem;
        font-weight: 700;
        color: {TEXT_PRIMARY};
        margin-bottom: 0.2rem;
        letter-spacing: -0.5px;
    }}
    .page-subtitle {{
        font-size: 0.95rem;
        color: {TEXT_SECONDARY};
        margin-bottom: 1.5rem;
    }}

    .data-card {{
        background: {CARD_BG};
        border: 1px solid {CARD_BORDER};
        border-radius: 12px;
        padding: 1.5rem;
        box-shadow: 0 4px 6px -1px rgba(0,0,0,0.3);
    }}

    .status-badge {{
        display: inline-block;
        padding: 0.25rem 0.75rem;
        border-radius: 20px;
        font-size: 0.75rem;
        font-weight: 600;
    }}
    .badge-ok {{ background: rgba(16,185,129,0.15); color: #34d399; }}
    .badge-error {{ background: rgba(239,68,68,0.15); color: #f87171; }}
    .badge-warning {{ background: rgba(245,158,11,0.15); color: #fbbf24; }}
    .badge-info {{ background: rgba(59,130,246,0.15); color: #60a5fa; }}

    .stDataFrame {{
        border-radius: 8px;
        overflow: hidden;
    }}
    .stDataFrame [data-testid="stDataFrameResizable"] {{
        border: 1px solid {CARD_BORDER};
        border-radius: 8px;
    }}

    .stButton > button[kind="primary"] {{
        background: linear-gradient(135deg, {ACCENT_BLUE}, #2563eb);
        border: none;
        border-radius: 8px;
        padding: 0.5rem 1.5rem;
        font-weight: 600;
        color: white;
    }}
    .stButton > button {{
        border-radius: 8px;
        font-weight: 500;
        border-color: {CARD_BORDER};
    }}

    .sync-result-ok {{
        background: rgba(16,185,129,0.1);
        border-left: 4px solid {ACCENT_GREEN};
        padding: 0.8rem 1rem;
        border-radius: 0 8px 8px 0;
        margin: 0.3rem 0;
        color: {TEXT_PRIMARY};
    }}
    .sync-result-error {{
        background: rgba(239,68,68,0.1);
        border-left: 4px solid {ACCENT_RED};
        padding: 0.8rem 1rem;
        border-radius: 0 8px 8px 0;
        margin: 0.3rem 0;
        color: {TEXT_PRIMARY};
    }}

    .api-card {{
        background: {CARD_BG};
        border: 1px solid {CARD_BORDER};
        border-radius: 10px;
        padding: 1rem 1.2rem;
        margin: 0.4rem 0;
        box-shadow: 0 2px 4px rgba(0,0,0,0.2);
    }}

    .stTabs [data-baseweb="tab-list"] {{
        gap: 8px;
    }}
    .stTabs [data-baseweb="tab"] {{
        border-radius: 8px;
        padding: 8px 16px;
    }}

    div.stAlert {{
        border-radius: 8px;
    }}
</style>
""", unsafe_allow_html=True)


def get_db():
    return SessionLocal()


# ════════════════════════════════════════════════════════════════════════
# CACHE LAYER (Fase 1 de optimizacion).
#
# Streamlit re-ejecuta TODO el script en cada interaccion del usuario
# (cambio de filtro, cambio de pestana, click en grafico). Antes el
# dashboard repetia ~10-20 consultas a Postgres en cada rerun.
#
# Estos helpers encapsulan las consultas pesadas y las cachean por 5 min.
# Reglas importantes:
#   - NO aceptar Session como parametro (no es hashable).
#   - Crear/cerrar la propia Session adentro.
#   - Devolver tipos primitivos (dict, list, tuple), NO objetos ORM
#     (los objetos detached entre reruns rompen lazy-loading).
#   - vendor_ids debe llegar como tuple (hashable).
#
# TTL=300s significa que un cambio reciente en BD se ve maximo 5 min
# despues. Si el usuario quiere ver datos frescos antes, puede hacer
# "Sync ahora" desde la sidebar (eso invalida sesiones).
# ════════════════════════════════════════════════════════════════════════

CACHE_TTL_SECONDS = 300


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _cached_load_empleados():
    """Lista de empleados activos (usada en multiselects de varias pestanas).
    Devuelve dicts simples para evitar objetos ORM detached entre reruns."""
    db = SessionLocal()
    try:
        rows = db.query(Empleado).filter(Empleado.activo == True).order_by(Empleado.nombre).all()
        return [
            {"obuma_id": e.obuma_id, "nombre": e.nombre, "cargo": e.cargo}
            for e in rows
        ]
    finally:
        db.close()


def _apply_dash_filters_q(query, date_from, date_to, vendor_ids, model=None):
    """Helper interno: aplica filtros de fecha + vendor_ids a una query."""
    if model is None:
        model = VentaHistorico
    query = query.filter(*date_range_filters(model.fecha, date_from, date_to))
    if vendor_ids and model is VentaHistorico:
        query = query.filter(model.vendedor_id.in_(list(vendor_ids)))
    return query


def _neto_sum_expr():
    """Expresion SQL: NC resta del neto, ND/Facturas suman positivo."""
    return func.sum(
        sql_case(
            (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -VentaHistorico.subtotal),
            else_=VentaHistorico.subtotal
        )
    )


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _cached_dashboard_kpis(date_from, date_to, vendor_ids):
    """KPIs del centro de mando. Antes eran ~9 consultas separadas en cada rerun."""
    db = SessionLocal()
    try:
        base_q = _apply_dash_filters_q(
            db.query(VentaHistorico).filter(
                VentaHistorico.anulada == False,
                VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G)
            ),
            date_from, date_to, vendor_ids,
        )
        total_ventas = base_q.with_entities(_neto_sum_expr()).scalar() or 0
        total_margen = base_q.with_entities(func.sum(VentaHistorico.margen_neto)).scalar() or 0
        n_ventas = base_q.count()
        ventas_anuladas = _apply_dash_filters_q(
            db.query(VentaHistorico).filter(VentaHistorico.anulada == True),
            date_from, date_to, vendor_ids,
        ).count()
        total_pagado = base_q.with_entities(func.sum(VentaHistorico.total_pagado)).scalar() or 0
        total_por_pagar = base_q.with_entities(func.sum(VentaHistorico.total_por_pagar)).scalar() or 0

        compras_q = db.query(func.sum(CompraHistorico.total)).filter(
            *date_range_filters(CompraHistorico.fecha, date_from, date_to)
        )
        total_compras = compras_q.scalar() or 0

        n_clientes = db.query(ClienteFinal).filter(ClienteFinal.activo == True).count()
        n_productos = db.query(Producto).filter(Producto.activo == True).count()
        total_cobros_count = base_q.with_entities(
            func.count(VentaHistorico.id)
        ).filter(VentaHistorico.total_pagado > 0).scalar() or 0

        return {
            "total_ventas": float(total_ventas or 0),
            "total_margen": float(total_margen or 0),
            "n_ventas": int(n_ventas or 0),
            "ventas_anuladas": int(ventas_anuladas or 0),
            "total_pagado": float(total_pagado or 0),
            "total_por_pagar": float(total_por_pagar or 0),
            "total_compras": float(total_compras or 0),
            "n_clientes": int(n_clientes or 0),
            "n_productos": int(n_productos or 0),
            "total_cobros_count": int(total_cobros_count or 0),
        }
    finally:
        db.close()


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _cached_dashboard_charts(date_from, date_to, vendor_ids):
    """Datos de los 8 graficos del Dashboard. Devuelve listas de tuplas
    (datos primitivos) para que Streamlit pueda cachear sin problemas."""
    db = SessionLocal()
    try:
        # 1. Ventas mensuales (con filtro de tipos validos + no anuladas)
        monthly_sales = _apply_dash_filters_q(
            db.query(
                extract('year', VentaHistorico.fecha).label("anio"),
                extract('month', VentaHistorico.fecha).label("mes"),
                _neto_sum_expr().label("total"),
            ).filter(
                VentaHistorico.fecha.isnot(None),
                VentaHistorico.anulada == False,
                VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
            ),
            date_from, date_to, vendor_ids,
        ).group_by("anio", "mes").order_by("anio", "mes").all()

        # 2. Top 10 vendedores
        top_vend = _apply_dash_filters_q(
            db.query(
                VentaHistorico.vendedor_id,
                _neto_sum_expr().label("total"),
            ).filter(
                VentaHistorico.vendedor_id.isnot(None),
                VentaHistorico.fecha.isnot(None),
                VentaHistorico.anulada == False,
                VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
            ),
            date_from, date_to, vendor_ids,
        ).group_by(VentaHistorico.vendedor_id).order_by(_neto_sum_expr().desc()).limit(10).all()

        # 3. Segmentacion ABC de clientes
        client_sales = _apply_dash_filters_q(
            db.query(
                VentaHistorico.cliente_id,
                _neto_sum_expr().label("total"),
            ).filter(
                VentaHistorico.cliente_id.isnot(None),
                VentaHistorico.fecha.isnot(None),
                VentaHistorico.anulada == False,
                VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
            ),
            date_from, date_to, vendor_ids,
        ).group_by(VentaHistorico.cliente_id).order_by(_neto_sum_expr().desc()).all()

        # 4. Rentabilidad por vendedor (sin filtros de doc type, como el original)
        rent_vend = _apply_dash_filters_q(
            db.query(
                VentaHistorico.vendedor_id,
                func.sum(VentaHistorico.margen_neto).label("margen"),
            ).filter(
                VentaHistorico.vendedor_id.isnot(None),
                VentaHistorico.fecha.isnot(None),
            ),
            date_from, date_to, vendor_ids,
        ).group_by(VentaHistorico.vendedor_id).order_by(
            func.sum(VentaHistorico.margen_neto).desc()
        ).limit(10).all()

        # 5. Ventas mensuales (alias para chart 7) y compras mensuales
        compras_q_monthly = db.query(
            extract('year', CompraHistorico.fecha).label("anio"),
            extract('month', CompraHistorico.fecha).label("mes"),
            func.sum(CompraHistorico.total).label("total"),
        ).filter(
            CompraHistorico.fecha.isnot(None),
            *date_range_filters(CompraHistorico.fecha, date_from, date_to),
        )
        monthly_compras = compras_q_monthly.group_by("anio", "mes").order_by("anio", "mes").all()

        # 6. Top 15 productos vendidos
        top_products_q = db.query(
            VentaItem.producto_nombre,
            func.sum(VentaItem.cantidad).label("cantidad"),
            func.sum(VentaItem.total).label("total"),
        ).join(
            VentaHistorico, VentaHistorico.obuma_id == VentaItem.venta_id_obuma,
        ).filter(VentaItem.producto_nombre.isnot(None))
        top_products_q = _apply_dash_filters_q(
            top_products_q, date_from, date_to, vendor_ids, model=VentaHistorico,
        )
        top_products = top_products_q.group_by(VentaItem.producto_nombre).order_by(
            func.sum(VentaItem.total).desc()
        ).limit(15).all()

        # 7. Distribucion por tipo de documento
        doc_types = _apply_dash_filters_q(
            db.query(
                VentaHistorico.tipo_documento,
                func.count(VentaHistorico.id).label("cantidad"),
                func.sum(VentaHistorico.subtotal).label("total"),
            ).filter(
                VentaHistorico.tipo_documento.isnot(None),
                VentaHistorico.fecha.isnot(None),
                VentaHistorico.anulada == False,
                VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
            ),
            date_from, date_to, vendor_ids,
        ).group_by(VentaHistorico.tipo_documento).order_by(
            func.sum(VentaHistorico.subtotal).desc()
        ).all()

        return {
            "monthly_sales": [(int(r.anio), int(r.mes), float(r.total or 0)) for r in monthly_sales],
            "top_vend": [(v.vendedor_id, float(v.total or 0)) for v in top_vend],
            "client_sales": [(c.cliente_id, float(c.total or 0)) for c in client_sales],
            "rent_vend": [(v.vendedor_id, float(v.margen or 0)) for v in rent_vend],
            "monthly_compras": [(int(r.anio), int(r.mes), float(r.total or 0)) for r in monthly_compras],
            "top_products": [
                ((p.producto_nombre or "")[:30], float(p.total or 0)) for p in top_products
            ],
            "doc_types": [(d.tipo_documento or "Otro", float(d.total or 0), int(d.cantidad or 0)) for d in doc_types],
        }
    finally:
        db.close()


@st.cache_data(ttl=CACHE_TTL_SECONDS, show_spinner=False)
def _cached_dashboard_recent_and_top(date_from, date_to, vendor_ids):
    """Tablas inferiores del Dashboard: ultimas transacciones + top clientes."""
    db = SessionLocal()
    try:
        recent_q = _apply_dash_filters_q(
            db.query(VentaHistorico).filter(VentaHistorico.fecha.isnot(None)),
            date_from, date_to, vendor_ids,
        )
        recent_ventas = recent_q.order_by(VentaHistorico.fecha.desc()).limit(15).all()
        recent_data = [
            {
                "fecha": str(v.fecha)[:10] if v.fecha else "-",
                "tipo": v.tipo_documento or "-",
                "folio": v.folio or "-",
                "total": float(v.total or 0),
                "vendedor_id": v.vendedor_id,
                "anulada": bool(v.anulada),
            }
            for v in recent_ventas
        ]

        top_clientes_q = db.query(
            ClienteFinal.nombre,
            ClienteFinal.rut,
            _neto_sum_expr().label("total"),
            func.sum(VentaHistorico.margen_neto).label("margen"),
            func.count(VentaHistorico.id).label("transacciones"),
        ).join(
            VentaHistorico, VentaHistorico.cliente_id == ClienteFinal.id,
        ).filter(
            VentaHistorico.fecha.isnot(None),
            VentaHistorico.anulada == False,
            VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
        )
        top_clientes_q = _apply_dash_filters_q(
            top_clientes_q, date_from, date_to, vendor_ids, model=VentaHistorico,
        )
        top_clientes = top_clientes_q.group_by(
            ClienteFinal.nombre, ClienteFinal.rut,
        ).order_by(_neto_sum_expr().desc()).limit(15).all()
        top_clientes_data = [
            {
                "nombre": c.nombre,
                "rut": c.rut or "-",
                "total": float(c.total or 0),
                "margen": float(c.margen or 0),
                "transacciones": int(c.transacciones or 0),
            }
            for c in top_clientes
        ]

        return {"recent": recent_data, "top_clientes": top_clientes_data}
    finally:
        db.close()


def run_async(coro):
    loop = asyncio.new_event_loop()
    asyncio.set_event_loop(loop)
    try:
        return loop.run_until_complete(coro)
    finally:
        loop.close()


_FASTAPI_BASE = "http://127.0.0.1:8000"


def _api_sync_start():
    try:
        r = _requests.post(f"{_FASTAPI_BASE}/api/sync/start", timeout=5)
        return r.json()
    except Exception:
        return {"status": "error"}


def _api_sync_status():
    try:
        r = _requests.get(f"{_FASTAPI_BASE}/api/sync/status", timeout=5)
        return r.json()
    except Exception:
        return {"running": False, "done": False, "step": 0, "total": 0, "label": "", "results": []}


def _api_sync_reset():
    try:
        _requests.post(f"{_FASTAPI_BASE}/api/sync/reset", timeout=5)
    except Exception:
        pass


def _run_full_sync_ui():
    """Non-blocking sync UI using FastAPI background task + st.fragment polling."""
    if not st.session_state.get("_sync_started"):
        status = _api_sync_status()
        if not status.get("running") and not status.get("done"):
            _api_sync_start()
        st.session_state._sync_started = True

    _sync_progress_fragment()


@st.fragment(run_every=timedelta(seconds=3))
def _sync_progress_fragment():
    """Fragment that auto-refreshes every 3s to poll FastAPI sync status."""
    progress = _api_sync_status()

    if progress.get("running"):
        step = progress.get("step", 0)
        total = progress.get("total", 8)
        label = progress.get("label", "...")
        results_so_far = progress.get("results", [])
        pct = step / total if total > 0 else 0

        st.markdown(f"""
        <div style="text-align:center;padding:24px 0 8px;">
            <span style="font-size:2.5rem;">🔄</span>
            <h1 style="color:{TEXT_PRIMARY};font-size:1.8rem;margin:8px 0 4px;">Actualizando Sistema</h1>
            <p style="color:{TEXT_SECONDARY};">Sincronizando todos los módulos desde Obuma ERP...</p>
        </div>
        """, unsafe_allow_html=True)

        st.progress(pct, text=f"Paso {step+1}/{total}: {label}")
        st.markdown(f"""
        <div style="background:{CARD_BG};border:1px solid {ACCENT_BLUE};border-radius:12px;
                    padding:18px 24px;margin:12px 0;display:flex;align-items:center;gap:16px;">
            <span style="font-size:1.8rem;">⏳</span>
            <div>
                <p style="color:{ACCENT_BLUE};font-weight:700;font-size:1.05rem;margin:0;">{label}</p>
                <p style="color:{TEXT_SECONDARY};font-size:0.85rem;margin:4px 0 0;">
                    Consultando API de Obuma... ({int(pct*100)}% completado)
                </p>
            </div>
        </div>
        """, unsafe_allow_html=True)

        if results_so_far:
            rows = ""
            for r in results_so_far:
                if r.get("ok"):
                    rows += f'<div style="color:{ACCENT_GREEN};font-size:0.85rem;padding:2px 0;">✅ {r["label"]}: {r.get("synced",0):,} registros</div>'
                else:
                    rows += f'<div style="color:{ACCENT_RED};font-size:0.85rem;padding:2px 0;">❌ {r["label"]}: {r.get("error","")}</div>'
            st.markdown(f'<div style="background:{CARD_BG};border:1px solid {CARD_BORDER};border-radius:8px;padding:12px 16px;margin-top:12px;">{rows}</div>', unsafe_allow_html=True)

    elif progress.get("done"):
        results = progress.get("results", [])
        ok_count = sum(1 for r in results if r.get("ok"))
        total_synced = sum(r.get("synced", 0) for r in results if r.get("ok"))

        st.markdown(f"""
        <div style="text-align:center;padding:24px 0 8px;">
            <span style="font-size:2.5rem;">✅</span>
            <h1 style="color:{ACCENT_GREEN};font-size:1.8rem;margin:8px 0 4px;">¡Actualización Completada!</h1>
            <p style="color:{TEXT_SECONDARY};">{ok_count}/{len(results)} módulos actualizados · {total_synced:,} registros totales</p>
        </div>
        """, unsafe_allow_html=True)

        rows_html = ""
        for r in results:
            if r.get("ok"):
                rows_html += f"""
                <tr>
                    <td style="padding:8px 12px;color:{TEXT_PRIMARY};">{r['label']}</td>
                    <td style="padding:8px 12px;color:{ACCENT_GREEN};font-weight:600;">✅ OK</td>
                    <td style="padding:8px 12px;color:{TEXT_SECONDARY};text-align:right;">{r.get('synced', 0):,} registros</td>
                </tr>"""
            else:
                rows_html += f"""
                <tr>
                    <td style="padding:8px 12px;color:{TEXT_PRIMARY};">{r['label']}</td>
                    <td style="padding:8px 12px;color:{ACCENT_RED};font-weight:600;">❌ Error</td>
                    <td style="padding:8px 12px;color:{ACCENT_RED};font-size:0.82rem;">{r.get('error','')}</td>
                </tr>"""

        st.markdown(f"""
        <div style="background:{CARD_BG};border:1px solid {CARD_BORDER};border-radius:12px;padding:20px 24px;margin:16px 0;">
            <table style="width:100%;border-collapse:collapse;">
                <thead><tr style="border-bottom:1px solid {CARD_BORDER};">
                    <th style="padding:6px 12px;color:{TEXT_SECONDARY};text-align:left;">Módulo</th>
                    <th style="padding:6px 12px;color:{TEXT_SECONDARY};text-align:left;">Estado</th>
                    <th style="padding:6px 12px;color:{TEXT_SECONDARY};text-align:right;">Registros</th>
                </tr></thead>
                <tbody>{rows_html}</tbody>
            </table>
        </div>
        """, unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)
        if st.button("📊 Ver Dashboard Actualizado", type="primary", use_container_width=True, key="btn_goto_dashboard"):
            _api_sync_reset()
            st.session_state.syncing_now = False
            st.session_state._sync_started = False
            st.session_state.sync_done = True
            # Sync recien terminado: invalida el cache de KPIs/graficos para
            # que el usuario vea datos frescos sin esperar al TTL de 5 min.
            st.cache_data.clear()
            st.rerun()

    else:
        st.markdown(f"""
        <div style="text-align:center;padding:48px 0;">
            <span style="font-size:2rem;">⏳</span>
            <p style="color:{TEXT_SECONDARY};">Conectando con el servidor...</p>
        </div>
        """, unsafe_allow_html=True)


def format_clp(value):
    if value is None:
        return "$0"
    return f"${value:,.0f}".replace(",", ".")


def render_metric(label, value, icon="", color=TEXT_PRIMARY, delta=None, delta_color=None):
    delta_html = ""
    if delta is not None:
        dc = delta_color or (ACCENT_GREEN if str(delta).startswith("+") or (isinstance(delta, (int, float)) and delta > 0) else ACCENT_RED)
        delta_html = f'<p class="metric-delta" style="color:{dc};">{delta}</p>'
    st.markdown(f"""
    <div class="metric-card">
        <div class="metric-icon">{icon}</div>
        <p class="metric-label">{label}</p>
        <p class="metric-value" style="color:{color};">{value}</p>
        {delta_html}
    </div>
    """, unsafe_allow_html=True)


def chart_layout(height=340):
    return dict(
        height=height,
        margin=dict(t=20, b=40, l=50, r=20),
        legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1,
                    font=dict(color=TEXT_SECONDARY)),
        plot_bgcolor="rgba(0,0,0,0)",
        paper_bgcolor="rgba(0,0,0,0)",
        font=dict(family="Inter", color=TEXT_SECONDARY),
        xaxis=dict(showgrid=False, color=TEXT_SECONDARY),
        yaxis=dict(showgrid=True, gridcolor="rgba(45,53,72,0.5)", color=TEXT_SECONDARY),
    )


MONTH_LABELS = ["Ene", "Feb", "Mar", "Abr", "May", "Jun", "Jul", "Ago", "Sep", "Oct", "Nov", "Dic"]
CHART_COLORS = [ACCENT_BLUE, ACCENT_GREEN, ACCENT_AMBER, ACCENT_RED, ACCENT_PURPLE,
                "#06b6d4", "#ec4899", "#84cc16", "#f97316", "#6366f1"]


st.sidebar.markdown("### 📊 BI Platform")
st.sidebar.markdown("**Gabriel Hoyos**")
st.sidebar.caption("Centro de Mando Empresarial")
st.sidebar.markdown("---")

st.sidebar.markdown("""
<style>
div[data-testid="stSidebar"] button[kind="primary"] {
    background: linear-gradient(135deg, #10b981, #059669) !important;
    color: white !important;
    font-weight: 700 !important;
    font-size: 0.95rem !important;
    border: none !important;
    border-radius: 10px !important;
    padding: 10px !important;
    letter-spacing: 0.03em !important;
    box-shadow: 0 0 12px rgba(16,185,129,0.35) !important;
}
div[data-testid="stSidebar"] button[kind="primary"]:hover {
    background: linear-gradient(135deg, #059669, #047857) !important;
    box-shadow: 0 0 20px rgba(16,185,129,0.55) !important;
    transform: translateY(-1px);
}
</style>
""", unsafe_allow_html=True)

if st.sidebar.button("🔄 ACTUALIZAR AHORA", use_container_width=True, type="primary", key="btn_actualizar_sidebar"):
    status = _api_sync_status()
    if not status.get("running"):
        st.session_state.syncing_now = True
        st.session_state._sync_started = False
        st.rerun()

st.sidebar.markdown("---")

page = st.sidebar.radio(
    "Navegacion",
    ["Dashboard", "Vendedores", "Ventas", "Clientes", "Proveedores", "Productos",
     "Empleados", "Compras", "Contabilidad", "CRM",
     "API Obuma", "Reportes", "Sincronizacion", "Auditoria"],
    index=0,
    label_visibility="collapsed"
)

st.sidebar.markdown("---")

_admin_alert_cfg = check_admin_alert_config()
if _admin_alert_cfg["configured"]:
    _alert_emails_str = ", ".join(_admin_alert_cfg["emails"])
    st.sidebar.markdown(
        f"""
        <div style="padding:8px 10px;border-radius:8px;background:rgba(16,185,129,0.10);
                    border-left:3px solid #10b981;font-size:0.78rem;line-height:1.35;">
            <div style="font-weight:700;color:#10b981;">Alertas admin: ON</div>
            <div style="color:#94a3b8;margin-top:2px;" title="{_alert_emails_str}">
                {len(_admin_alert_cfg["emails"])} destinatario{'' if len(_admin_alert_cfg["emails"]) == 1 else 's'}
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )
else:
    st.sidebar.markdown(
        """
        <div style="padding:8px 10px;border-radius:8px;background:rgba(239,68,68,0.10);
                    border-left:3px solid #ef4444;font-size:0.78rem;line-height:1.35;">
            <div style="font-weight:700;color:#ef4444;">Alertas admin: OFF</div>
            <div style="color:#94a3b8;margin-top:2px;">
                Define <code>ADMIN_ALERT_EMAILS</code> para recibir avisos
                cuando un reporte automatico no se envie.
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

st.sidebar.markdown("---")
st.sidebar.caption("v2.0 | Powered by Obuma ERP")
if st.sidebar.button("🔒 Cerrar Sesión", use_container_width=True):
    st.session_state.authenticated = False
    st.session_state.user_role = None
    st.session_state.user_name = None
    st.query_params.clear()
    st.rerun()


# ============================================================
# ACTUALIZAR AHORA - Intercepta toda la navegación si está activo
# ============================================================
_active_sync = _api_sync_status()
if st.session_state.get("syncing_now", False) or _active_sync.get("running") or _active_sync.get("done"):
    _run_full_sync_ui()
    st.stop()

# ============================================================
# DASHBOARD
# ============================================================
if page == "Dashboard":
    st.markdown('<p class="page-title">Centro de Mando</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Vista consolidada del rendimiento empresarial</p>', unsafe_allow_html=True)

    # Filtros globales (UI rapida, no consulta BD)
    st.markdown('<p class="section-header">Filtros Globales</p>', unsafe_allow_html=True)
    fc1, fc2, fc3 = st.columns([2, 2, 3])
    with fc1:
        dash_date_from = st.date_input("Desde", value=date.today() - timedelta(days=365), key="dash_from")
    with fc2:
        dash_date_to = st.date_input("Hasta", value=date.today(), key="dash_to")
    with fc3:
        # Lista de empleados cacheada (5 min) — antes era una query nueva
        # en cada interaccion del filtro.
        empleados_list = _cached_load_empleados()
        vendedor_names = {e["obuma_id"]: e["nombre"] for e in empleados_list}
        tracked_empleados = [e for e in empleados_list if e["obuma_id"] in TRACKED_VENDEDOR_IDS]
        other_empleados = [e for e in empleados_list if e["obuma_id"] not in TRACKED_VENDEDOR_IDS]
        vendedor_options_all = ["Todos (5 Vendedores)"] + [
            f"{e['nombre']} ({e['cargo'] or 'Sin cargo'})" for e in tracked_empleados
        ]
        if other_empleados:
            vendedor_options_all += ["---"] + [
                f"{e['nombre']} ({e['cargo'] or 'Sin cargo'})" for e in other_empleados
            ]
        vendedor_sel_dash = st.multiselect(
            "Vendedores", vendedor_options_all,
            default=["Todos (5 Vendedores)"], key="dash_vend",
        )

    selected_vendedor_ids = []
    if "Todos (5 Vendedores)" in vendedor_sel_dash:
        selected_vendedor_ids = list(TRACKED_VENDEDOR_IDS)
    elif vendedor_sel_dash:
        for sel in vendedor_sel_dash:
            if sel == "---":
                continue
            for e in empleados_list:
                if f"{e['nombre']} ({e['cargo'] or 'Sin cargo'})" == sel:
                    selected_vendedor_ids.append(e["obuma_id"])

    # tuple inmutable y hashable para el cache key
    vendor_ids_key = tuple(sorted(selected_vendedor_ids)) if selected_vendedor_ids else tuple()

    st.markdown("---")

    # ── KPIs (cacheado) ───────────────────────────────────────────────
    kpis = _cached_dashboard_kpis(dash_date_from, dash_date_to, vendor_ids_key)
    margen_pct = (kpis["total_margen"] / kpis["total_ventas"] * 100) if kpis["total_ventas"] else 0

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        render_metric("Ventas Totales", format_clp(kpis["total_ventas"]), "💰", ACCENT_GREEN)
    with c2:
        render_metric("Compras Totales", format_clp(kpis["total_compras"]), "🛒", ACCENT_AMBER)
    with c3:
        render_metric("Margen Neto", format_clp(kpis["total_margen"]), "📈", ACCENT_BLUE)
    with c4:
        render_metric("Margen %", f"{margen_pct:.1f}%", "📊", ACCENT_PURPLE)

    st.markdown("")
    c5, c6, c7, c8 = st.columns(4)
    with c5:
        render_metric("Clientes Activos", str(kpis["n_clientes"]), "👥")
    with c6:
        render_metric("Documentos", str(kpis["n_ventas"]), "📄")
    with c7:
        render_metric("Productos", str(kpis["n_productos"]), "📦")
    with c8:
        render_metric("Cobros", str(kpis["total_cobros_count"]), "🧾")

    st.markdown("---")

    # ── Graficos (cacheado) ───────────────────────────────────────────
    charts = _cached_dashboard_charts(dash_date_from, dash_date_to, vendor_ids_key)

    col_chart1, col_chart2 = st.columns(2)

    with col_chart1:
        st.markdown('<p class="section-header">Ventas Mensuales</p>', unsafe_allow_html=True)
        if charts["monthly_sales"]:
            df_ms = pd.DataFrame([{
                "Periodo": f"{anio}-{mes:02d}",
                "Mes": MONTH_LABELS[mes - 1] + f" {anio}",
                "Total": total,
            } for anio, mes, total in charts["monthly_sales"]])
            fig = go.Figure()
            fig.add_trace(go.Bar(x=df_ms["Mes"], y=df_ms["Total"], name="Ventas",
                                 marker_color=ACCENT_BLUE, opacity=0.85))
            if len(df_ms) > 2:
                fig.add_trace(go.Scatter(x=df_ms["Mes"],
                                         y=df_ms["Total"].rolling(3, min_periods=1).mean(),
                                         mode="lines", name="Tendencia",
                                         line=dict(color=ACCENT_AMBER, width=2.5, dash="dot")))
            fig.update_layout(**chart_layout())
            st.plotly_chart(fig)
        else:
            st.info("Sin datos de ventas para el rango seleccionado.")

    with col_chart2:
        st.markdown('<p class="section-header">Top 10 Vendedores</p>', unsafe_allow_html=True)
        if charts["top_vend"]:
            df_tv = pd.DataFrame([{
                "Vendedor": vendedor_names.get(vid, vid or "Desconocido"),
                "Total": total,
            } for vid, total in charts["top_vend"]])
            fig2 = go.Figure(go.Bar(
                x=df_tv["Total"], y=df_tv["Vendedor"], orientation='h',
                marker_color=ACCENT_GREEN,
                text=df_tv["Total"].apply(lambda x: format_clp(x)),
                textposition="auto",
            ))
            fig2.update_layout(**chart_layout(height=380))
            fig2.update_layout(yaxis=dict(autorange="reversed", showgrid=False))
            st.plotly_chart(fig2)
        else:
            st.info("Sin datos de vendedores.")

    st.markdown("")
    col_chart3, col_chart4 = st.columns(2)

    with col_chart3:
        st.markdown('<p class="section-header">Segmentacion ABC de Clientes</p>', unsafe_allow_html=True)
        if charts["client_sales"]:
            grand_total_abc = sum(total for _, total in charts["client_sales"])
            segments = {"A": 0, "B": 0, "C": 0, "D": 0}
            cumulative = 0
            for _, val in charts["client_sales"]:
                cumulative += val
                pct = cumulative / grand_total_abc if grand_total_abc else 0
                if pct <= 0.80:
                    segments["A"] += val
                elif pct <= 0.95:
                    segments["B"] += val
                elif pct <= 0.99:
                    segments["C"] += val
                else:
                    segments["D"] += val

            df_abc = pd.DataFrame([{"Segmento": k, "Total": v} for k, v in segments.items() if v > 0])
            abc_colors = {"A": ACCENT_GREEN, "B": ACCENT_BLUE, "C": ACCENT_AMBER, "D": ACCENT_RED}
            fig3 = go.Figure(go.Pie(
                labels=df_abc["Segmento"], values=df_abc["Total"],
                hole=0.45,
                marker=dict(colors=[abc_colors.get(s, ACCENT_BLUE) for s in df_abc["Segmento"]]),
                textinfo="label+percent", textfont=dict(color="white"),
            ))
            fig3.update_layout(**chart_layout())
            st.plotly_chart(fig3)
        else:
            st.info("Sin datos de clientes para segmentacion.")

    with col_chart4:
        st.markdown('<p class="section-header">Rentabilidad por Vendedor</p>', unsafe_allow_html=True)
        if charts["rent_vend"]:
            df_rv = pd.DataFrame([{
                "Vendedor": vendedor_names.get(vid, vid or "Desconocido"),
                "Margen": margen,
            } for vid, margen in charts["rent_vend"]])
            colors_rv = [ACCENT_GREEN if m >= 0 else ACCENT_RED for m in df_rv["Margen"]]
            fig4 = go.Figure(go.Bar(
                x=df_rv["Vendedor"], y=df_rv["Margen"],
                marker_color=colors_rv,
                text=df_rv["Margen"].apply(lambda x: format_clp(x)),
                textposition="auto",
            ))
            fig4.update_layout(**chart_layout())
            st.plotly_chart(fig4)
        else:
            st.info("Sin datos de rentabilidad.")

    st.markdown("")
    col_chart5, col_chart6 = st.columns(2)

    with col_chart5:
        st.markdown('<p class="section-header">Estado de Cobranza</p>', unsafe_allow_html=True)
        if kpis["total_pagado"] or kpis["total_por_pagar"]:
            fig5 = go.Figure(go.Pie(
                labels=["Pagado", "Por Pagar"],
                values=[kpis["total_pagado"], kpis["total_por_pagar"]],
                hole=0.5,
                marker=dict(colors=[ACCENT_GREEN, ACCENT_RED]),
                textinfo="label+percent+value",
                textfont=dict(color="white"),
                texttemplate="%{label}<br>%{percent}<br>$%{value:,.0f}",
            ))
            fig5.update_layout(**chart_layout())
            st.plotly_chart(fig5)
        else:
            st.info("Sin datos de cobranza.")

    with col_chart6:
        st.markdown('<p class="section-header">Top 15 Productos Vendidos</p>', unsafe_allow_html=True)
        if charts["top_products"]:
            df_tp = pd.DataFrame([{
                "Producto": nombre, "Total": total,
            } for nombre, total in charts["top_products"]])
            fig6 = go.Figure(go.Bar(
                x=df_tp["Total"], y=df_tp["Producto"], orientation='h',
                marker_color=ACCENT_PURPLE,
                text=df_tp["Total"].apply(lambda x: format_clp(x)),
                textposition="auto",
            ))
            fig6.update_layout(**chart_layout(height=420))
            fig6.update_layout(yaxis=dict(autorange="reversed", showgrid=False))
            st.plotly_chart(fig6)
        else:
            st.info("Sin datos de productos vendidos.")

    st.markdown("")
    col_chart7, col_chart8 = st.columns(2)

    with col_chart7:
        st.markdown('<p class="section-header">Evolucion Ventas vs Compras</p>', unsafe_allow_html=True)
        ventas_map, compras_map, all_periods = {}, {}, set()
        for anio, mes, total in charts["monthly_sales"]:
            key = f"{anio}-{mes:02d}"
            all_periods.add(key); ventas_map[key] = total
        for anio, mes, total in charts["monthly_compras"]:
            key = f"{anio}-{mes:02d}"
            all_periods.add(key); compras_map[key] = total

        if all_periods:
            sorted_periods = sorted(all_periods)
            labels = [
                MONTH_LABELS[int(p.split("-")[1]) - 1] + f" {p.split('-')[0]}"
                for p in sorted_periods
            ]
            fig7 = go.Figure()
            fig7.add_trace(go.Scatter(
                x=labels, y=[ventas_map.get(p, 0) for p in sorted_periods],
                mode="lines+markers", name="Ventas",
                line=dict(color=ACCENT_GREEN, width=2.5), marker=dict(size=6),
            ))
            fig7.add_trace(go.Scatter(
                x=labels, y=[compras_map.get(p, 0) for p in sorted_periods],
                mode="lines+markers", name="Compras",
                line=dict(color=ACCENT_RED, width=2.5), marker=dict(size=6),
            ))
            fig7.update_layout(**chart_layout())
            st.plotly_chart(fig7)
        else:
            st.info("Sin datos para comparacion ventas vs compras.")

    with col_chart8:
        st.markdown('<p class="section-header">Distribucion por Tipo Documento</p>', unsafe_allow_html=True)
        if charts["doc_types"]:
            df_dt = pd.DataFrame([{
                "Tipo": tipo, "Total": total, "Cantidad": cant,
            } for tipo, total, cant in charts["doc_types"]])
            fig8 = go.Figure(go.Pie(
                labels=df_dt["Tipo"], values=df_dt["Total"],
                hole=0.4, marker=dict(colors=CHART_COLORS[:len(df_dt)]),
                textinfo="label+percent", textfont=dict(color="white"),
            ))
            fig8.update_layout(**chart_layout())
            st.plotly_chart(fig8)
        else:
            st.info("Sin datos de tipos de documento.")

    st.markdown("---")
    col_tbl1, col_tbl2 = st.columns(2)

    # ── Tablas inferiores (cacheado) ──────────────────────────────────
    rec_top = _cached_dashboard_recent_and_top(dash_date_from, dash_date_to, vendor_ids_key)

    with col_tbl1:
        st.markdown('<p class="section-header">Ultimas Transacciones</p>', unsafe_allow_html=True)
        if rec_top["recent"]:
            data = []
            for v in rec_top["recent"]:
                vend_name = vendedor_names.get(v["vendedor_id"], "-") if v["vendedor_id"] else "-"
                data.append({
                    "Fecha": v["fecha"],
                    "Tipo": v["tipo"],
                    "Folio": v["folio"],
                    "Total": format_clp(v["total"]),
                    "Vendedor": vend_name,
                    "Estado": "Anulada" if v["anulada"] else "Vigente",
                })
            st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
        else:
            st.info("Sin transacciones recientes.")

    with col_tbl2:
        st.markdown('<p class="section-header">Top Clientes por Ingresos</p>', unsafe_allow_html=True)
        if rec_top["top_clientes"]:
            df_top = pd.DataFrame([{
                "Cliente": c["nombre"],
                "RUT": c["rut"],
                "Total": format_clp(c["total"]),
                "Margen": format_clp(c["margen"]),
                "Docs": c["transacciones"],
            } for c in rec_top["top_clientes"]])
            st.dataframe(df_top, hide_index=True, height=400)
        else:
            st.info("Sin datos de clientes.")


# ============================================================
# VENDEDORES
# ============================================================
elif page == "Vendedores":
    st.markdown('<p class="page-title">Gestion de Vendedores</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Rendimiento, metas y cartera de clientes por vendedor</p>', unsafe_allow_html=True)

    TRACKED_VENDEDORES = TRACKED_VENDEDOR_IDS

    db = get_db()
    try:
        tab_rend, tab_cartera, tab_cruce, tab_metas = st.tabs([
            "📊 Rendimiento vs Metas", "👥 Cartera de Clientes", "🔍 Cruce Cartera vs Ventas", "🎯 Configurar Metas"
        ])

        current_year = date.today().year

        vendedores_map = {}
        for vid in TRACKED_VENDEDORES:
            emp = db.query(Empleado).filter(Empleado.obuma_id == vid).first()
            if emp:
                vendedores_map[vid] = emp

        # ── TAB 1: Rendimiento vs Metas ──
        with tab_rend:
            st.markdown('<p class="section-header">Rendimiento Mensual vs Metas</p>', unsafe_allow_html=True)

            rc1, rc2 = st.columns(2)
            with rc1:
                rend_mes = st.selectbox("Mes", list(range(1, 13)),
                                        index=date.today().month - 1,
                                        format_func=lambda m: MONTH_LABELS[m - 1],
                                        key="vend_rend_mes")
            with rc2:
                rend_anio = st.selectbox("Año", list(range(current_year - 2, current_year + 2)),
                                         index=2, key="vend_rend_anio")

            st.markdown("---")

            summary_meta_rep = 0
            summary_meta_maq = 0
            summary_actual_rep = 0
            summary_actual_maq = 0
            chart_names = []
            chart_metas = []
            chart_actuals = []

            # ── Pre-fetch GROUP BY (Fase 2: antes eran 5 queries x N vendedores) ──
            # 1) Metas del periodo, indexadas por vendedor.
            metas_rows = db.query(VendedorMeta).filter(
                VendedorMeta.empleado_obuma_id.in_(TRACKED_VENDEDORES),
                VendedorMeta.anio == rend_anio,
                VendedorMeta.mes == rend_mes,
            ).all()
            metas_map = {m.empleado_obuma_id: m for m in metas_rows}

            # Rango de fecha equivalente a (year=rend_anio, month=rend_mes). Fase 3:
            # se usa col >= start AND col < end_excl en vez de extract('year/month',
            # col) == ... para que el planner aproveche ix_ventas_vendedor_fecha.
            _period_start, _period_end = year_month_range(rend_anio, rend_mes)

            # 2) Neto total por vendedor (NC resta).
            actual_neto_rows = db.query(
                VentaHistorico.vendedor_id,
                func.sum(
                    sql_case(
                        (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -VentaHistorico.subtotal),
                        else_=VentaHistorico.subtotal,
                    )
                ).label("neto"),
            ).filter(
                VentaHistorico.vendedor_id.in_(TRACKED_VENDEDORES),
                VentaHistorico.fecha >= _period_start,
                VentaHistorico.fecha < _period_end,
                VentaHistorico.anulada == False,
                VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
            ).group_by(VentaHistorico.vendedor_id).all()
            actual_neto_map = {r.vendedor_id: float(r.neto or 0) for r in actual_neto_rows}

            # 3) Maquinaria por vendedor (sku LIKE 'mq-%', NC resta).
            _sign_mq = sql_case(
                (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -1),
                else_=1,
            )
            actual_maq_rows = db.query(
                VentaHistorico.vendedor_id,
                func.coalesce(func.sum(_sign_mq * VentaItem.total), 0).label("maq"),
            ).select_from(VentaItem).join(
                VentaHistorico, VentaHistorico.obuma_id == VentaItem.venta_id_obuma,
            ).filter(
                VentaHistorico.vendedor_id.in_(TRACKED_VENDEDORES),
                VentaHistorico.fecha >= _period_start,
                VentaHistorico.fecha < _period_end,
                VentaHistorico.anulada == False,
                VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
                func.lower(VentaItem.producto_sku).like('mq-%'),
            ).group_by(VentaHistorico.vendedor_id).all()
            actual_maq_map = {r.vendedor_id: float(r.maq or 0) for r in actual_maq_rows}

            # 4) Cartera total por vendedor (excluye dummies OBU-* sin nombre).
            cartera_rows = db.query(
                VendedorCartera.empleado_obuma_id,
                func.count(VendedorCartera.id).label("total"),
            ).join(
                ClienteFinal, VendedorCartera.cliente_id == ClienteFinal.id,
            ).filter(
                VendedorCartera.empleado_obuma_id.in_(TRACKED_VENDEDORES),
                VendedorCartera.activo == True,
                ~(
                    ClienteFinal.rut.like('OBU-%') &
                    (func.coalesce(func.trim(ClienteFinal.nombre), '') == '')
                ),
            ).group_by(VendedorCartera.empleado_obuma_id).all()
            cartera_count_map = {r.empleado_obuma_id: int(r.total or 0) for r in cartera_rows}

            # 5) Clientes atendidos por vendedor: solo cuentan los que tienen
            #    neto > 0 en el periodo (regla original — NCs no cuentan como atencion).
            #    Subquery: neto por (vendedor_id, cliente_id) HAVING > 0.
            #    Outer: COUNT por vendedor_id.
            _net_per_vc_subq = (
                db.query(
                    VentaHistorico.vendedor_id.label('vid'),
                    VentaHistorico.cliente_id.label('cli_id'),
                    func.sum(
                        sql_case(
                            (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -VentaHistorico.subtotal),
                            else_=VentaHistorico.subtotal,
                        )
                    ).label('neto'),
                )
                .join(ClienteFinal, VentaHistorico.cliente_id == ClienteFinal.id)
                .filter(
                    VentaHistorico.vendedor_id.in_(TRACKED_VENDEDORES),
                    VentaHistorico.fecha >= _period_start,
                    VentaHistorico.fecha < _period_end,
                    VentaHistorico.anulada == False,
                    VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
                    ~(
                        ClienteFinal.rut.like('OBU-%') &
                        (func.coalesce(func.trim(ClienteFinal.nombre), '') == '')
                    ),
                )
                .group_by(VentaHistorico.vendedor_id, VentaHistorico.cliente_id)
                .having(func.sum(
                    sql_case(
                        (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -VentaHistorico.subtotal),
                        else_=VentaHistorico.subtotal,
                    )
                ) > 0)
                .subquery()
            )
            atendidos_rows = db.query(
                _net_per_vc_subq.c.vid,
                func.count(_net_per_vc_subq.c.cli_id).label("n"),
            ).group_by(_net_per_vc_subq.c.vid).all()
            atendidos_map = {r.vid: int(r.n or 0) for r in atendidos_rows}

            for vid in TRACKED_VENDEDORES:
                emp = vendedores_map.get(vid)
                if not emp:
                    continue

                meta = metas_map.get(vid)
                meta_rep = meta.meta_repuestos if meta else 0
                meta_maq = meta.meta_maquinaria if meta else 0
                meta_total = meta_rep + meta_maq

                actual_total_neto = actual_neto_map.get(vid, 0)
                actual_maq = actual_maq_map.get(vid, 0)
                actual_rep_result = actual_total_neto - actual_maq
                actual_total = actual_total_neto

                rep_negativo = actual_rep_result < 0
                maq_negativo = actual_maq < 0
                actual_rep_for_pct = max(actual_rep_result, 0)
                actual_maq_for_pct = max(actual_maq, 0)
                actual_total_for_pct = max(actual_total, 0)

                total_cartera_count = cartera_count_map.get(vid, 0)
                clientes_atendidos = atendidos_map.get(vid, 0)
                cobertura_pct = (clientes_atendidos / total_cartera_count * 100) if total_cartera_count > 0 else 0

                summary_meta_rep += meta_rep
                summary_meta_maq += meta_maq
                summary_actual_rep += actual_rep_for_pct
                summary_actual_maq += actual_maq_for_pct

                chart_names.append(emp.nombre.split(" ")[0] if emp.nombre else vid)
                chart_metas.append(meta_total)
                chart_actuals.append(actual_total_for_pct)

                pct_rep = (actual_rep_for_pct / meta_rep * 100) if meta_rep > 0 else 0
                pct_maq = (actual_maq_for_pct / meta_maq * 100) if meta_maq > 0 else 0
                pct_total = (actual_total_for_pct / meta_total * 100) if meta_total > 0 else 0

                if pct_total >= 100:
                    border_color = ACCENT_GREEN
                elif pct_total >= 70:
                    border_color = ACCENT_AMBER
                else:
                    border_color = ACCENT_RED

                rep_display = f'<span style="color:#F59E0B;font-weight:600;">{format_clp(actual_rep_result)} ⚠</span>' if rep_negativo else format_clp(actual_rep_result)
                maq_display = f'<span style="color:#F59E0B;font-weight:600;">{format_clp(actual_maq)} ⚠</span>' if maq_negativo else format_clp(actual_maq)

                pct_rep_color = ACCENT_GREEN if pct_rep >= 100 else (ACCENT_AMBER if pct_rep >= 70 else ACCENT_RED)
                pct_maq_color = ACCENT_GREEN if pct_maq >= 100 else (ACCENT_AMBER if pct_maq >= 70 else ACCENT_RED)

                st.markdown(f"""
                <div style="background:{CARD_BG}; border:1px solid {border_color}; border-radius:12px;
                            padding:1rem 1.5rem; margin-bottom:0.4rem; border-left:4px solid {border_color};">
                    <div style="display:flex; justify-content:space-between; align-items:center; flex-wrap:wrap; gap:0.5rem;">
                        <div>
                            <span style="font-size:1.1rem; font-weight:600; color:{TEXT_PRIMARY};">{emp.nombre}</span>
                            <span style="font-size:0.8rem; color:{TEXT_SECONDARY}; margin-left:0.5rem;">{emp.cargo or ''}</span>
                        </div>
                        <div style="display:flex; gap:1.2rem; align-items:center;">
                            <div style="text-align:center;">
                                <div style="font-size:0.72rem; color:{TEXT_SECONDARY}; margin-bottom:2px;">🔧 Repuestos</div>
                                <div style="font-size:1.25rem; font-weight:700; color:{pct_rep_color};">{pct_rep:.0f}%</div>
                                <div style="font-size:0.7rem; color:{TEXT_SECONDARY};">{rep_display} / {format_clp(meta_rep)}</div>
                            </div>
                            <div style="width:1px; height:40px; background:{CARD_BORDER};"></div>
                            <div style="text-align:center;">
                                <div style="font-size:0.72rem; color:{TEXT_SECONDARY}; margin-bottom:2px;">🏗️ Maquinaria</div>
                                <div style="font-size:1.25rem; font-weight:700; color:{pct_maq_color};">{pct_maq:.0f}%</div>
                                <div style="font-size:0.7rem; color:{TEXT_SECONDARY};">{maq_display} / {format_clp(meta_maq)}</div>
                            </div>
                            <div style="width:1px; height:40px; background:{CARD_BORDER};"></div>
                            <div style="text-align:center;">
                                <div style="font-size:0.72rem; color:{TEXT_SECONDARY}; margin-bottom:2px;">📊 Total</div>
                                <div style="font-size:1.25rem; font-weight:700; color:{border_color};">{pct_total:.0f}%</div>
                                <div style="font-size:0.7rem; color:{TEXT_SECONDARY};">{format_clp(actual_total_for_pct)} / {format_clp(meta_total)}</div>
                            </div>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

                vc1, vc2, vc3, vc4 = st.columns(4)
                with vc1:
                    st.caption("Avance Repuestos")
                    prog_rep = min(pct_rep / 100, 1.0)
                    st.progress(prog_rep if prog_rep >= 0 else 0)
                with vc2:
                    st.caption("Avance Maquinaria")
                    prog_maq = min(pct_maq / 100, 1.0)
                    st.progress(prog_maq if prog_maq >= 0 else 0)
                with vc3:
                    st.caption("Avance Total")
                    prog_total = min(pct_total / 100, 1.0)
                    st.progress(prog_total if prog_total >= 0 else 0)
                with vc4:
                    st.caption(f"Clientes: {clientes_atendidos} / {total_cartera_count} ({cobertura_pct:.0f}%)")
                    prog_cob = min(cobertura_pct / 100, 1.0)
                    st.progress(prog_cob if prog_cob >= 0 else 0)

                st.markdown("")

            st.markdown("---")
            st.markdown('<p class="section-header">Resumen General</p>', unsafe_allow_html=True)

            sc1, sc2, sc3, sc4 = st.columns(4)
            with sc1:
                render_metric("Meta Repuestos", format_clp(summary_meta_rep), "🔧", ACCENT_BLUE)
            with sc2:
                render_metric("Actual Repuestos", format_clp(summary_actual_rep), "✅", ACCENT_GREEN)
            with sc3:
                render_metric("Meta Maquinaria", format_clp(summary_meta_maq), "🏗️", ACCENT_AMBER)
            with sc4:
                render_metric("Actual Maquinaria", format_clp(summary_actual_maq), "📦", ACCENT_PURPLE)

            st.markdown("")
            st.markdown('<p class="section-header">Meta vs Real por Vendedor</p>', unsafe_allow_html=True)

            if chart_names:
                fig_rend = go.Figure()
                fig_rend.add_trace(go.Bar(
                    x=chart_names, y=chart_metas, name="Meta",
                    marker_color=ACCENT_BLUE, opacity=0.7,
                    text=[format_clp(v) for v in chart_metas], textposition="auto"
                ))
                fig_rend.add_trace(go.Bar(
                    x=chart_names, y=chart_actuals, name="Real",
                    marker_color=ACCENT_GREEN, opacity=0.85,
                    text=[format_clp(v) for v in chart_actuals], textposition="auto"
                ))
                fig_rend.update_layout(**chart_layout(height=400))
                fig_rend.update_layout(barmode="group")
                st.plotly_chart(fig_rend)

        # ── TAB 2: Cartera de Clientes ──
        with tab_cartera:
            st.markdown('<p class="section-header">Cartera de Clientes por Vendedor</p>', unsafe_allow_html=True)

            vendedor_options_cart = [(vid, vendedores_map[vid].nombre) for vid in TRACKED_VENDEDORES if vid in vendedores_map]
            if vendedor_options_cart:
                sel_cart_label = st.selectbox(
                    "Seleccionar Vendedor",
                    [f"{name} ({vid})" for vid, name in vendedor_options_cart],
                    key="vend_cart_sel"
                )
                sel_cart_vid = sel_cart_label.split("(")[-1].replace(")", "").strip()

                st.markdown("---")

                cartera_items = db.query(VendedorCartera, ClienteFinal).join(
                    ClienteFinal, VendedorCartera.cliente_id == ClienteFinal.id
                ).filter(
                    VendedorCartera.empleado_obuma_id == sel_cart_vid,
                    VendedorCartera.activo == True
                ).all()

                twelve_months_ago = date.today() - timedelta(days=365)

                if cartera_items:
                    st.markdown(f"**{len(cartera_items)} clientes activos asignados**")

                    # Fase 2: una sola query GROUP BY para los 12M de TODOS los
                    # clientes de la cartera (antes era 1 query por cliente,
                    # ~100+ roundtrips).
                    cliente_ids = [cli.id for _, cli in cartera_items]
                    ventas_12m_rows = db.query(
                        VentaHistorico.cliente_id,
                        func.sum(
                            sql_case(
                                (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -VentaHistorico.subtotal),
                                else_=VentaHistorico.subtotal,
                            )
                        ).label("neto"),
                    ).filter(
                        VentaHistorico.cliente_id.in_(cliente_ids),
                        VentaHistorico.vendedor_id == sel_cart_vid,
                        VentaHistorico.anulada == False,
                        *date_range_filters(VentaHistorico.fecha, twelve_months_ago, None),
                        VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
                    ).group_by(VentaHistorico.cliente_id).all() if cliente_ids else []
                    ventas_12m_map = {r.cliente_id: float(r.neto or 0) for r in ventas_12m_rows}

                    rows = []
                    for vc, cli in cartera_items:
                        ventas_12m = ventas_12m_map.get(cli.id, 0)
                        rows.append({
                            "Cliente": cli.nombre,
                            "RUT": cli.rut or "—",
                            "Asignado": str(vc.fecha_asignacion) if vc.fecha_asignacion else "—",
                            "Ventas 12M": format_clp(ventas_12m),
                            "ID": cli.id
                        })

                    df_cart = pd.DataFrame(rows)
                    st.dataframe(df_cart[["Cliente", "RUT", "Asignado", "Ventas 12M"]],
                                 hide_index=True, height=400)
                else:
                    st.info("Este vendedor no tiene clientes asignados.")

                st.markdown("---")
                st.markdown('<p class="section-header">Gestionar Cartera</p>', unsafe_allow_html=True)

                gc1, gc2 = st.columns(2)

                with gc1:
                    st.markdown("**Agregar Cliente**")
                    assigned_ids = [vc.cliente_id for vc, _ in cartera_items] if cartera_items else []
                    unassigned = db.query(ClienteFinal).filter(
                        ClienteFinal.activo == True,
                        ~ClienteFinal.id.in_(assigned_ids) if assigned_ids else True
                    ).order_by(ClienteFinal.nombre).limit(200).all()

                    if unassigned:
                        new_client_label = st.selectbox(
                            "Cliente a agregar",
                            [f"{c.nombre} ({c.rut or 'Sin RUT'})" for c in unassigned],
                            key="vend_add_client"
                        )
                        if st.button("➕ Agregar a Cartera", key="btn_add_client"):
                            idx = [f"{c.nombre} ({c.rut or 'Sin RUT'})" for c in unassigned].index(new_client_label)
                            new_vc = VendedorCartera(
                                empleado_obuma_id=sel_cart_vid,
                                cliente_id=unassigned[idx].id,
                                fecha_asignacion=date.today(),
                                activo=True
                            )
                            db.add(new_vc)
                            db.commit()
                            st.success(f"Cliente asignado exitosamente.")
                            st.rerun()
                    else:
                        st.caption("No hay clientes disponibles para asignar.")

                with gc2:
                    st.markdown("**Remover Cliente**")
                    if cartera_items:
                        remove_label = st.selectbox(
                            "Cliente a remover",
                            [f"{cli.nombre} ({cli.rut or 'Sin RUT'})" for _, cli in cartera_items],
                            key="vend_remove_client"
                        )
                        if st.button("❌ Remover de Cartera", key="btn_remove_client"):
                            idx = [f"{cli.nombre} ({cli.rut or 'Sin RUT'})" for _, cli in cartera_items].index(remove_label)
                            vc_to_remove = cartera_items[idx][0]
                            vc_to_remove.activo = False
                            vc_to_remove.fecha_baja = date.today()
                            db.commit()
                            st.success("Cliente removido de la cartera.")
                            st.rerun()
                    else:
                        st.caption("No hay clientes para remover.")

                st.markdown("---")
                st.markdown('<p class="section-header">Clientes por Vendedor</p>', unsafe_allow_html=True)

                # Fase 2: una sola GROUP BY en vez de 1 query por vendedor.
                cart_count_rows = db.query(
                    VendedorCartera.empleado_obuma_id,
                    func.count(VendedorCartera.id).label("n"),
                ).filter(
                    VendedorCartera.empleado_obuma_id.in_(TRACKED_VENDEDORES),
                    VendedorCartera.activo == True,
                ).group_by(VendedorCartera.empleado_obuma_id).all()
                cart_count_map = {r.empleado_obuma_id: int(r.n or 0) for r in cart_count_rows}

                chart_vend_names = []
                chart_vend_counts = []
                for vid in TRACKED_VENDEDORES:
                    emp = vendedores_map.get(vid)
                    if not emp:
                        continue
                    chart_vend_names.append(emp.nombre.split(" ")[0] if emp.nombre else vid)
                    chart_vend_counts.append(cart_count_map.get(vid, 0))

                if chart_vend_names:
                    fig_cart = go.Figure(go.Bar(
                        x=chart_vend_names, y=chart_vend_counts,
                        marker_color=CHART_COLORS[:len(chart_vend_names)],
                        text=chart_vend_counts, textposition="auto"
                    ))
                    fig_cart.update_layout(**chart_layout())
                    st.plotly_chart(fig_cart)
            else:
                st.warning("No se encontraron vendedores registrados.")

        # ── TAB 3: Cruce Cartera vs Ventas ──
        with tab_cruce:
            st.markdown('<p class="section-header">Cruce Cartera vs Ventas</p>', unsafe_allow_html=True)

            cruce_vendedor_opts = [(vid, vendedores_map[vid].nombre) for vid in TRACKED_VENDEDORES if vid in vendedores_map]
            if cruce_vendedor_opts:
                cc1, cc2, cc3 = st.columns(3)
                with cc1:
                    cruce_sel_label = st.selectbox(
                        "Vendedor",
                        [f"{name} ({vid})" for vid, name in cruce_vendedor_opts],
                        key="vend_cruce_sel"
                    )
                    cruce_sel_vid = cruce_sel_label.split("(")[-1].replace(")", "").strip()
                with cc2:
                    cruce_anio = st.selectbox("Año", list(range(current_year - 2, current_year + 2)),
                                              index=2, key="vend_cruce_anio")
                with cc3:
                    cruce_mes_options = ["Todo el año"] + MONTH_LABELS
                    cruce_mes_sel = st.selectbox("Mes", cruce_mes_options, index=0, key="vend_cruce_mes")

                st.markdown("---")

                cartera_cruce = db.query(VendedorCartera, ClienteFinal).join(
                    ClienteFinal, VendedorCartera.cliente_id == ClienteFinal.id
                ).filter(
                    VendedorCartera.empleado_obuma_id == cruce_sel_vid,
                    VendedorCartera.activo == True
                ).all()

                if cartera_cruce:
                    compraron_rows = []
                    no_compraron_rows = []

                    # Fase 2: 2 queries GROUP BY en vez de 2 queries por cliente
                    # (~200+ roundtrips antes en una cartera de 100 clientes).
                    cliente_ids_cruce = [cli.id for _, cli in cartera_cruce]

                    # Query A: ventas del periodo seleccionado, por cliente.
                    # Fase 3: rango de fecha (col >= start AND col < end_excl) en
                    # vez de extract('year/month', col) para usar ix_ventas_*_fecha.
                    _cruce_mes = (MONTH_LABELS.index(cruce_mes_sel) + 1) if cruce_mes_sel != "Todo el año" else None
                    _cruce_start, _cruce_end = year_month_range(cruce_anio, _cruce_mes)
                    ventas_q_all = db.query(
                        VentaHistorico.cliente_id,
                        func.sum(
                            sql_case(
                                (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -VentaHistorico.subtotal),
                                else_=VentaHistorico.subtotal,
                            )
                        ).label("total_ventas"),
                        func.count(VentaHistorico.id).label("num_docs"),
                    ).filter(
                        VentaHistorico.cliente_id.in_(cliente_ids_cruce),
                        VentaHistorico.vendedor_id == cruce_sel_vid,
                        VentaHistorico.anulada == False,
                        VentaHistorico.fecha >= _cruce_start,
                        VentaHistorico.fecha < _cruce_end,
                        VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G),
                    )
                    ventas_periodo_rows = ventas_q_all.group_by(VentaHistorico.cliente_id).all() if cliente_ids_cruce else []
                    ventas_periodo_map = {
                        r.cliente_id: (float(r.total_ventas or 0), int(r.num_docs or 0))
                        for r in ventas_periodo_rows
                    }

                    # Query B: ultima compra historica (max fecha) por cliente,
                    # solo para los que NO compraron en el periodo. Una sola query.
                    no_compraron_ids = [
                        cli.id for _, cli in cartera_cruce
                        if ventas_periodo_map.get(cli.id, (0, 0))[1] == 0
                    ]
                    ultima_rows = db.query(
                        VentaHistorico.cliente_id,
                        func.max(VentaHistorico.fecha).label("ult"),
                    ).filter(
                        VentaHistorico.cliente_id.in_(no_compraron_ids),
                        VentaHistorico.vendedor_id == cruce_sel_vid,
                        VentaHistorico.anulada == False,
                    ).group_by(VentaHistorico.cliente_id).all() if no_compraron_ids else []
                    ultima_compra_map = {r.cliente_id: r.ult for r in ultima_rows}

                    for vc, cli in cartera_cruce:
                        total_ventas, num_docs = ventas_periodo_map.get(cli.id, (0, 0))

                        if num_docs > 0:
                            compraron_rows.append({
                                "Cliente": cli.nombre,
                                "RUT": cli.rut or "—",
                                "Total Ventas": total_ventas,
                                "Total Ventas Fmt": format_clp(total_ventas),
                                "Documentos": num_docs
                            })
                        else:
                            ultima_compra = ultima_compra_map.get(cli.id)
                            dias_sin = (date.today() - ultima_compra.date()).days if ultima_compra else None
                            no_compraron_rows.append({
                                "Cliente": cli.nombre,
                                "RUT": cli.rut or "—",
                                "Última Compra": str(ultima_compra.date()) if ultima_compra else "Sin registro",
                                "Días sin Comprar": str(dias_sin) if dias_sin is not None else "—"
                            })

                    total_cartera = len(cartera_cruce)
                    total_compraron = len(compraron_rows)
                    total_no_compraron = len(no_compraron_rows)
                    cobertura = (total_compraron / total_cartera * 100) if total_cartera > 0 else 0

                    mc1, mc2, mc3, mc4 = st.columns(4)
                    with mc1:
                        render_metric("Total Cartera", str(total_cartera), "👥", ACCENT_BLUE)
                    with mc2:
                        render_metric("Compraron", str(total_compraron), "✅", ACCENT_GREEN)
                    with mc3:
                        render_metric("No Compraron", str(total_no_compraron), "❌", ACCENT_RED)
                    with mc4:
                        render_metric("Cobertura", f"{cobertura:.1f}%", "📊", ACCENT_AMBER)

                    st.markdown("")

                    col_si, col_no = st.columns(2)
                    with col_si:
                        st.markdown(f'<p class="section-header" style="color:{ACCENT_GREEN};">✅ Clientes que Compraron ({total_compraron})</p>', unsafe_allow_html=True)
                        if compraron_rows:
                            df_si = pd.DataFrame(compraron_rows)
                            df_si_sorted = df_si.sort_values("Total Ventas", ascending=False)
                            st.dataframe(
                                df_si_sorted[["Cliente", "RUT", "Total Ventas Fmt", "Documentos"]].rename(
                                    columns={"Total Ventas Fmt": "Total Ventas"}
                                ),
                                hide_index=True, height=400
                            )
                        else:
                            st.info("Ningún cliente de la cartera compró en este periodo.")

                    with col_no:
                        st.markdown(f'<p class="section-header" style="color:{ACCENT_RED};">❌ Clientes que NO Compraron ({total_no_compraron})</p>', unsafe_allow_html=True)
                        if no_compraron_rows:
                            df_no = pd.DataFrame(no_compraron_rows)
                            st.dataframe(
                                df_no[["Cliente", "RUT", "Última Compra", "Días sin Comprar"]],
                                hide_index=True, height=400
                            )
                        else:
                            st.info("Todos los clientes de la cartera compraron en este periodo.")

                    st.markdown("")
                    st.markdown('<p class="section-header">Comparación Visual</p>', unsafe_allow_html=True)

                    fig_cruce = go.Figure(go.Bar(
                        x=["Compraron", "No Compraron"],
                        y=[total_compraron, total_no_compraron],
                        marker_color=[ACCENT_GREEN, ACCENT_RED],
                        text=[total_compraron, total_no_compraron],
                        textposition="auto",
                        textfont=dict(size=16, color="white")
                    ))
                    fig_cruce.update_layout(**chart_layout(height=350))
                    fig_cruce.update_layout(
                        xaxis_title="",
                        yaxis_title="Cantidad de Clientes"
                    )
                    st.plotly_chart(fig_cruce)

                    st.markdown("")
                    periodo_label = f"{cruce_anio}" if cruce_mes_sel == "Todo el año" else f"{cruce_mes_sel} {cruce_anio}"
                    emp_cruce = vendedores_map.get(cruce_sel_vid)
                    nombre_vendedor = emp_cruce.nombre if emp_cruce else cruce_sel_vid

                    if st.button("📥 Exportar a Excel", key="btn_cruce_export"):
                        from openpyxl import Workbook
                        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

                        wb = Workbook()

                        ws_si = wb.active
                        ws_si.title = "Compraron"
                        header_font = Font(bold=True, color="FFFFFF", size=11)
                        green_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                        red_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
                        thin_border = Border(
                            left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin')
                        )

                        ws_si.append([f"Cruce Cartera vs Ventas - {nombre_vendedor} - {periodo_label}"])
                        ws_si.merge_cells('A1:D1')
                        ws_si['A1'].font = Font(bold=True, size=14)
                        ws_si.append([])
                        ws_si.append(["Cliente", "RUT", "Total Ventas", "Documentos"])
                        for cell in ws_si[3]:
                            cell.font = header_font
                            cell.fill = green_fill
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border
                        for row in compraron_rows:
                            ws_si.append([row["Cliente"], row["RUT"], row["Total Ventas"], row["Documentos"]])
                        ws_si.column_dimensions['A'].width = 40
                        ws_si.column_dimensions['B'].width = 15
                        ws_si.column_dimensions['C'].width = 18
                        ws_si.column_dimensions['D'].width = 14

                        ws_no = wb.create_sheet("No Compraron")
                        ws_no.append([f"Clientes que NO Compraron - {nombre_vendedor} - {periodo_label}"])
                        ws_no.merge_cells('A1:D1')
                        ws_no['A1'].font = Font(bold=True, size=14)
                        ws_no.append([])
                        ws_no.append(["Cliente", "RUT", "Última Compra", "Días sin Comprar"])
                        for cell in ws_no[3]:
                            cell.font = header_font
                            cell.fill = red_fill
                            cell.alignment = Alignment(horizontal='center')
                            cell.border = thin_border
                        for row in no_compraron_rows:
                            ws_no.append([row["Cliente"], row["RUT"], row["Última Compra"], row["Días sin Comprar"]])
                        ws_no.column_dimensions['A'].width = 40
                        ws_no.column_dimensions['B'].width = 15
                        ws_no.column_dimensions['C'].width = 18
                        ws_no.column_dimensions['D'].width = 18

                        ws_resumen = wb.create_sheet("Resumen", 0)
                        ws_resumen.append(["Resumen Cruce Cartera vs Ventas"])
                        ws_resumen.merge_cells('A1:B1')
                        ws_resumen['A1'].font = Font(bold=True, size=14)
                        ws_resumen.append([])
                        ws_resumen.append(["Vendedor", nombre_vendedor])
                        ws_resumen.append(["Periodo", periodo_label])
                        ws_resumen.append(["Total Cartera", total_cartera])
                        ws_resumen.append(["Compraron", total_compraron])
                        ws_resumen.append(["No Compraron", total_no_compraron])
                        ws_resumen.append(["Cobertura", f"{cobertura:.1f}%"])
                        ws_resumen.column_dimensions['A'].width = 20
                        ws_resumen.column_dimensions['B'].width = 30

                        buffer = BytesIO()
                        wb.save(buffer)
                        buffer.seek(0)

                        st.download_button(
                            label="⬇️ Descargar Excel",
                            data=buffer.getvalue(),
                            file_name=f"cruce_cartera_{cruce_sel_vid}_{periodo_label.replace(' ', '_')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key="download_cruce_excel"
                        )
                else:
                    st.info("Este vendedor no tiene clientes asignados en su cartera.")
            else:
                st.warning("No se encontraron vendedores registrados.")

        # ── TAB 4: Configurar Metas ──
        with tab_metas:
            st.markdown('<p class="section-header">Configuracion de Metas Mensuales</p>', unsafe_allow_html=True)

            mc1, mc2 = st.columns(2)
            with mc1:
                meta_vendedor_opts = [(vid, vendedores_map[vid].nombre) for vid in TRACKED_VENDEDORES if vid in vendedores_map]
                if meta_vendedor_opts:
                    meta_sel_label = st.selectbox(
                        "Vendedor",
                        [f"{name} ({vid})" for vid, name in meta_vendedor_opts],
                        key="vend_meta_sel"
                    )
                    meta_sel_vid = meta_sel_label.split("(")[-1].replace(")", "").strip()
                else:
                    meta_sel_vid = None
                    st.warning("No hay vendedores registrados.")
            with mc2:
                meta_anio = st.selectbox("Año", list(range(current_year - 1, current_year + 3)),
                                          index=1, key="vend_meta_anio")

            if meta_sel_vid:
                st.markdown("---")
                st.info("💡 Haz clic en cualquier celda de la tabla para editar el valor. Luego presiona **💾 Guardar Metas**.")

                existing_metas = db.query(VendedorMeta).filter(
                    VendedorMeta.empleado_obuma_id == meta_sel_vid,
                    VendedorMeta.anio == meta_anio
                ).all()

                meta_dict = {m.mes: m for m in existing_metas}

                editor_data = []
                for m in range(1, 13):
                    meta_obj = meta_dict.get(m)
                    editor_data.append({
                        "Mes": MONTH_LABELS[m - 1],
                        "Meta Repuestos": int(meta_obj.meta_repuestos) if meta_obj else 0,
                        "Meta Maquinaria": int(meta_obj.meta_maquinaria) if meta_obj else 0,
                    })

                df_metas = pd.DataFrame(editor_data)

                edited_df = st.data_editor(
                    df_metas,
                    hide_index=True,
                    disabled=["Mes"],
                    num_rows="fixed",
                    use_container_width=True,
                    column_config={
                        "Mes": st.column_config.TextColumn("Mes", disabled=True),
                        "Meta Repuestos": st.column_config.NumberColumn("Meta Repuestos ($)", min_value=0, step=100000, format="$ %d"),
                        "Meta Maquinaria": st.column_config.NumberColumn("Meta Maquinaria ($)", min_value=0, step=100000, format="$ %d"),
                    },
                    key=f"meta_editor_{meta_sel_vid}_{meta_anio}"
                )

                btn_c1, btn_c2 = st.columns(2)

                with btn_c1:
                    if st.button("💾 Guardar Metas", type="primary", key="btn_save_metas"):
                        for i, row in edited_df.iterrows():
                            mes_num = i + 1
                            existing = meta_dict.get(mes_num)
                            if existing:
                                existing.meta_repuestos = float(row["Meta Repuestos"])
                                existing.meta_maquinaria = float(row["Meta Maquinaria"])
                            else:
                                new_meta = VendedorMeta(
                                    empleado_obuma_id=meta_sel_vid,
                                    anio=meta_anio,
                                    mes=mes_num,
                                    meta_repuestos=float(row["Meta Repuestos"]),
                                    meta_maquinaria=float(row["Meta Maquinaria"])
                                )
                                db.add(new_meta)
                        db.commit()
                        st.success("Metas guardadas exitosamente.")
                        st.rerun()

                with btn_c2:
                    copy_from = st.selectbox("Copiar desde mes", list(range(1, 13)),
                                              format_func=lambda m: MONTH_LABELS[m - 1],
                                              key="vend_copy_from")
                    if st.button("📋 Copiar a meses restantes", key="btn_copy_metas"):
                        source_row = edited_df.iloc[copy_from - 1]
                        rep_val = float(source_row["Meta Repuestos"])
                        maq_val = float(source_row["Meta Maquinaria"])
                        for mes_num in range(copy_from + 1, 13):
                            existing = meta_dict.get(mes_num)
                            if existing:
                                existing.meta_repuestos = rep_val
                                existing.meta_maquinaria = maq_val
                            else:
                                new_meta = VendedorMeta(
                                    empleado_obuma_id=meta_sel_vid,
                                    anio=meta_anio,
                                    mes=mes_num,
                                    meta_repuestos=rep_val,
                                    meta_maquinaria=maq_val
                                )
                                db.add(new_meta)
                        db.commit()
                        st.success(f"Metas copiadas desde {MONTH_LABELS[copy_from - 1]} a los meses siguientes.")
                        st.rerun()

    finally:
        db.close()


# ============================================================
# VENTAS
# ============================================================
elif page == "Ventas":
    st.markdown('<p class="page-title">Gestion de Ventas</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Documentos de venta, items, cotizaciones, cobros y DTE</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        tab_docs, tab_items, tab_cotiz, tab_cobros, tab_dte = st.tabs([
            "Documentos", "Items de Venta", "Cotizaciones", "Cobros", "DTE"
        ])

        with tab_docs:
            empleados_v = db.query(Empleado).filter(Empleado.activo == True).order_by(Empleado.nombre).all()
            vend_map_v = {e.obuma_id: e.nombre for e in empleados_v}

            col1, col2, col3, col4 = st.columns([2, 2, 2, 2])
            with col1:
                fecha_desde = st.date_input("Desde", value=date.today() - timedelta(days=365), key="vd1")
            with col2:
                fecha_hasta = st.date_input("Hasta", value=date.today(), key="vd2")
            with col3:
                filtro_estado = st.selectbox("Estado", ["Todos", "Vigentes", "Anuladas"])
            with col4:
                vend_options_v = ["Todos"] + [e.nombre for e in empleados_v]
                filtro_vendedor = st.selectbox("Vendedor", vend_options_v, key="vd_vend")

            query = db.query(VentaHistorico).filter(
                *date_range_filters(VentaHistorico.fecha, fecha_desde, fecha_hasta)
            )
            if filtro_estado == "Vigentes":
                query = query.filter(VentaHistorico.anulada == False)
            elif filtro_estado == "Anuladas":
                query = query.filter(VentaHistorico.anulada == True)
            if filtro_vendedor != "Todos":
                sel_emp = next((e for e in empleados_v if e.nombre == filtro_vendedor), None)
                if sel_emp:
                    query = query.filter(VentaHistorico.vendedor_id == sel_emp.obuma_id)

            ventas = query.order_by(VentaHistorico.fecha.desc()).all()

            if ventas:
                total_neto = sum(
                    (-(v.subtotal or 0) if v.tipo_documento in NC_DOC_TYPES_G else (v.subtotal or 0))
                    for v in ventas if v.tipo_documento in VALID_DOC_TYPES_G
                )
                total_total = sum(v.total or 0 for v in ventas)
                total_margen = sum(v.margen_neto or 0 for v in ventas)

                m1, m2, m3, m4 = st.columns(4)
                with m1:
                    render_metric("Documentos", str(len(ventas)), "📄")
                with m2:
                    render_metric("Total Neto", format_clp(total_neto), "💵", ACCENT_BLUE)
                with m3:
                    render_metric("Total Bruto", format_clp(total_total), "💰", ACCENT_GREEN)
                with m4:
                    render_metric("Margen", format_clp(total_margen), "📈", ACCENT_PURPLE)

                st.markdown("")

                col_vc1, col_vc2 = st.columns(2)
                with col_vc1:
                    st.markdown('<p class="section-header">Ventas Mensuales</p>', unsafe_allow_html=True)
                    df_v_all = pd.DataFrame([{
                        "fecha": v.fecha,
                        "total": v.total or 0,
                        "mes": v.fecha.month if v.fecha else None,
                        "anio": v.fecha.year if v.fecha else None
                    } for v in ventas if v.fecha])
                    if not df_v_all.empty:
                        monthly = df_v_all.groupby(["anio", "mes"])["total"].sum().reset_index()
                        monthly["label"] = monthly.apply(lambda r: f"{MONTH_LABELS[int(r['mes'])-1]} {int(r['anio'])}", axis=1)
                        fig_vm = go.Figure(go.Bar(x=monthly["label"], y=monthly["total"],
                                                   marker_color=ACCENT_BLUE))
                        fig_vm.update_layout(**chart_layout())
                        st.plotly_chart(fig_vm)

                with col_vc2:
                    st.markdown('<p class="section-header">Tipo de Documento</p>', unsafe_allow_html=True)
                    df_tipos = pd.DataFrame([{"tipo": v.tipo_documento or "Otro", "total": v.total or 0} for v in ventas])
                    if not df_tipos.empty:
                        tipos_agg = df_tipos.groupby("tipo")["total"].sum().reset_index()
                        fig_tp = go.Figure(go.Pie(
                            labels=tipos_agg["tipo"], values=tipos_agg["total"],
                            hole=0.4, marker=dict(colors=CHART_COLORS[:len(tipos_agg)]),
                            textinfo="label+percent", textfont=dict(color="white")
                        ))
                        fig_tp.update_layout(**chart_layout())
                        st.plotly_chart(fig_tp)

                st.markdown("")
                data = []
                for v in ventas:
                    data.append({
                        "Fecha": str(v.fecha)[:10] if v.fecha else "-",
                        "Tipo": v.tipo_documento or "-",
                        "Folio": v.folio or "-",
                        "Neto": format_clp(v.subtotal),
                        "IVA": format_clp(v.impuestos),
                        "Total": format_clp(v.total),
                        "Costo": format_clp(v.costo_total),
                        "Margen": format_clp(v.margen_neto),
                        "Vendedor": vend_map_v.get(v.vendedor_id, "-") if v.vendedor_id else "-",
                        "Estado": "Anulada" if v.anulada else "Vigente",
                    })
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("No hay ventas en el rango seleccionado.")

        with tab_items:
            items = db.query(VentaItem).order_by(VentaItem.id.desc()).limit(500).all()
            if items:
                st.metric("Total Items", len(items))
                data = [{
                    "Venta ID": i.venta_id_obuma or "-",
                    "Producto": i.producto_nombre or "-",
                    "SKU": i.producto_sku or "-",
                    "Cantidad": i.cantidad or 0,
                    "Precio Unit.": format_clp(i.precio_unitario),
                    "Total": format_clp(i.total),
                } for i in items]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin items de venta. Sincronice con Obuma.")

        with tab_cotiz:
            cotizaciones = db.query(VentaCotizacion).order_by(VentaCotizacion.id.desc()).limit(200).all()
            if cotizaciones:
                st.metric("Total Cotizaciones", len(cotizaciones))
                data = [{
                    "Folio": c.folio or "-",
                    "Fecha": str(c.fecha)[:10] if c.fecha else "-",
                    "Cliente": c.cliente_nombre or "-",
                    "Total": format_clp(c.total),
                    "Estado": c.estado or "-",
                } for c in cotizaciones]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin cotizaciones. Sincronice con Obuma.")

        with tab_cobros:
            cobros = db.query(VentaCobro).order_by(VentaCobro.id.desc()).limit(200).all()
            if cobros:
                total_cobros = sum(c.monto or 0 for c in cobros)
                c1, c2 = st.columns(2)
                with c1:
                    render_metric("Total Cobros", str(len(cobros)), "🧾")
                with c2:
                    render_metric("Monto Total", format_clp(total_cobros), "💵", ACCENT_GREEN)
                st.markdown("")
                data = [{
                    "Fecha": str(c.fecha)[:10] if c.fecha else "-",
                    "Monto": format_clp(c.monto),
                    "Forma Pago": c.forma_pago or "-",
                    "Estado": c.estado or "-",
                } for c in cobros]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin cobros registrados. Sincronice con Obuma.")

        with tab_dte:
            dtes = db.query(VentaDte).order_by(VentaDte.id.desc()).limit(200).all()
            if dtes:
                st.metric("Total DTE Emitidos", len(dtes))
                data = [{
                    "Tipo DTE": d.tipo_dcto or "-",
                    "Folio": d.folio or "-",
                    "Fecha": str(d.fecha)[:10] if d.fecha else "-",
                    "Total": format_clp(d.monto_total),
                    "Estado SII": d.estado_sii or "-",
                } for d in dtes]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin DTE emitidos. Sincronice con Obuma.")
    finally:
        db.close()


# ============================================================
# CLIENTES
# ============================================================
elif page == "Clientes":
    st.markdown('<p class="page-title">Gestion de Clientes</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Clientes, contactos, direcciones y analisis de actividad</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        tab_clientes, tab_contactos, tab_direcciones = st.tabs([
            "Clientes", "Contactos", "Direcciones"
        ])

        with tab_clientes:
            fc1, fc2 = st.columns([3, 1])
            with fc1:
                search_term = st.text_input("Buscar por nombre o RUT", "", key="cli_search")
            with fc2:
                show_inactive = st.checkbox("Mostrar inactivos", value=False, key="cli_inactive")

            total_activos = db.query(func.count(ClienteFinal.id)).filter(ClienteFinal.activo == True).scalar() or 0
            total_inactivos = db.query(func.count(ClienteFinal.id)).filter(ClienteFinal.activo == False).scalar() or 0
            total_db = total_activos + total_inactivos

            clientes_q = db.query(ClienteFinal)
            if not show_inactive:
                clientes_q = clientes_q.filter(ClienteFinal.activo == True)
            if search_term:
                clientes_q = clientes_q.filter(
                    (ClienteFinal.nombre.ilike(f"%{search_term}%")) |
                    (ClienteFinal.rut.ilike(f"%{search_term}%"))
                )
            clientes = clientes_q.order_by(ClienteFinal.nombre).all()

            if clientes:
                m1, m2, m3 = st.columns(3)
                with m1:
                    render_metric("Total en DB", f"{total_db:,}", "🗄️", ACCENT_BLUE)
                with m2:
                    render_metric("Activos", f"{total_activos:,}", "✅", "#27AE60")
                with m3:
                    render_metric("Inactivos", f"{total_inactivos:,}", "⛔", "#E74C3C")
                st.markdown("")

                col_cc1, col_cc2 = st.columns(2)

                with col_cc1:
                    st.markdown('<p class="section-header">Top Clientes por Facturacion</p>', unsafe_allow_html=True)
                    top_cl = db.query(
                        ClienteFinal.nombre,
                        func.sum(
                            sql_case(
                                (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -VentaHistorico.subtotal),
                                else_=VentaHistorico.subtotal
                            )
                        ).label("total")
                    ).join(VentaHistorico, VentaHistorico.cliente_id == ClienteFinal.id
                    ).filter(VentaHistorico.anulada == False,
                             VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G)
                    ).group_by(ClienteFinal.nombre
                    ).order_by(func.sum(
                        sql_case(
                            (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -VentaHistorico.subtotal),
                            else_=VentaHistorico.subtotal
                        )
                    ).desc()).limit(10).all()

                    if top_cl:
                        df_tcl = pd.DataFrame([{"Cliente": c.nombre, "Total": c.total or 0} for c in top_cl])
                        fig_cl = go.Figure(go.Bar(
                            x=df_tcl["Total"], y=df_tcl["Cliente"], orientation='h',
                            marker_color=ACCENT_BLUE,
                            text=df_tcl["Total"].apply(lambda x: format_clp(x)),
                            textposition="auto"
                        ))
                        fig_cl.update_layout(**chart_layout(height=350))
                        fig_cl.update_layout(yaxis=dict(autorange="reversed", showgrid=False))
                        st.plotly_chart(fig_cl)

                with col_cc2:
                    st.markdown('<p class="section-header">Actividad de Clientes (Meses con Compras)</p>', unsafe_allow_html=True)
                    activity = db.query(
                        ClienteFinal.nombre,
                        func.count(distinct(extract('month', VentaHistorico.fecha))).label("meses")
                    ).join(VentaHistorico, VentaHistorico.cliente_id == ClienteFinal.id
                    ).filter(VentaHistorico.fecha.isnot(None)
                    ).group_by(ClienteFinal.nombre
                    ).order_by(func.count(distinct(extract('month', VentaHistorico.fecha))).desc()).limit(10).all()

                    if activity:
                        df_act = pd.DataFrame([{"Cliente": a.nombre, "Meses": a.meses or 0} for a in activity])
                        fig_act = go.Figure(go.Bar(
                            x=df_act["Meses"], y=df_act["Cliente"], orientation='h',
                            marker_color=ACCENT_GREEN,
                            text=df_act["Meses"], textposition="auto"
                        ))
                        fig_act.update_layout(**chart_layout(height=350))
                        fig_act.update_layout(yaxis=dict(autorange="reversed", showgrid=False),
                                              xaxis=dict(title="Meses con compras"))
                        st.plotly_chart(fig_act)

                st.markdown("")
                data = []
                for c in clientes:
                    n_ventas = db.query(VentaHistorico).filter(VentaHistorico.cliente_id == c.id).count()
                    total = db.query(func.sum(
                        sql_case(
                            (VentaHistorico.tipo_documento.in_(NC_DOC_TYPES_G), -VentaHistorico.subtotal),
                            else_=VentaHistorico.subtotal
                        )
                    )).filter(VentaHistorico.cliente_id == c.id, VentaHistorico.anulada == False,
                              VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES_G)).scalar() or 0
                    data.append({
                        "Nombre": c.nombre,
                        "RUT": c.rut or "-",
                        "Email": c.email or "-",
                        "Telefono": c.telefono or "-",
                        "Ventas": n_ventas,
                        "Total Facturado": format_clp(total),
                        "Activo": "Si" if c.activo else "No",
                    })
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("No hay clientes que coincidan con la busqueda.")

        with tab_contactos:
            contactos = db.query(ClienteContacto).all()
            if contactos:
                st.metric("Total Contactos", len(contactos))
                data = [{
                    "Nombre": c.nombre or "-",
                    "Email": c.email or "-",
                    "Telefono": c.telefono or "-",
                    "Cargo": c.cargo or "-",
                    "Cliente ID": c.cliente_id_obuma or "-",
                } for c in contactos]
                st.dataframe(pd.DataFrame(data), hide_index=True)
            else:
                st.info("Sin contactos de clientes. Sincronice con Obuma.")

        with tab_direcciones:
            direcciones = db.query(ClienteDireccion).all()
            if direcciones:
                st.metric("Total Direcciones", len(direcciones))
                data = [{
                    "Direccion": d.direccion or "-",
                    "Ciudad": d.ciudad or "-",
                    "Comuna": d.comuna or "-",
                    "Region": d.region or "-",
                    "Cliente ID": d.cliente_id_obuma or "-",
                } for d in direcciones]
                st.dataframe(pd.DataFrame(data), hide_index=True)
            else:
                st.info("Sin direcciones de clientes. Sincronice con Obuma.")
    finally:
        db.close()


# ============================================================
# PROVEEDORES
# ============================================================
elif page == "Proveedores":
    st.markdown('<p class="page-title">Proveedores</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Proveedores registrados desde Obuma ERP</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        proveedores = db.query(Proveedor).filter(Proveedor.activo == True).all()
        total_prov = len(proveedores)

        render_metric("Total Proveedores", str(total_prov), "🏭", ACCENT_BLUE)
        st.markdown("")

        if proveedores:
            data = [{
                "Razon Social": p.razon_social or p.nombre_fantasia or "-",
                "RUT": p.rut or "-",
                "Email": p.email or "-",
                "Telefono": p.telefono or "-",
                "Direccion": (p.direccion or "-")[:60],
            } for p in proveedores]
            st.dataframe(pd.DataFrame(data), hide_index=True)
        else:
            st.info("Sin proveedores registrados. Sincronice con Obuma.")
    finally:
        db.close()


# ============================================================
# PRODUCTOS
# ============================================================
elif page == "Productos":
    st.markdown('<p class="page-title">Productos</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Productos, categorias, precios y fabricantes</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        tab_prod, tab_cat, tab_subcat, tab_fab, tab_precios = st.tabs([
            "Productos", "Categorias", "Subcategorias", "Fabricantes", "Precios"
        ])

        with tab_prod:
            productos = db.query(Producto).all()
            activos = [p for p in productos if p.activo]
            c1, c2, c3 = st.columns(3)
            with c1:
                render_metric("Total Productos", str(len(productos)), "📦")
            with c2:
                render_metric("Activos", str(len(activos)), "✅", ACCENT_GREEN)
            with c3:
                alertas = [p for p in activos if p.stock_actual <= p.stock_minimo]
                render_metric("Alerta Stock", str(len(alertas)), "⚠️", ACCENT_RED)

            st.markdown("")
            if productos:
                data = [{
                    "Nombre": p.nombre,
                    "SKU": p.sku or "-",
                    "Categoria": p.categoria or "-",
                    "Precio Venta": format_clp(p.precio_venta),
                    "Costo": format_clp(p.costo),
                    "Stock": p.stock_actual or 0,
                    "Activo": "Si" if p.activo else "No",
                } for p in productos]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin productos. Sincronice con Obuma.")

        with tab_cat:
            categorias = db.query(ProductoCategoria).all()
            if categorias:
                st.metric("Total Categorias", len(categorias))
                data = [{
                    "ID": c.obuma_id or "-",
                    "Nombre": c.nombre or "-",
                } for c in categorias]
                st.dataframe(pd.DataFrame(data), hide_index=True)
            else:
                st.info("Sin categorias. Sincronice con Obuma.")

        with tab_subcat:
            subcategorias = db.query(ProductoSubCategoria).all()
            if subcategorias:
                st.metric("Total Subcategorias", len(subcategorias))
                data = [{
                    "ID": s.obuma_id or "-",
                    "Nombre": s.nombre or "-",
                    "Categoria ID": s.categoria_id_obuma or "-",
                } for s in subcategorias]
                st.dataframe(pd.DataFrame(data), hide_index=True)
            else:
                st.info("Sin subcategorias. Sincronice con Obuma.")

        with tab_fab:
            fabricantes = db.query(ProductoFabricante).all()
            if fabricantes:
                st.metric("Total Fabricantes", len(fabricantes))
                data = [{
                    "ID": f.obuma_id or "-",
                    "Nombre": f.nombre or "-",
                } for f in fabricantes]
                st.dataframe(pd.DataFrame(data), hide_index=True)
            else:
                st.info("Sin fabricantes. Sincronice con Obuma.")

        with tab_precios:
            precios = db.query(ProductoPrecio).all()
            if precios:
                st.metric("Total Listas de Precios", len(precios))
                data = [{
                    "Producto ID": p.producto_id_obuma or "-",
                    "Producto": p.producto_nombre or "-",
                    "Lista": p.lista_precio or "-",
                    "Precio": format_clp(p.precio),
                } for p in precios]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin listas de precios. Sincronice con Obuma.")
    finally:
        db.close()


# ============================================================
# EMPLEADOS
# ============================================================
elif page == "Empleados":
    st.markdown('<p class="page-title">Empleados</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Empleados y remuneraciones desde Obuma ERP</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        tab_emp, tab_rem = st.tabs(["Empleados", "Remuneraciones"])

        with tab_emp:
            empleados = db.query(Empleado).all()
            activos = [e for e in empleados if e.activo]
            c1, c2 = st.columns(2)
            with c1:
                render_metric("Total Empleados", str(len(empleados)), "👷")
            with c2:
                render_metric("Activos", str(len(activos)), "✅", ACCENT_GREEN)

            st.markdown("")
            if empleados:
                data = [{
                    "Nombre": e.nombre or "-",
                    "RUT": e.rut or "-",
                    "Email": e.email or "-",
                    "Cargo": e.cargo or "-",
                    "Activo": "Si" if e.activo else "No",
                } for e in empleados]
                st.dataframe(pd.DataFrame(data), hide_index=True)
            else:
                st.info("Sin empleados. Sincronice con Obuma.")

        with tab_rem:
            remuneraciones = db.query(Remuneracion).order_by(Remuneracion.id.desc()).limit(200).all()
            if remuneraciones:
                total_haberes = sum(r.total_haberes or 0 for r in remuneraciones)
                total_liquido = sum(r.liquido or 0 for r in remuneraciones)
                c1, c2, c3 = st.columns(3)
                with c1:
                    render_metric("Registros", str(len(remuneraciones)), "📋")
                with c2:
                    render_metric("Total Haberes", format_clp(total_haberes), "💵", ACCENT_BLUE)
                with c3:
                    render_metric("Total Liquido", format_clp(total_liquido), "💰", ACCENT_GREEN)

                st.markdown("")
                data = [{
                    "Empleado RUT": r.empleado_rut or "-",
                    "Periodo": r.periodo or "-",
                    "Haberes": format_clp(r.total_haberes),
                    "Descuentos": format_clp(r.total_descuentos),
                    "Liquido": format_clp(r.liquido),
                } for r in remuneraciones]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin remuneraciones. Sincronice con Obuma.")
    finally:
        db.close()


# ============================================================
# COMPRAS
# ============================================================
elif page == "Compras":
    st.markdown('<p class="page-title">Compras</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Compras, ordenes de compra, pagos y DTE recibidos</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        tab_compras, tab_oc, tab_pagos = st.tabs([
            "Compras", "Ordenes de Compra", "Pagos"
        ])

        with tab_compras:
            compras = db.query(CompraHistorico).order_by(CompraHistorico.fecha.desc()).all()
            if compras:
                total_compras = sum(c.total or 0 for c in compras)
                c1, c2 = st.columns(2)
                with c1:
                    render_metric("Total Compras", str(len(compras)), "🛒")
                with c2:
                    render_metric("Monto Total", format_clp(total_compras), "💵", ACCENT_AMBER)

                st.markdown("")
                data = [{
                    "Fecha": str(c.fecha)[:10] if c.fecha else "-",
                    "Proveedor": c.proveedor or "-",
                    "Folio": c.folio or "-",
                    "Total": format_clp(c.total),
                    "Estado": c.estado or "-",
                } for c in compras]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin compras registradas. Sincronice con Obuma.")

        with tab_oc:
            ocs = db.query(CompraOC).order_by(CompraOC.id.desc()).limit(200).all()
            if ocs:
                st.metric("Total OC", len(ocs))
                data = [{
                    "Folio": o.folio or "-",
                    "Fecha": str(o.fecha)[:10] if o.fecha else "-",
                    "Proveedor": o.proveedor or "-",
                    "Total": format_clp(o.total),
                    "Estado": o.estado or "-",
                } for o in ocs]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin ordenes de compra. Sincronice con Obuma.")

        with tab_pagos:
            pagos = db.query(CompraPago).order_by(CompraPago.id.desc()).limit(200).all()
            if pagos:
                total_pagos = sum(p.monto or 0 for p in pagos)
                c1, c2 = st.columns(2)
                with c1:
                    render_metric("Total Pagos", str(len(pagos)), "💳")
                with c2:
                    render_metric("Monto Total", format_clp(total_pagos), "💵", ACCENT_RED)

                st.markdown("")
                data = [{
                    "Fecha": str(p.fecha)[:10] if p.fecha else "-",
                    "Monto": format_clp(p.monto),
                    "Forma Pago": p.forma_pago or "-",
                    "Origen": p.origen or "-",
                } for p in pagos]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
            else:
                st.info("Sin pagos a proveedores. Sincronice con Obuma.")

    finally:
        db.close()


# ============================================================
# CONTABILIDAD
# ============================================================
elif page == "Contabilidad":
    st.markdown('<p class="page-title">Contabilidad</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Libro diario y registros contables</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        total_debe = db.query(func.sum(ContabilidadHistorico.debe)).scalar() or 0
        total_haber = db.query(func.sum(ContabilidadHistorico.haber)).scalar() or 0
        balance = total_haber - total_debe

        c1, c2, c3 = st.columns(3)
        with c1:
            render_metric("Total Ingresos (Haber)", format_clp(total_haber), "📥", ACCENT_GREEN)
        with c2:
            render_metric("Total Egresos (Debe)", format_clp(total_debe), "📤", ACCENT_RED)
        with c3:
            color = ACCENT_GREEN if balance >= 0 else ACCENT_RED
            render_metric("Balance", format_clp(balance), "⚖️", color)

        st.markdown("---")

        entries = db.query(ContabilidadHistorico).order_by(
            ContabilidadHistorico.fecha.desc()
        ).limit(200).all()

        if entries:
            df = pd.DataFrame([{
                "Fecha": str(e.fecha) if e.fecha else "-",
                "Cuenta": e.cuenta or "-",
                "Descripcion": (e.descripcion or "-")[:60],
                "Debe": format_clp(e.debe),
                "Haber": format_clp(e.haber),
            } for e in entries])
            st.dataframe(df, hide_index=True, height=400)
        else:
            st.info("No hay registros de contabilidad. Sincronice con Obuma primero.")
    finally:
        db.close()


# ============================================================
# CRM
# ============================================================
elif page == "CRM":
    st.markdown('<p class="page-title">CRM - Leads</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Oportunidades comerciales y seguimiento de leads</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        leads = db.query(CrmLead).order_by(CrmLead.id.desc()).all()

        if leads:
            total_monto = sum(l.monto_estimado or 0 for l in leads)
            c1, c2, c3 = st.columns(3)
            with c1:
                render_metric("Total Leads", str(len(leads)), "🎯")
            with c2:
                render_metric("Monto Estimado", format_clp(total_monto), "💰", ACCENT_GREEN)
            with c3:
                activos = [l for l in leads if l.estado and l.estado.lower() not in ("cerrado", "perdido")]
                render_metric("Activos", str(len(activos)), "🔥", ACCENT_AMBER)

            st.markdown("")
            data = [{
                "Nombre": l.nombre or "-",
                "Empresa": l.empresa or "-",
                "Email": l.email or "-",
                "Telefono": l.telefono or "-",
                "Estado": l.estado or "-",
                "Monto Est.": format_clp(l.monto_estimado),
            } for l in leads]
            st.dataframe(pd.DataFrame(data), hide_index=True, height=400)
        else:
            st.info("Sin leads CRM. Sincronice con Obuma.")
    finally:
        db.close()


# ============================================================
# API OBUMA
# ============================================================
elif page == "API Obuma":
    st.markdown('<p class="page-title">Catalogo API Obuma</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">30 endpoints registrados para automatizaciones</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        endpoints = db.query(ObumaApiEndpoint).order_by(
            ObumaApiEndpoint.categoria_orden, ObumaApiEndpoint.id
        ).all()

        total_eps = len(endpoints)
        implementados = sum(1 for e in endpoints if e.implementado)
        sync_activos = sum(1 for e in endpoints if e.sync_habilitado)
        no_disp = sum(1 for e in endpoints if e.estado == "no_disponible")

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            render_metric("Total Endpoints", str(total_eps), "🔌")
        with c2:
            render_metric("Implementados", str(implementados), "✅", ACCENT_GREEN)
        with c3:
            render_metric("Sync Activos", str(sync_activos), "🔄", ACCENT_BLUE)
        with c4:
            render_metric("No Disponibles", str(no_disp), "❌", ACCENT_RED)

        st.markdown("---")

        categorias_vistas = []
        for ep in endpoints:
            if ep.categoria not in categorias_vistas:
                categorias_vistas.append(ep.categoria)
                st.markdown(f'<p class="section-header">{ep.categoria_orden}. {ep.categoria}</p>', unsafe_allow_html=True)

            if ep.estado == "no_disponible":
                badge = f'<span style="background:rgba(239,68,68,0.15);color:#f87171;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600;">NO DISPONIBLE</span>'
            elif ep.implementado and ep.sync_habilitado:
                badge = f'<span style="background:rgba(16,185,129,0.15);color:#34d399;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600;">SYNC ACTIVO</span>'
            elif ep.implementado:
                badge = f'<span style="background:rgba(59,130,246,0.15);color:#60a5fa;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600;">IMPLEMENTADO</span>'
            elif ep.endpoint_url:
                badge = f'<span style="background:rgba(245,158,11,0.15);color:#fbbf24;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600;">DISPONIBLE</span>'
            else:
                badge = f'<span style="background:rgba(148,163,184,0.15);color:#94a3b8;padding:2px 10px;border-radius:12px;font-size:0.75rem;font-weight:600;">REFERENCIA</span>'

            metodo_color = {"GET": ACCENT_GREEN, "POST": ACCENT_AMBER, "-": TEXT_SECONDARY}.get(ep.metodo_http, TEXT_SECONDARY)
            endpoint_display = f'<code style="background:rgba(59,130,246,0.1);color:#60a5fa;padding:2px 8px;border-radius:4px;font-size:0.8rem;">{ep.endpoint_url}</code>' if ep.endpoint_url else '<span style="color:#64748b;font-size:0.8rem;">Sin endpoint directo</span>'

            sync_info = ""
            if ep.sync_habilitado and ep.registros_sync and ep.registros_sync > 0:
                sync_info = f' | <strong style="color:{ACCENT_GREEN};">{ep.registros_sync}</strong> registros'
            if ep.ultima_sync:
                sync_info += f' | Sync: {str(ep.ultima_sync)[:16]}'

            st.markdown(f"""
            <div class="api-card">
                <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:0.4rem;">
                    <div style="display:flex;align-items:center;gap:0.6rem;">
                        <span style="background:{metodo_color};color:white;padding:2px 8px;border-radius:4px;font-size:0.7rem;font-weight:700;">{ep.metodo_http}</span>
                        <strong style="font-size:0.95rem;color:{TEXT_PRIMARY};">{ep.nombre}</strong>
                    </div>
                    {badge}
                </div>
                <div style="margin:0.3rem 0;">{endpoint_display}</div>
                <div style="font-size:0.82rem;color:{TEXT_SECONDARY};margin:0.3rem 0;">{ep.descripcion or ''}</div>
                <div style="font-size:0.75rem;color:#64748b;margin-top:0.4rem;">{sync_info}</div>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("---")
        st.markdown('<p class="section-header">Probar Endpoint</p>', unsafe_allow_html=True)
        ep_options = [e for e in endpoints if e.endpoint_url and "{" not in (e.endpoint_url or "")]
        if ep_options:
            selected_ep = st.selectbox(
                "Seleccionar endpoint",
                options=ep_options,
                format_func=lambda e: f"{e.categoria} > {e.nombre} ({e.endpoint_url})"
            )
            if st.button("Probar Conexion", type="primary"):
                with st.spinner(f"Consultando {selected_ep.endpoint_url}..."):
                    client = ObumaClient()
                    result = run_async(client.test_endpoint(selected_ep.endpoint_url))
                    if "error" in result:
                        st.error(f"Error: {result.get('error', '')[:200]}")
                    else:
                        total_items = result.get("data-total-items", result.get("data-actual-total", "?"))
                        data_list = result.get("data", [])
                        st.success(f"Conexion exitosa. Total registros: {total_items}")
                        if data_list and len(data_list) > 0:
                            st.json(data_list[:3])
    finally:
        db.close()


# ============================================================
# REPORTES
# ============================================================
elif page == "Reportes":
    st.markdown('<p class="page-title">Centro de Reportes</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Genera, programa y gestiona reportes de ventas por vendedor</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        vendedores_all = db.query(Empleado).filter(Empleado.activo == True).order_by(Empleado.nombre).all()
        tracked_vends = [v for v in vendedores_all if v.obuma_id in TRACKED_VENDEDOR_IDS]
        other_vends = [v for v in vendedores_all if v.obuma_id not in TRACKED_VENDEDOR_IDS]
        sorted_vendedores = tracked_vends + other_vends
        vendedor_map = {f"{v.nombre} ({v.cargo or 'Vendedor'})": v.obuma_id for v in sorted_vendedores}
        vendedor_name_map = {v.obuma_id: v.nombre for v in vendedores_all}
        vendedor_display_list = list(vendedor_map.keys())

        def calc_proxima_ejecucion(freq, dia_semana, dia_mes, hora, minuto, ultima_ej, created):
            now = datetime.now()
            base_time = now.replace(hour=hora, minute=minuto, second=0, microsecond=0)
            freq_lower = (freq or "").lower()
            if freq_lower == "diario":
                nxt = base_time + timedelta(days=1) if now >= base_time else base_time
            elif freq_lower == "semanal":
                ds = dia_semana if dia_semana is not None else 0
                days_ahead = ds - now.weekday()
                if days_ahead < 0 or (days_ahead == 0 and now >= base_time):
                    days_ahead += 7
                nxt = (now + timedelta(days=days_ahead)).replace(hour=hora, minute=minuto, second=0, microsecond=0)
            elif freq_lower == "quincenal":
                ref = ultima_ej or created or now
                if isinstance(ref, date) and not isinstance(ref, datetime):
                    ref = datetime.combine(ref, datetime.min.time())
                nxt = ref + timedelta(days=14)
                nxt = nxt.replace(hour=hora, minute=minuto, second=0, microsecond=0)
                while nxt <= now:
                    nxt += timedelta(days=14)
            elif freq_lower == "mensual":
                dm = dia_mes if dia_mes else 1
                try:
                    nxt = now.replace(day=dm, hour=hora, minute=minuto, second=0, microsecond=0)
                except ValueError:
                    nxt = now.replace(day=28, hour=hora, minute=minuto, second=0, microsecond=0)
                if nxt <= now:
                    month = nxt.month + 1
                    year = nxt.year
                    if month > 12:
                        month = 1
                        year += 1
                    try:
                        nxt = nxt.replace(year=year, month=month, day=dm)
                    except ValueError:
                        nxt = nxt.replace(year=year, month=month, day=28)
            else:
                nxt = base_time + timedelta(days=1)
            return nxt

        tab_generar, tab_programar, tab_historial, tab_config = st.tabs([
            "📄 Generar Ahora", "⏰ Programar Envio", "📋 Historial", "⚙️ Configuracion Email"
        ])

        with tab_generar:
            col_ind, col_all = st.columns(2, gap="large")

            with col_ind:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-icon">📊</div>
                    <p class="metric-label">REPORTE INDIVIDUAL</p>
                    <p style="color:{TEXT_SECONDARY};font-size:0.85rem;margin:0;">Genera reporte para un vendedor especifico</p>
                </div>
                """, unsafe_allow_html=True)
                st.markdown("")

                rpt_vendedor_sel = st.selectbox("Vendedor", vendedor_display_list, key="rpt_ind_vendedor")
                rpt_ind_mode = st.radio("Periodo", ["Por Ano", "Rango Personalizado"], horizontal=True, key="rpt_ind_mode")

                if rpt_ind_mode == "Por Ano":
                    rpt_ind_year = st.number_input("Ano", value=date.today().year, min_value=2020, max_value=2030, key="rpt_ind_year")
                    rpt_ind_from = date(int(rpt_ind_year), 1, 1)
                    rpt_ind_to = date(int(rpt_ind_year), 12, 31)
                else:
                    rc1, rc2 = st.columns(2)
                    with rc1:
                        rpt_ind_from = st.date_input("Desde", value=date(date.today().year, 1, 1), key="rpt_ind_from")
                    with rc2:
                        rpt_ind_to = st.date_input("Hasta", value=date.today(), key="rpt_ind_to")

                if st.button("Generar Reporte Individual", type="primary", use_container_width=True, key="rpt_gen_ind"):
                    vid = vendedor_map[rpt_vendedor_sel]
                    with st.spinner(f"Generando reporte para {rpt_vendedor_sel}..."):
                        try:
                            fp = generate_vendedor_report(db, vid, rpt_ind_from, rpt_ind_to)
                            if fp:
                                st.session_state["rpt_last_files"] = [fp]
                                st.session_state["rpt_last_type"] = "individual"
                                st.session_state["rpt_last_date_from"] = rpt_ind_from
                                st.session_state["rpt_last_date_to"] = rpt_ind_to
                                st.success(f"Reporte generado: {os.path.basename(fp)}")
                            else:
                                st.warning("Este vendedor no tiene ventas en el periodo seleccionado.")
                        except Exception as e:
                            st.error(f"Error: {str(e)}")

            with col_all:
                st.markdown(f"""
                <div class="metric-card">
                    <div class="metric-icon">👥</div>
                    <p class="metric-label">REPORTE TODOS LOS VENDEDORES</p>
                    <p style="color:{TEXT_SECONDARY};font-size:0.85rem;margin:0;">Genera reportes para los 5 vendedores principales</p>
                </div>
                """, unsafe_allow_html=True)
                st.markdown("")

                rpt_all_mode = st.radio("Periodo", ["Por Ano", "Rango Personalizado"], horizontal=True, key="rpt_all_mode")

                if rpt_all_mode == "Por Ano":
                    rpt_all_year = st.number_input("Ano", value=date.today().year, min_value=2020, max_value=2030, key="rpt_all_year")
                    rpt_all_from = date(int(rpt_all_year), 1, 1)
                    rpt_all_to = date(int(rpt_all_year), 12, 31)
                else:
                    ac1, ac2 = st.columns(2)
                    with ac1:
                        rpt_all_from = st.date_input("Desde", value=date(date.today().year, 1, 1), key="rpt_all_from")
                    with ac2:
                        rpt_all_to = st.date_input("Hasta", value=date.today(), key="rpt_all_to")

                if st.button("Generar Todos los Reportes", type="primary", use_container_width=True, key="rpt_gen_all"):
                    with st.spinner("Generando reportes para todos los vendedores..."):
                        try:
                            fps = generate_all_vendedor_reports(db, rpt_all_from, rpt_all_to)
                            if fps:
                                st.session_state["rpt_last_files"] = fps
                                st.session_state["rpt_last_type"] = "todos"
                                st.session_state["rpt_last_date_from"] = rpt_all_from
                                st.session_state["rpt_last_date_to"] = rpt_all_to
                                st.success(f"Se generaron {len(fps)} reportes exitosamente")
                            else:
                                st.warning("No se encontraron ventas en el periodo seleccionado.")
                        except Exception as e:
                            st.error(f"Error: {str(e)}")

            if st.session_state.get("rpt_last_files"):
                st.markdown("---")
                st.markdown('<p class="section-header">Reportes Generados</p>', unsafe_allow_html=True)

                for idx, fp in enumerate(st.session_state["rpt_last_files"]):
                    if os.path.exists(fp):
                        dc1, dc2 = st.columns([3, 1])
                        with dc1:
                            st.markdown(f"📄 **{os.path.basename(fp)}**")
                        with dc2:
                            with open(fp, "rb") as f:
                                st.download_button(
                                    label="⬇️ Descargar",
                                    data=f.read(),
                                    file_name=os.path.basename(fp),
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"rpt_dl_{idx}",
                                )

                with st.expander("📧 Enviar por Email", expanded=False):
                    rpt_emails = st.text_input(
                        "Emails (separados por coma)",
                        placeholder="gabriel@vlsur.cl, jhonatan@vlsur.cl",
                        key="rpt_send_emails"
                    )
                    if st.button("Enviar Reportes por Email", type="primary", key="rpt_send_email_btn"):
                        if not rpt_emails.strip():
                            st.error("Ingrese al menos un email.")
                        else:
                            email_list = [e.strip() for e in rpt_emails.replace("\n", ",").split(",") if e.strip()]
                            with st.spinner("Enviando email..."):
                                stored_from = st.session_state.get("rpt_last_date_from", "")
                                stored_to = st.session_state.get("rpt_last_date_to", "")
                                date_range_str = f"{stored_from} - {stored_to}"
                                body = build_report_email_html(
                                    "Vendedores",
                                    st.session_state.get("rpt_last_type", "reporte"),
                                    date_range_str
                                )
                                result = send_report_email(
                                    email_list,
                                    "Reporte de Ventas - BI Platform",
                                    body,
                                    st.session_state["rpt_last_files"]
                                )
                                if result.get("success"):
                                    st.success(f"Email enviado a {', '.join(email_list)}")
                                else:
                                    st.error(f"Error enviando email: {result.get('error', 'Error desconocido')}")

        with tab_programar:
            st.markdown('<p class="section-header">Crear Nueva Programacion</p>', unsafe_allow_html=True)

            with st.form("rpt_schedule_form", clear_on_submit=True):
                rpt_sched_name = st.text_input("Nombre del Reporte", value="Reporte Semanal Ventas", key="rpt_sched_name")

                sc1, sc2 = st.columns(2)
                with sc1:
                    rpt_sched_tipo = st.selectbox("Tipo de Reporte", ["Individual (1 vendedor)", "Todos los vendedores"], key="rpt_sched_tipo")
                with sc2:
                    rpt_sched_vendedor = st.selectbox("Vendedor (solo si individual)", vendedor_display_list, key="rpt_sched_vendedor")

                sc3, sc4 = st.columns(2)
                with sc3:
                    rpt_sched_freq = st.selectbox("Frecuencia", ["Diario", "Semanal", "Quincenal", "Mensual"], index=1, key="rpt_sched_freq")
                with sc4:
                    rpt_sched_dia_semana = st.selectbox("Dia de la Semana (si semanal)", ["Lunes", "Martes", "Miercoles", "Jueves", "Viernes"], key="rpt_sched_dia_semana")

                sc5, sc6, sc7 = st.columns(3)
                with sc5:
                    rpt_sched_dia_mes = st.number_input("Dia del Mes (si mensual)", min_value=1, max_value=28, value=1, key="rpt_sched_dia_mes")
                with sc6:
                    rpt_sched_hora = st.number_input("Hora", min_value=0, max_value=23, value=8, key="rpt_sched_hora")
                with sc7:
                    rpt_sched_min = st.selectbox("Minuto", [0, 15, 30, 45], key="rpt_sched_min")

                rpt_sched_emails = st.text_area(
                    "Emails Destino (uno por linea o separados por coma)",
                    placeholder="gabriel@vlsur.cl\njhonatan@vlsur.cl",
                    key="rpt_sched_emails"
                )

                sc8, sc9 = st.columns(2)
                with sc8:
                    rpt_sched_periodo = st.selectbox("Periodo del Reporte", [
                        "Mes Actual", "Mes Anterior", "Ultimo Trimestre", "Ano Actual", "Rango Personalizado"
                    ], key="rpt_sched_periodo")
                with sc9:
                    rpt_sched_custom_from = st.date_input("Desde (si personalizado)", value=date(date.today().year, 1, 1), key="rpt_sched_custom_from")

                rpt_sched_custom_to = st.date_input("Hasta (si personalizado)", value=date.today(), key="rpt_sched_custom_to")

                submitted = st.form_submit_button("✅ Crear Programacion", type="primary", use_container_width=True)

                if submitted:
                    if not rpt_sched_name.strip():
                        st.error("Ingrese un nombre para el reporte.")
                    elif not rpt_sched_emails.strip():
                        st.error("Ingrese al menos un email destino.")
                    else:
                        dia_sem_map = {"Lunes": 0, "Martes": 1, "Miercoles": 2, "Jueves": 3, "Viernes": 4}
                        periodo_map = {
                            "Mes Actual": "mes_actual",
                            "Mes Anterior": "mes_anterior",
                            "Ultimo Trimestre": "ultimo_trimestre",
                            "Ano Actual": "ano_actual",
                            "Rango Personalizado": "personalizado"
                        }
                        FRECUENCIA_MAP = {
                            "Diario": "diario",
                            "Semanal": "semanal",
                            "Quincenal": "quincenal",
                            "Mensual": "mensual",
                        }
                        tipo_val = "individual" if "Individual" in rpt_sched_tipo else "todos"
                        vid_val = vendedor_map.get(rpt_sched_vendedor) if tipo_val == "individual" else None
                        freq_val = FRECUENCIA_MAP.get(rpt_sched_freq, rpt_sched_freq.lower())

                        nuevo = ReporteProgramado(
                            tenant_id=1,
                            nombre=rpt_sched_name.strip(),
                            tipo_reporte=tipo_val,
                            vendedor_obuma_id=vid_val,
                            frecuencia=freq_val,
                            dia_semana=dia_sem_map.get(rpt_sched_dia_semana) if rpt_sched_freq == "Semanal" else None,
                            dia_mes=rpt_sched_dia_mes if rpt_sched_freq == "Mensual" else None,
                            hora=rpt_sched_hora,
                            minuto=rpt_sched_min,
                            emails_destino=rpt_sched_emails.strip(),
                            filtro_fecha_tipo=periodo_map.get(rpt_sched_periodo, "mes_actual"),
                            filtro_fecha_desde=rpt_sched_custom_from if rpt_sched_periodo == "Rango Personalizado" else None,
                            filtro_fecha_hasta=rpt_sched_custom_to if rpt_sched_periodo == "Rango Personalizado" else None,
                            activo=True,
                            proxima_ejecucion=calc_proxima_ejecucion(
                                rpt_sched_freq,
                                dia_sem_map.get(rpt_sched_dia_semana),
                                rpt_sched_dia_mes,
                                rpt_sched_hora,
                                rpt_sched_min,
                                None, None
                            )
                        )
                        db.add(nuevo)
                        db.commit()
                        st.success(f"Programacion '{rpt_sched_name}' creada exitosamente.")
                        st.rerun()

            st.markdown("---")
            st.markdown('<p class="section-header">Programaciones Existentes</p>', unsafe_allow_html=True)

            schedules = db.query(ReporteProgramado).filter(ReporteProgramado.tenant_id == 1).order_by(ReporteProgramado.created_at.desc()).all()

            if not schedules:
                st.info("No hay reportes programados. Use el formulario de arriba para crear uno.")
            else:
                for sched in schedules:
                    freq_display = (sched.frecuencia or "N/A").capitalize()
                    dia_sem_names = {0: "Lunes", 1: "Martes", 2: "Miercoles", 3: "Jueves", 4: "Viernes"}
                    freq_detail = ""
                    freq_lower = (sched.frecuencia or "").lower()
                    if freq_lower == "semanal" and sched.dia_semana is not None:
                        freq_detail = f" - {dia_sem_names.get(sched.dia_semana, 'N/A')}"
                    elif freq_lower == "mensual" and sched.dia_mes:
                        freq_detail = f" - Dia {sched.dia_mes}"
                    hora_str = f"{sched.hora or 0:02d}:{sched.minuto or 0:02d}"

                    tipo_display = "Individual" if sched.tipo_reporte == "individual" else "Todos"
                    if sched.tipo_reporte == "individual" and sched.vendedor_obuma_id:
                        vname = vendedor_name_map.get(sched.vendedor_obuma_id, sched.vendedor_obuma_id)
                        tipo_display = f"Individual - {vname}"

                    emails_preview = (sched.emails_destino or "")[:60]
                    if len(sched.emails_destino or "") > 60:
                        emails_preview += "..."

                    badge_class = "badge-ok" if sched.activo else "badge-warning"
                    badge_text = "Activo" if sched.activo else "Inactivo"

                    prox = sched.proxima_ejecucion or calc_proxima_ejecucion(
                        sched.frecuencia, sched.dia_semana, sched.dia_mes,
                        sched.hora or 0, sched.minuto or 0,
                        sched.ultima_ejecucion, sched.created_at
                    )
                    prox_str = prox.strftime("%d/%m/%Y %H:%M") if prox else "N/A"

                    st.markdown(f"""
                    <div class="metric-card" style="margin-bottom:0.8rem;">
                        <div style="display:flex;justify-content:space-between;align-items:center;">
                            <div>
                                <p style="color:{TEXT_PRIMARY};font-weight:600;font-size:1rem;margin:0;">{sched.nombre}</p>
                                <p style="color:{TEXT_SECONDARY};font-size:0.82rem;margin:4px 0 0;">
                                    {tipo_display} · {freq_display}{freq_detail} · {hora_str} hrs
                                </p>
                                <p style="color:{TEXT_SECONDARY};font-size:0.78rem;margin:2px 0 0;">
                                    📧 {emails_preview}
                                </p>
                                <p style="color:{TEXT_SECONDARY};font-size:0.78rem;margin:2px 0 0;">
                                    Proxima: {prox_str} · Enviados: {sched.total_enviados or 0}
                                </p>
                            </div>
                            <div>
                                <span class="status-badge {badge_class}">{badge_text}</span>
                            </div>
                        </div>
                    </div>
                    """, unsafe_allow_html=True)

                    btn_c1, btn_c2 = st.columns(2)
                    with btn_c1:
                        toggle_label = "⏸ Desactivar" if sched.activo else "▶️ Activar"
                        if st.button(toggle_label, key=f"rpt_toggle_{sched.id}", use_container_width=True):
                            sched.activo = not sched.activo
                            sched.proxima_ejecucion = calc_proxima_ejecucion(
                                sched.frecuencia, sched.dia_semana, sched.dia_mes,
                                sched.hora or 0, sched.minuto or 0,
                                sched.ultima_ejecucion, sched.created_at
                            ) if sched.activo else None
                            db.commit()
                            st.rerun()
                    with btn_c2:
                        if st.button("🗑 Eliminar", key=f"rpt_del_{sched.id}", use_container_width=True):
                            st.session_state[f"rpt_confirm_del_{sched.id}"] = True

                    if st.session_state.get(f"rpt_confirm_del_{sched.id}"):
                        st.warning(f"¿Seguro que desea eliminar '{sched.nombre}'?")
                        cc1, cc2 = st.columns(2)
                        with cc1:
                            if st.button("Si, eliminar", key=f"rpt_confirm_yes_{sched.id}", use_container_width=True):
                                db.delete(sched)
                                db.commit()
                                del st.session_state[f"rpt_confirm_del_{sched.id}"]
                                st.rerun()
                        with cc2:
                            if st.button("Cancelar", key=f"rpt_confirm_no_{sched.id}", use_container_width=True):
                                del st.session_state[f"rpt_confirm_del_{sched.id}"]
                                st.rerun()
                    st.markdown("")

        with tab_historial:
            st.markdown('<p class="section-header">Historial de Reportes Generados</p>', unsafe_allow_html=True)

            hc1, hc2, hc3 = st.columns([2, 2, 1])
            with hc1:
                hist_from = st.date_input("Desde", value=date.today() - timedelta(days=90), key="rpt_hist_from")
            with hc2:
                hist_to = st.date_input("Hasta", value=date.today(), key="rpt_hist_to")
            with hc3:
                hist_tipo = st.selectbox("Tipo", ["Todos", "vendedor", "diario"], key="rpt_hist_tipo")

            hist_q = db.query(ReporteGenerado).filter(
                ReporteGenerado.fecha_reporte >= hist_from,
                ReporteGenerado.fecha_reporte <= hist_to
            )
            if hist_tipo != "Todos":
                hist_q = hist_q.filter(ReporteGenerado.tipo == hist_tipo)

            reportes_hist = hist_q.order_by(ReporteGenerado.generado_at.desc()).limit(100).all()

            if reportes_hist:
                st.markdown(f"<p style='color:{TEXT_SECONDARY};font-size:0.85rem;'>Mostrando {len(reportes_hist)} reportes</p>", unsafe_allow_html=True)
                for r in reportes_hist:
                    hrc1, hrc2, hrc3, hrc4 = st.columns([2, 3, 1, 1])
                    with hrc1:
                        fecha_str = r.generado_at.strftime("%d/%m/%Y %H:%M") if r.generado_at else "N/A"
                        st.markdown(f"<p style='color:{TEXT_SECONDARY};font-size:0.85rem;margin:0;'>{fecha_str}</p>", unsafe_allow_html=True)
                    with hrc2:
                        st.markdown(f"**{r.nombre_archivo}**")
                    with hrc3:
                        badge = "badge-info" if r.tipo == "vendedor" else "badge-ok"
                        st.markdown(f'<span class="status-badge {badge}">{r.tipo}</span>', unsafe_allow_html=True)
                    with hrc4:
                        if r.ruta_archivo and os.path.exists(r.ruta_archivo):
                            with open(r.ruta_archivo, "rb") as f:
                                st.download_button(
                                    label="⬇️",
                                    data=f.read(),
                                    file_name=r.nombre_archivo,
                                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    key=f"rpt_hist_dl_{r.id}"
                                )
                        else:
                            st.markdown(f"<span style='color:{TEXT_SECONDARY};font-size:0.8rem;'>N/A</span>", unsafe_allow_html=True)
            else:
                st.info("No se encontraron reportes para los filtros seleccionados.")

        with tab_config:
            st.markdown('<p class="section-header">Configuracion de Email</p>', unsafe_allow_html=True)

            email_cfg = check_email_config()

            if email_cfg["configured"]:
                is_sandbox = email_cfg.get("sandbox", False)
                status_color = ACCENT_AMBER if is_sandbox else ACCENT_GREEN
                status_icon = "⚠️" if is_sandbox else "✅"
                status_label = "Resend Activo — Dominio sin verificar (modo sandbox)" if is_sandbox else f"Email Configurado · {email_cfg['method']}"
                status_detail = (
                    f"Remitente actual: <code>onboarding@resend.dev</code> (sandbox). Solo puede enviar a la cuenta del propietario de Resend. Para enviar a cualquier destinatario, debes verificar el dominio <strong>vlsur.cl</strong> en Resend."
                    if is_sandbox else
                    f"Metodo: {email_cfg['method']} · {email_cfg['detail']} · Remitente: {email_cfg.get('from_email','')}"
                )
                st.markdown(f"""
                <div class="metric-card" style="border-left:4px solid {status_color};">
                    <div style="display:flex;align-items:center;gap:12px;">
                        <span style="font-size:2rem;">{status_icon}</span>
                        <div>
                            <p style="color:{TEXT_PRIMARY};font-weight:600;margin:0;">{status_label}</p>
                            <p style="color:{TEXT_SECONDARY};font-size:0.85rem;margin:4px 0 0;">{status_detail}</p>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class="metric-card" style="border-left:4px solid {ACCENT_RED};">
                    <div style="display:flex;align-items:center;gap:12px;">
                        <span style="font-size:2rem;">❌</span>
                        <div>
                            <p style="color:{TEXT_PRIMARY};font-weight:600;margin:0;">Email No Configurado</p>
                            <p style="color:{TEXT_SECONDARY};font-size:0.85rem;margin:4px 0 0;">{email_cfg['detail']}</p>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("")

            # Estado de las alertas a admin (ADMIN_ALERT_EMAILS)
            _admin_cfg_tab = check_admin_alert_config()
            if _admin_cfg_tab["configured"]:
                _emails_html = ", ".join(_admin_cfg_tab["emails"])
                st.markdown(f"""
                <div class="metric-card" style="border-left:4px solid {ACCENT_GREEN};">
                    <div style="display:flex;align-items:center;gap:12px;">
                        <span style="font-size:2rem;">🔔</span>
                        <div>
                            <p style="color:{TEXT_PRIMARY};font-weight:600;margin:0;">Alertas a Admin: ON</p>
                            <p style="color:{TEXT_SECONDARY};font-size:0.85rem;margin:4px 0 0;">
                                Recibiran aviso si un envio automatico se aborta por fallo de Obuma:
                                <strong>{_emails_html}</strong>
                            </p>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                <div class="metric-card" style="border-left:4px solid {ACCENT_RED};">
                    <div style="display:flex;align-items:center;gap:12px;">
                        <span style="font-size:2rem;">🔕</span>
                        <div>
                            <p style="color:{TEXT_PRIMARY};font-weight:600;margin:0;">Alertas a Admin: OFF</p>
                            <p style="color:{TEXT_SECONDARY};font-size:0.85rem;margin:4px 0 0;">
                                {_admin_cfg_tab['reason']}. Si Obuma falla y un reporte automatico se aborta,
                                <strong>nadie recibira aviso por correo</strong> (solo quedara en logs).
                                Define <code>ADMIN_ALERT_EMAILS</code> (CSV de correos) en el entorno
                                para habilitar esta alerta.
                            </p>
                        </div>
                    </div>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("")

            if email_cfg.get("sandbox"):
                with st.expander("🔑 Pasos para activar envio a cualquier destinatario (Resend)", expanded=True):
                    st.markdown(f"""
                    <div class="data-card" style="border-left:4px solid {ACCENT_AMBER};">
                        <p style="color:{TEXT_PRIMARY};font-weight:700;margin:0 0 12px;font-size:1rem;">Como verificar el dominio vlsur.cl en Resend</p>
                        <ol style="color:{TEXT_SECONDARY};font-size:0.9rem;margin:0;padding-left:20px;line-height:1.8;">
                            <li>Inicia sesion en <a href="https://resend.com" target="_blank" style="color:{ACCENT_BLUE};">resend.com</a> con la cuenta <strong>reportesvlsurspa@gmail.com</strong></li>
                            <li>Ve a <strong>Domains</strong> → <strong>Add Domain</strong></li>
                            <li>Escribe <strong>vlsur.cl</strong> y haz clic en <strong>Add</strong></li>
                            <li>Resend te mostrara registros DNS (DKIM, SPF, DMARC) — agregalos en tu proveedor DNS</li>
                            <li>Espera 5-30 minutos y haz clic en <strong>Verify</strong></li>
                            <li>Una vez verificado, agrega la variable de entorno:<br>
                                <code style="background:#f5f5f5;padding:2px 6px;border-radius:3px;">EMAIL_FROM = reportes@vlsur.cl</code>
                            </li>
                            <li>Reinicia la aplicacion — los reportes comenzaran a llegar a <strong>gabrielhoyos@vlsur.cl</strong></li>
                        </ol>
                    </div>
                    """, unsafe_allow_html=True)

            with st.expander("📋 Variables de Entorno", expanded=not email_cfg["configured"]):
                st.markdown(f"""
                <div class="data-card">
                    <p style="color:{ACCENT_BLUE};font-weight:600;margin:0 0 8px;">Resend (activo) — para activar dominio propio</p>
                    <p style="color:{TEXT_SECONDARY};font-size:0.85rem;margin:0;">
                        <code>EMAIL_FROM</code> — Email remitente una vez verificado el dominio (ej: <code>reportes@vlsur.cl</code>)<br>
                        <code>EMAIL_FROM_NAME</code> — Nombre remitente (default: BI Platform - VLSur)
                    </p>
                </div>
                """, unsafe_allow_html=True)
                st.markdown("")
                st.markdown(f"""
                <div class="data-card">
                    <p style="color:{ACCENT_AMBER};font-weight:600;margin:0 0 8px;">Alternativa: SMTP (Gmail Workspace, etc.)</p>
                    <p style="color:{TEXT_SECONDARY};font-size:0.85rem;margin:0;">
                        <code>SMTP_HOST</code> — Servidor SMTP (ej: smtp.gmail.com)<br>
                        <code>SMTP_PORT</code> — Puerto (default: 587)<br>
                        <code>SMTP_USER</code> — Usuario SMTP<br>
                        <code>SMTP_PASS</code> — Contrasena SMTP
                    </p>
                </div>
                """, unsafe_allow_html=True)

            st.markdown("")
            st.markdown('<p class="section-header">Enviar Email de Prueba</p>', unsafe_allow_html=True)

            tc1, tc2 = st.columns([3, 1])
            with tc1:
                test_email = st.text_input("Email de prueba", value="gabrielhoyos@vlsur.cl", key="rpt_test_email")
            with tc2:
                st.markdown("")
                if st.button("Enviar Test", type="primary", use_container_width=True, key="rpt_test_send"):
                    if not test_email.strip():
                        st.error("Ingrese un email.")
                    elif not email_cfg["configured"]:
                        st.error("Configure el servicio de email primero.")
                    else:
                        with st.spinner("Enviando email de prueba..."):
                            result = test_email_delivery(test_email.strip())
                            if result.get("success"):
                                st.success(f"✅ Email de prueba enviado correctamente a {test_email}. Revisa tu bandeja de entrada.")
                            elif result.get("sandbox"):
                                st.warning(f"⚠️ Dominio sin verificar. {result.get('error', '')}")
                            else:
                                st.error(f"Error: {result.get('error', 'Error desconocido')}")
    finally:
        db.close()


# ============================================================
# SINCRONIZACION
# ============================================================
elif page == "Sincronizacion":
    st.markdown('<p class="page-title">Sincronizacion con Obuma</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Importe datos desde el ERP Obuma a la base de datos local</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        col1, col2 = st.columns(2)
        with col1:
            if st.button("Sincronizar Todo", type="primary", use_container_width=True):
                st.session_state.syncing_now = True
                st.session_state._sync_started = False
                st.rerun()

        with col2:
            all_modules = [
                "clientes", "clientes_contactos", "clientes_direcciones",
                "proveedores", "productos", "producto_categorias", "producto_subcategorias",
                "producto_fabricantes", "producto_precios", "empleados", "remuneraciones",
                "ventas", "ventas_items", "ventas_cotizaciones", "ventas_cobros", "ventas_dte",
                "compras", "compras_oc", "compras_pagos",
                "contabilidad", "crm_leads"
            ]
            endpoint = st.selectbox("Modulo especifico", all_modules)
            if st.button("Sincronizar Modulo", use_container_width=True):
                with st.spinner(f"Sincronizando {endpoint}..."):
                    service = SyncService(db)
                    method_name = f"sync_{endpoint}"
                    method = getattr(service, method_name, None)
                    if method:
                        result = run_async(method())
                        if isinstance(result, dict) and "error" not in result:
                            st.success(f"{endpoint.title()}: {result.get('synced', 0)} registros sincronizados")
                        else:
                            st.warning(f"Resultado: {result}")
                    else:
                        st.error(f"Metodo sync_{endpoint} no encontrado")

        st.markdown("---")
        st.markdown('<p class="section-header">Historial de Sincronizacion</p>', unsafe_allow_html=True)

        logs = db.query(SyncLog).order_by(SyncLog.ejecutado_at.desc()).limit(30).all()
        if logs:
            data = [{
                "Fecha": str(l.ejecutado_at)[:16],
                "Modulo": l.endpoint,
                "Registros API": l.registros_api,
                "Registros BD": l.registros_db,
                "Estado": l.estado.upper(),
            } for l in logs]
            st.dataframe(pd.DataFrame(data), hide_index=True)
        else:
            st.info("Sin historial de sincronizacion.")
    finally:
        db.close()


# ============================================================
# AUDITORIA
# ============================================================
elif page == "Auditoria":
    st.markdown('<p class="page-title">Auditoria de Datos</p>', unsafe_allow_html=True)
    st.markdown('<p class="page-subtitle">Diagnostico completo de datos disponibles en Obuma y base de datos local</p>', unsafe_allow_html=True)

    db = get_db()
    try:
        tab_diag, tab_tables, tab_logs = st.tabs(["Diagnostico API en Vivo", "Resumen de Tablas", "Log de Sincronizacion"])

        with tab_diag:
            st.markdown('<p class="section-header">Consultar API Obuma en Tiempo Real</p>', unsafe_allow_html=True)
            st.markdown(f'<p style="color:{TEXT_SECONDARY};font-size:0.9rem;">Este diagnostico consulta directamente la API de Obuma para verificar cuantos registros existen en cada endpoint. Permite identificar que datos estan disponibles para importar.</p>', unsafe_allow_html=True)

            if st.button("Ejecutar Diagnostico Completo", type="primary", use_container_width=True):
                with st.spinner("Consultando todos los endpoints de la API Obuma..."):
                    client = ObumaClient()
                    endpoints_to_test = [
                        ("Clientes", "clientes.list.json"),
                        ("Contactos Clientes", "clientesContactos.listAll.json"),
                        ("Direcciones Clientes", "clientesDirecciones.listAll.json"),
                        ("Proveedores", "proveedores.list.json"),
                        ("Productos", "productos.list.json"),
                        ("Categorias Productos", "productosCategorias.list.json"),
                        ("Subcategorias Productos", "productosSubCategorias.list.json"),
                        ("Fabricantes Productos", "productosFabricantes.list.json"),
                        ("Precios Productos", "productosConsultaPrecios.list.json"),
                        ("Empleados", "empleados.list.json"),
                        ("Remuneraciones", "remuneraciones.list.json"),
                        ("Ventas", "ventas.list.json"),
                        ("Items de Venta", "ventas.listItems.json"),
                        ("Cotizaciones", "ventasCotizaciones.list.json"),
                        ("Cobros", "ventasCobros.list.json"),
                        ("DTE Emitidos", "ventas.listDte.json"),
                        ("Compras", "compras.list.json"),
                        ("Ordenes de Compra", "comprasOc.list.json"),
                        ("Pagos Proveedores", "comprasPagos.list.json"),
                        ("Contabilidad", "contabilidad.listDiario.json"),
                        ("CRM Leads", "crm.list.json"),
                    ]

                    diag_results = []
                    for name, endpoint in endpoints_to_test:
                        try:
                            result = run_async(client.test_endpoint(endpoint))
                            if "error" in result:
                                status_code = result.get("status_code", "?")
                                diag_results.append({
                                    "Endpoint": name,
                                    "API Obuma": f"Error ({status_code})",
                                    "Estado": "NO DISPONIBLE",
                                    "Detalle": str(result.get("error", ""))[:80]
                                })
                            else:
                                total = result.get("data-total-items", 0)
                                data_list = result.get("data") or []
                                actual = len(data_list) if isinstance(data_list, list) else 0
                                estado = "CON DATOS" if total and int(total) > 0 else "SIN DATOS"
                                diag_results.append({
                                    "Endpoint": name,
                                    "API Obuma": f"{total} registros",
                                    "Estado": estado,
                                    "Detalle": f"{actual} en ultima pagina"
                                })
                        except Exception as e:
                            diag_results.append({
                                "Endpoint": name,
                                "API Obuma": "Error",
                                "Estado": "ERROR",
                                "Detalle": str(e)[:80]
                            })

                    df_diag = pd.DataFrame(diag_results)

                    con_datos = sum(1 for r in diag_results if r["Estado"] == "CON DATOS")
                    sin_datos = sum(1 for r in diag_results if r["Estado"] == "SIN DATOS")
                    errores = sum(1 for r in diag_results if r["Estado"] in ("ERROR", "NO DISPONIBLE"))

                    c1, c2, c3 = st.columns(3)
                    with c1:
                        render_metric("Con Datos", str(con_datos), "✅", ACCENT_GREEN)
                    with c2:
                        render_metric("Sin Datos", str(sin_datos), "⚠️", ACCENT_AMBER)
                    with c3:
                        render_metric("No Disponible", str(errores), "❌", ACCENT_RED)

                    st.markdown("")

                    def color_estado(val):
                        if val == "CON DATOS":
                            return f"background-color: rgba(16,185,129,0.15); color: #34d399;"
                        elif val == "SIN DATOS":
                            return f"background-color: rgba(245,158,11,0.15); color: #fbbf24;"
                        else:
                            return f"background-color: rgba(239,68,68,0.15); color: #f87171;"

                    styled_df = df_diag.style.map(color_estado, subset=["Estado"])
                    st.dataframe(styled_df, hide_index=True, height=600)

                    if sin_datos > 10:
                        st.warning(f"Su cuenta Obuma tiene datos solo en {con_datos} de {len(diag_results)} endpoints. Los endpoints sin datos ({sin_datos}) no tienen registros cargados en su ERP Obuma. Para que aparezcan aqui, primero debe cargar esos datos en Obuma (productos, compras, contabilidad, etc).")

        with tab_tables:
            st.markdown('<p class="section-header">Registros en Base de Datos Local</p>', unsafe_allow_html=True)
            table_counts = [
                ("Clientes", db.query(ClienteFinal).count(), "clientes.list.json"),
                ("Contactos", db.query(ClienteContacto).count(), "clientesContactos.listAll.json"),
                ("Direcciones", db.query(ClienteDireccion).count(), "clientesDirecciones.listAll.json"),
                ("Proveedores", db.query(Proveedor).count(), "proveedores.list.json"),
                ("Productos", db.query(Producto).count(), "productos.list.json"),
                ("Categorias", db.query(ProductoCategoria).count(), "productosCategorias.list.json"),
                ("Subcategorias", db.query(ProductoSubCategoria).count(), "productosSubCategorias.list.json"),
                ("Fabricantes", db.query(ProductoFabricante).count(), "productosFabricantes.list.json"),
                ("Precios", db.query(ProductoPrecio).count(), "productosConsultaPrecios.list.json"),
                ("Empleados", db.query(Empleado).count(), "empleados.list.json"),
                ("Remuneraciones", db.query(Remuneracion).count(), "remuneraciones.list.json"),
                ("Ventas", db.query(VentaHistorico).count(), "ventas.list.json"),
                ("Items Venta", db.query(VentaItem).count(), "ventas.listItems.json"),
                ("Cotizaciones", db.query(VentaCotizacion).count(), "ventasCotizaciones.list.json"),
                ("Cobros", db.query(VentaCobro).count(), "ventasCobros.list.json"),
                ("DTE Emitidos", db.query(VentaDte).count(), "ventas.listDte.json"),
                ("Compras", db.query(CompraHistorico).count(), "compras.list.json"),
                ("Ordenes Compra", db.query(CompraOC).count(), "comprasOc.list.json"),
                ("Pagos Proveedores", db.query(CompraPago).count(), "comprasPagos.list.json"),
                ("Contabilidad", db.query(ContabilidadHistorico).count(), "contabilidad.listDiario.json"),
                ("CRM Leads", db.query(CrmLead).count(), "crm.list.json"),
                ("Costos Historicos", db.query(CostoHistorico).count(), "-"),
            ]

            total_registros = sum(t[1] for t in table_counts)
            tablas_con_datos = sum(1 for t in table_counts if t[1] > 0)

            c1, c2 = st.columns(2)
            with c1:
                render_metric("Total Registros BD", str(total_registros), "🗄️", ACCENT_BLUE)
            with c2:
                render_metric("Tablas con Datos", f"{tablas_con_datos}/{len(table_counts)}", "📊", ACCENT_GREEN)

            st.markdown("")
            df_tables = pd.DataFrame(table_counts, columns=["Tabla", "Registros", "Endpoint API"])

            def color_registros(val):
                if isinstance(val, (int, float)) and val > 0:
                    return "color: #34d399; font-weight: bold;"
                elif isinstance(val, (int, float)):
                    return "color: #64748b;"
                return ""

            styled = df_tables.style.map(color_registros, subset=["Registros"])
            st.dataframe(styled, hide_index=True, height=700)

        with tab_logs:
            st.markdown('<p class="section-header">Historial de Sincronizacion</p>', unsafe_allow_html=True)
            logs = db.query(SyncLog).order_by(SyncLog.ejecutado_at.desc()).limit(30).all()
            if logs:
                data = [{
                    "Fecha": str(l.ejecutado_at)[:16],
                    "Modulo": l.endpoint,
                    "Registros API": l.registros_api,
                    "Registros BD": l.registros_db,
                    "Discrepancias": l.discrepancias,
                    "Estado": l.estado.upper(),
                } for l in logs]
                st.dataframe(pd.DataFrame(data), hide_index=True, height=600)
            else:
                st.info("Sin historial de sincronizacion.")
    finally:
        db.close()
