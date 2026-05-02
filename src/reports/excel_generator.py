import os
import json
import logging
from datetime import datetime, date
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from sqlalchemy import func, extract, distinct, case as sql_case
from src.models.models import (
    VentaHistorico,
    ClienteFinal,
    Empleado,
    ReporteGenerado,
    VendedorCartera,
)
from src.utils.date_filters import date_range_filters

logger = logging.getLogger(__name__)

HEADER_FONT = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
CURRENCY_FORMAT = "#,##0"
PERCENT_FORMAT = "0.00%"
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="2F5496")
SUBTITLE_FONT = Font(name="Calibri", bold=True, size=11, color="404040")

YELLOW_FILL = PatternFill(
    start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid"
)
GREEN_FILL = PatternFill(
    start_color="FF92D050", end_color="FF92D050", fill_type="solid"
)
RED_FILL = PatternFill(start_color="FFFF6B6B", end_color="FFFF6B6B", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(
    start_color="FFE8F5E9", end_color="FFE8F5E9", fill_type="solid"
)
LIGHT_RED_FILL = PatternFill(
    start_color="FFFFEBEE", end_color="FFFFEBEE", fill_type="solid"
)
SEGMENTO_FONT = Font(name="Calibri", bold=True, size=11)

MONTH_NAMES = [
    "Ene",
    "Feb",
    "Mar",
    "Abr",
    "May",
    "Jun",
    "Jul",
    "Ago",
    "Sep",
    "Oct",
    "Nov",
    "Dic",
]

BILLING_DOC_TYPES = ["Factura Electr.", "Factura Exenta", "Boleta Electr."]
NC_DOC_TYPES = ["Nota Credito"]
# Notas de Debito: cargos adicionales al cliente (intereses, recargos, ajustes
# a favor de la empresa). En Obuma "Facturas por Cobrar" aparecen como deuda
# pendiente igual que las facturas. Tratamiento:
#  - Cartera/Cobranza: SUMAN positivo en POR PAGAR (igual que Facturas).
#  - Ventas/Margen/Dashboard: SUMAN positivo (representan ingreso adicional).
# (Diferente de NC, que en ventas RESTA y en cartera tambien se muestra positiva
# pero por otra razon: el saldo a favor del cliente.)
ND_DOC_TYPES = ["Nota Debito"]
VALID_DOC_TYPES = BILLING_DOC_TYPES + NC_DOC_TYPES + ND_DOC_TYPES

# Totales de referencia capturados manualmente desde la pantalla
# "Facturas por Cobrar" de Obuma para cada vendedor trackeado.
# Se usan para reconciliar el reporte de cartera contra Obuma y loguear
# el % de diferencia (objetivo: <0.5%). Si el valor es None significa que
# no hay snapshot reciente; el log mostrara solo el total interno.
# Para actualizar: ir a Obuma > Facturas por Cobrar, filtrar por vendedor,
# copiar el total y la fecha de captura aqui (mantener la fecha cerca del
# valor para saber si esta vigente).
OBUMA_REFERENCE_TOTALS = {
    "28856": 42_502_743,  # Gabriel  (capturado 2026-04-27)
    "28886": None,        # Jhonatan (pendiente de captura)
    "28887": None,        # Ernesto  (pendiente de captura)
    "28891": None,        # Pablo    (pendiente de captura)
    "28892": None,        # Jesus    (pendiente de captura)
}


def _style_header(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_length + 4, 40)


def _get_last_3_months(reference_date):
    months = []
    year = reference_date.year
    month = reference_date.month
    for _ in range(3):
        month -= 1
        if month <= 0:
            month = 12
            year -= 1
        months.append(month)
    return months


def _classify_segmento(acumulado_pct):
    if acumulado_pct <= 0.80:
        return "A"
    elif acumulado_pct <= 0.95:
        return "B"
    elif acumulado_pct <= 0.99:
        return "C"
    else:
        return "D"


def _classify_riesgo(meses_con_venta, ventas_ultimos_3):
    if meses_con_venta >= 8:
        return "BAJO"
    if meses_con_venta >= 6 and ventas_ultimos_3 > 0:
        return "BAJO"
    if meses_con_venta >= 3 and ventas_ultimos_3 > 0:
        return "MEDIO"
    return "ALTO"


def _resolve_date_range(date_from, date_to):
    if date_from is None and date_to is None:
        current_year = date.today().year
        return date(current_year, 1, 1), date(current_year, 12, 31), False
    if date_from is None:
        date_from = date(date_to.year, 1, 1)
    if date_to is None:
        date_to = date.today()
    return date_from, date_to, True


def _parse_detalle_json(venta):
    """Parsea el campo VentaHistorico.detalle (text JSON) de forma segura.
    Retorna {} si está vacío o no es JSON válido."""
    if not venta or not getattr(venta, "detalle", None):
        return {}
    try:
        return json.loads(venta.detalle) or {}
    except (json.JSONDecodeError, TypeError, ValueError):
        return {}


def _get_emision_date(venta, detalle_json=None):
    """Obtiene la fecha de emisión de una VentaHistorico con COALESCE:
    1) Usa la columna VentaHistorico.fecha si está presente.
    2) Si está NULL, reconstruye la fecha desde el JSON detalle
       usando venta_dia/venta_mes/venta_ano.
    3) Si no se puede reconstruir, retorna None.
    """
    fecha_db = getattr(venta, "fecha", None) if venta else None
    if fecha_db is not None:
        if isinstance(fecha_db, datetime):
            return fecha_db.date()
        return fecha_db

    if detalle_json is None:
        detalle_json = _parse_detalle_json(venta)

    try:
        dia_raw = detalle_json.get("venta_dia")
        mes_raw = detalle_json.get("venta_mes")
        ano_raw = detalle_json.get("venta_ano")
        if not dia_raw or not mes_raw or not ano_raw:
            return None
        dia = int(str(dia_raw).strip())
        mes = int(str(mes_raw).strip())
        ano = int(str(ano_raw).strip())
        if dia <= 0 or mes <= 0 or ano <= 0:
            return None
        return date(ano, mes, dia)
    except (ValueError, TypeError, AttributeError):
        return None


def _get_vencimiento_date(venta, detalle_json=None):
    """Obtiene la fecha de vencimiento desde detalle.venta_fecha_vencimiento.
    Maneja '0000-00-00' y cadenas vacías como None.
    Acepta formato 'YYYY-MM-DD' o 'YYYY-MM-DD HH:MM:SS'.
    """
    if detalle_json is None:
        detalle_json = _parse_detalle_json(venta)

    raw = detalle_json.get("venta_fecha_vencimiento")
    if raw is None:
        return None
    raw_str = str(raw).strip()
    if not raw_str or raw_str.startswith("0000-00-00") or raw_str == "0":
        return None
    try:
        if " " in raw_str:
            raw_str = raw_str.split(" ")[0]
        parts = raw_str.split("-")
        if len(parts) != 3:
            return None
        ano, mes, dia = int(parts[0]), int(parts[1]), int(parts[2])
        if ano <= 0 or mes <= 0 or dia <= 0:
            return None
        return date(ano, mes, dia)
    except (ValueError, TypeError):
        return None


def _build_cartera_sheet(wb, db, vendedor_obuma_id, empleado, date_from, date_to):
    cartera_entries = (
        db.query(VendedorCartera, ClienteFinal)
        .join(ClienteFinal, VendedorCartera.cliente_id == ClienteFinal.id)
        .filter(
            VendedorCartera.empleado_obuma_id == vendedor_obuma_id,
            VendedorCartera.activo == True,
        )
        .all()
    )

    if not cartera_entries:
        return

    compraron = []
    no_compraron = []

    for vc, cli in cartera_entries:
        ventas_result = (
            db.query(
                func.sum(
                    sql_case(
                        (
                            VentaHistorico.tipo_documento.in_(NC_DOC_TYPES),
                            -VentaHistorico.subtotal,
                        ),
                        else_=VentaHistorico.subtotal,
                    )
                ).label("total_ventas"),
                func.count(VentaHistorico.id).label("num_docs"),
            )
            .filter(
                VentaHistorico.cliente_id == cli.id,
                VentaHistorico.vendedor_id == vendedor_obuma_id,
                VentaHistorico.anulada == False,
                *date_range_filters(VentaHistorico.fecha, date_from, date_to),
                VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES),
            )
            .first()
        )

        total_ventas = ventas_result.total_ventas or 0
        num_docs = ventas_result.num_docs or 0

        # Regla: si neto <= 0 (NCs >= facturas), NO cuenta como compra real.
        # Va a "no_compraron" para no inflar la cobertura.
        if num_docs > 0 and total_ventas > 0:
            compraron.append(
                {
                    "nombre": cli.nombre or "",
                    "rut": cli.rut or "",
                    "total_ventas": total_ventas,
                    "num_docs": num_docs,
                }
            )
        else:
            ultima_compra = (
                db.query(func.max(VentaHistorico.fecha))
                .filter(
                    VentaHistorico.cliente_id == cli.id,
                    VentaHistorico.vendedor_id == vendedor_obuma_id,
                    VentaHistorico.anulada == False,
                )
                .scalar()
            )
            dias_sin = None
            ultima_str = "Sin registro"
            if ultima_compra:
                if hasattr(ultima_compra, "date"):
                    ultima_date = ultima_compra.date()
                else:
                    ultima_date = ultima_compra
                dias_sin = (date.today() - ultima_date).days
                ultima_str = str(ultima_date)

            no_compraron.append(
                {
                    "nombre": cli.nombre or "",
                    "rut": cli.rut or "",
                    "ultima_compra": ultima_str,
                    "dias_sin_comprar": dias_sin,
                }
            )

    compraron.sort(key=lambda x: x["total_ventas"], reverse=True)
    no_compraron.sort(
        key=lambda x: (x["dias_sin_comprar"] is None, -(x["dias_sin_comprar"] or 0))
    )

    total_cartera = len(cartera_entries)
    total_compraron = len(compraron)
    total_no_compraron = len(no_compraron)
    cobertura = (total_compraron / total_cartera * 100) if total_cartera > 0 else 0

    GREEN_HEADER = PatternFill(
        start_color="10B981", end_color="10B981", fill_type="solid"
    )
    RED_HEADER = PatternFill(
        start_color="EF4444", end_color="EF4444", fill_type="solid"
    )
    BLUE_HEADER = PatternFill(
        start_color="3B82F6", end_color="3B82F6", fill_type="solid"
    )

    ws = wb.create_sheet("Cruce Cartera vs Ventas")

    ws.cell(
        row=1, column=1, value=f"Cruce Cartera vs Ventas - {empleado.nombre}"
    ).font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws.cell(
        row=2,
        column=1,
        value=f"Periodo: {date_from.strftime('%d/%m/%Y')} - {date_to.strftime('%d/%m/%Y')}",
    ).font = SUBTITLE_FONT
    ws.merge_cells("A2:F2")

    ws.cell(row=4, column=1, value="Total Cartera")
    ws.cell(row=4, column=2, value=total_cartera)
    ws.cell(row=5, column=1, value="Compraron")
    ws.cell(row=5, column=2, value=total_compraron)
    ws.cell(row=6, column=1, value="No Compraron")
    ws.cell(row=6, column=2, value=total_no_compraron)
    ws.cell(row=7, column=1, value="Cobertura")
    ws.cell(row=7, column=2, value=f"{cobertura:.1f}%")
    for r in range(4, 8):
        ws.cell(row=r, column=1).font = Font(bold=True)
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2).border = THIN_BORDER
    ws.cell(row=5, column=2).fill = LIGHT_GREEN_FILL
    ws.cell(row=6, column=2).fill = LIGHT_RED_FILL

    row = 9
    ws.cell(
        row=row, column=1, value=f"CLIENTES QUE COMPRARON ({total_compraron})"
    ).font = Font(bold=True, size=12, color="10B981")
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    comp_headers = ["N°", "RUT", "Razon Social", "Total Ventas", "Documentos"]
    for i, h in enumerate(comp_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = GREEN_HEADER
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    row += 1

    for idx, c in enumerate(compraron, 1):
        ws.cell(row=row, column=1, value=idx).border = THIN_BORDER
        ws.cell(row=row, column=2, value=c["rut"]).border = THIN_BORDER
        ws.cell(row=row, column=3, value=c["nombre"]).border = THIN_BORDER
        cell = ws.cell(row=row, column=4, value=c["total_ventas"])
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER
        ws.cell(row=row, column=5, value=c["num_docs"]).border = THIN_BORDER
        row_fill = LIGHT_RED_FILL if c["total_ventas"] < 0 else LIGHT_GREEN_FILL
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = row_fill
        if c["total_ventas"] < 0:
            cell.font = Font(name="Calibri", size=11, color="C00000", bold=True)
        row += 1

    row += 1
    ws.cell(
        row=row, column=1, value=f"CLIENTES QUE NO COMPRARON ({total_no_compraron})"
    ).font = Font(bold=True, size=12, color="EF4444")
    ws.merge_cells(f"A{row}:D{row}")
    row += 1

    no_headers = ["N°", "RUT", "Razon Social", "Ultima Compra", "Dias sin Comprar"]
    for i, h in enumerate(no_headers, 1):
        cell = ws.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = RED_HEADER
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER
    row += 1

    for idx, c in enumerate(no_compraron, 1):
        ws.cell(row=row, column=1, value=idx).border = THIN_BORDER
        ws.cell(row=row, column=2, value=c["rut"]).border = THIN_BORDER
        ws.cell(row=row, column=3, value=c["nombre"]).border = THIN_BORDER
        ws.cell(row=row, column=4, value=c["ultima_compra"]).border = THIN_BORDER
        dias = c["dias_sin_comprar"]
        cell_dias = ws.cell(
            row=row, column=5, value=dias if dias is not None else "Sin registro"
        )
        cell_dias.border = THIN_BORDER
        for col in range(1, 6):
            ws.cell(row=row, column=col).fill = LIGHT_RED_FILL
        row += 1

    _auto_width(ws)

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18


def generate_vendedor_report(
    db: Session, vendedor_obuma_id: str, date_from: date = None, date_to: date = None
) -> str:
    date_from, date_to, custom_range = _resolve_date_range(date_from, date_to)

    reference_date = date_to if date_to else date.today()
    last_3_months = _get_last_3_months(reference_date)

    empleado = db.query(Empleado).filter(Empleado.obuma_id == vendedor_obuma_id).first()
    if not empleado:
        logger.warning(f"Empleado not found for obuma_id: {vendedor_obuma_id}")
        return None

    ventas = (
        db.query(VentaHistorico)
        .filter(
            VentaHistorico.vendedor_id == vendedor_obuma_id,
            VentaHistorico.anulada != True,
            *date_range_filters(VentaHistorico.fecha, date_from, date_to),
            VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES),
        )
        .all()
    )

    cartera_clients = (
        db.query(VendedorCartera, ClienteFinal)
        .join(ClienteFinal, VendedorCartera.cliente_id == ClienteFinal.id)
        .filter(
            VendedorCartera.empleado_obuma_id == vendedor_obuma_id,
            VendedorCartera.activo == True,
        )
        .all()
    )

    if not ventas and not cartera_clients:
        logger.info(
            f"No sales or cartera found for vendedor {empleado.nombre} ({vendedor_obuma_id}) in range {date_from} to {date_to}"
        )
        return None

    active_months = set()
    for m in range(1, 13):
        month_start = date(date_from.year, m, 1)
        if m == 12:
            month_end = date(date_from.year, 12, 31)
        else:
            month_end = date(date_from.year, m + 1, 1)
        if month_start <= date_to and month_end > date_from:
            active_months.add(m)

    client_data = defaultdict(lambda: defaultdict(float))
    client_info = {}

    for vc, cli in cartera_clients:
        ckey = f"db_{cli.id}"
        if ckey not in client_info:
            client_info[ckey] = {"rut": cli.rut or "", "nombre": cli.nombre or ""}

    cartera_db_ids = (
        set(cli.id for _, cli in cartera_clients) if cartera_clients else set()
    )

    excluded_db_ids = set()
    inactive_entries = (
        db.query(VendedorCartera)
        .filter(
            VendedorCartera.empleado_obuma_id == vendedor_obuma_id,
            VendedorCartera.activo == False,
        )
        .all()
    )
    for ie in inactive_entries:
        excluded_db_ids.add(ie.cliente_id)

    reassigned_entries = (
        db.query(VendedorCartera)
        .filter(
            VendedorCartera.empleado_obuma_id != vendedor_obuma_id,
            VendedorCartera.activo == True,
        )
        .all()
    )
    reassigned_db_ids = set(re.cliente_id for re in reassigned_entries)

    def _parse_data_json(raw):
        if isinstance(raw, dict):
            return raw
        if isinstance(raw, str):
            try:
                return json.loads(raw)
            except (json.JSONDecodeError, TypeError):
                pass
        return {}

    for v in ventas:
        ckey = None
        if v.cliente_id:
            if v.cliente_id in excluded_db_ids:
                continue
            if v.cliente_id in reassigned_db_ids:
                continue
            if v.cliente_id not in cartera_db_ids:
                cliente = (
                    db.query(ClienteFinal)
                    .filter(ClienteFinal.id == v.cliente_id)
                    .first()
                )
                if cliente:
                    dj = _parse_data_json(cliente.data_json)
                    cli_vendedor = str(dj.get("rel_usuario_id", ""))
                    if cli_vendedor != str(vendedor_obuma_id):
                        continue
                    cartera_db_ids.add(v.cliente_id)
            ckey = f"db_{v.cliente_id}"
            if ckey not in client_info:
                cliente = (
                    db.query(ClienteFinal)
                    .filter(ClienteFinal.id == v.cliente_id)
                    .first()
                )
                if cliente:
                    client_info[ckey] = {
                        "rut": cliente.rut or "",
                        "nombre": cliente.nombre or "",
                    }
                else:
                    client_info[ckey] = {"rut": "", "nombre": ""}
        else:
            try:
                detalle = json.loads(v.detalle) if v.detalle else {}
            except (json.JSONDecodeError, TypeError):
                detalle = {}
            rel_id = str(detalle.get("rel_cliente_id", "0"))
            if rel_id and rel_id != "0":
                ckey = f"obuma_{rel_id}"
                if ckey not in client_info:
                    cliente = (
                        db.query(ClienteFinal)
                        .filter(ClienteFinal.obuma_id == rel_id)
                        .first()
                    )
                    if cliente:
                        if cliente.id in excluded_db_ids or cliente.id in reassigned_db_ids:
                            continue
                        dj = _parse_data_json(cliente.data_json)
                        cli_vendedor = str(dj.get("rel_usuario_id", ""))
                        if cli_vendedor != str(vendedor_obuma_id):
                            continue
                        client_info[ckey] = {
                            "rut": cliente.rut or "",
                            "nombre": cliente.nombre or "",
                        }
                    else:
                        client_info[ckey] = {
                            "rut": f"ID-{rel_id}",
                            "nombre": detalle.get(
                                "cliente_razon_social", f"Cliente {rel_id}"
                            ),
                        }
            else:
                continue

        if ckey and v.fecha:
            month = v.fecha.month
            amount = v.subtotal or 0
            if v.tipo_documento in NC_DOC_TYPES:
                client_data[ckey][month] -= amount
            else:
                client_data[ckey][month] += amount

    all_client_keys = set(client_info.keys()) | set(client_data.keys())

    rows = []
    for cid in all_client_keys:
        monthly = client_data.get(cid, defaultdict(float))
        info = client_info.get(cid, {"rut": "", "nombre": ""})
        rut_val = (info.get("rut") or "").strip()
        nombre_val = (info.get("nombre") or "").strip()
        if rut_val.startswith("OBU-") and not nombre_val:
            continue
        # Regla del cliente:
        # - Mostrar valor real (incluso negativo) y Total con NCs restadas (monto correcto).
        # - Para CONTEO de cobertura ("meses con venta"), un negativo cuenta como cero
        #   (no genero compra ese mes).
        month_values = [
            monthly.get(m, 0) if m in active_months else 0 for m in range(1, 13)
        ]
        total = sum(month_values)
        meses_con_venta = sum(
            1 for i, v in enumerate(month_values) if v > 0 and (i + 1) in active_months
        )
        ventas_ultimos_3 = sum(monthly.get(m, 0) for m in last_3_months)

        rows.append(
            {
                "rut": info["rut"],
                "nombre": info["nombre"],
                "months": month_values,
                "total": total,
                "meses_con_venta": meses_con_venta,
                "ventas_ultimos_3": ventas_ultimos_3,
            }
        )

    rows.sort(key=lambda r: r["total"], reverse=True)

    grand_total = sum(r["total"] for r in rows)

    acumulado = 0
    for r in rows:
        pct_venta = r["total"] / grand_total if grand_total > 0 else 0
        acumulado += pct_venta
        r["pct_venta"] = pct_venta
        r["pct_acumulado"] = acumulado
        r["segmento"] = _classify_segmento(acumulado)
        r["nivel_riesgo"] = _classify_riesgo(
            r["meses_con_venta"], r["ventas_ultimos_3"]
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte-Ventas-EvolucionPorClie"

    headers = [
        "N°",
        "CLIENTE Rut",
        "CLIENTE Razon Social",
        "CLIENTE Tipo",
        "Ene",
        "Feb",
        "Mar",
        "Abr",
        "May",
        "Jun",
        "Jul",
        "Ago",
        "Sep",
        "Oct",
        "Nov",
        "Dic",
        "TOTAL",
        "% Acumulado",
        "Segmento",
        "% de la venta",
        "Meses con venta",
        "Ventas últimos 3 meses",
        "Nivel de Riesgo",
    ]

    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    _style_header(ws, 1, len(headers))

    for idx, r in enumerate(rows, 1):
        row_num = idx + 1
        ws.cell(row=row_num, column=1, value=idx).border = THIN_BORDER
        ws.cell(row=row_num, column=2, value=r["rut"]).border = THIN_BORDER
        ws.cell(row=row_num, column=3, value=r["nombre"]).border = THIN_BORDER
        ws.cell(row=row_num, column=4, value="Distribuidor").border = THIN_BORDER

        for m in range(12):
            val = r["months"][m]
            cell = ws.cell(row=row_num, column=5 + m, value=val)
            cell.number_format = CURRENCY_FORMAT
            cell.border = THIN_BORDER
            if val < 0:
                cell.fill = LIGHT_RED_FILL
                cell.font = Font(name="Calibri", size=11, color="C00000", bold=True)
            elif val == 0:
                cell.fill = YELLOW_FILL

        total_val = r["total"]
        cell = ws.cell(row=row_num, column=17, value=total_val)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER
        if total_val < 0:
            cell.fill = LIGHT_RED_FILL
            cell.font = Font(name="Calibri", size=11, color="C00000", bold=True)

        cell = ws.cell(row=row_num, column=18, value=r["pct_acumulado"])
        cell.number_format = PERCENT_FORMAT
        cell.border = THIN_BORDER

        cell = ws.cell(row=row_num, column=19, value=r["segmento"])
        cell.fill = GREEN_FILL
        cell.font = SEGMENTO_FONT
        cell.border = THIN_BORDER

        cell = ws.cell(row=row_num, column=20, value=r["pct_venta"])
        cell.number_format = PERCENT_FORMAT
        cell.border = THIN_BORDER

        ws.cell(row=row_num, column=21, value=r["meses_con_venta"]).border = THIN_BORDER

        ult3_val = r["ventas_ultimos_3"]
        cell = ws.cell(row=row_num, column=22, value=ult3_val)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER
        if ult3_val < 0:
            cell.fill = LIGHT_RED_FILL
            cell.font = Font(name="Calibri", size=11, color="C00000", bold=True)
        elif ult3_val == 0:
            cell.fill = YELLOW_FILL

        riesgo_val = r["nivel_riesgo"]
        riesgo_cell = ws.cell(row=row_num, column=23, value=riesgo_val)
        riesgo_cell.border = THIN_BORDER
        if riesgo_val == "ALTO":
            riesgo_cell.fill = RED_FILL
            riesgo_cell.font = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
        elif riesgo_val == "MEDIO":
            riesgo_cell.fill = PatternFill(
                start_color="FFF3CD", end_color="FFF3CD", fill_type="solid"
            )
            riesgo_cell.font = Font(name="Calibri", size=11, color="856404", bold=True)
        elif riesgo_val == "BAJO":
            riesgo_cell.fill = LIGHT_GREEN_FILL
            riesgo_cell.font = Font(name="Calibri", size=11, color="155724", bold=True)

    _auto_width(ws)

    _build_cartera_sheet(wb, db, vendedor_obuma_id, empleado, date_from, date_to)

    os.makedirs("reports", exist_ok=True)
    vendedor_rut = (empleado.rut or vendedor_obuma_id).replace(".", "").replace("-", "")
    if custom_range:
        filename = f"vendedor_{vendedor_rut}_{date_from.strftime('%Y%m%d')}_{date_to.strftime('%Y%m%d')}.xlsx"
    else:
        filename = f"vendedor_{vendedor_rut}_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join("reports", filename)
    wb.save(filepath)

    reporte = ReporteGenerado(
        nombre_archivo=filename,
        tipo="vendedor",
        fecha_reporte=date.today(),
        ruta_archivo=filepath,
    )
    db.add(reporte)
    db.commit()

    logger.info(f"Vendedor report generated: {filepath} ({empleado.nombre})")
    return filepath


def generate_all_vendedor_reports(
    db: Session, date_from: date = None, date_to: date = None
) -> list:
    resolved_from, resolved_to, _ = _resolve_date_range(date_from, date_to)

    vendedor_ids = (
        db.query(distinct(VentaHistorico.vendedor_id))
        .filter(
            VentaHistorico.vendedor_id != None,
            VentaHistorico.anulada != True,
            VentaHistorico.fecha >= resolved_from,
            VentaHistorico.fecha <= resolved_to,
        )
        .all()
    )

    vendedor_ids = [vid[0] for vid in vendedor_ids if vid[0]]

    filepaths = []
    for vid in vendedor_ids:
        try:
            fp = generate_vendedor_report(db, vid, date_from, date_to)
            if fp:
                filepaths.append(fp)
        except Exception as e:
            logger.error(f"Error generating report for vendedor {vid}: {e}")

    logger.info(
        f"Generated {len(filepaths)} vendedor reports for range {resolved_from} to {resolved_to}"
    )
    return filepaths


def generate_daily_report(db: Session, report_date: date = None) -> str:
    if report_date is None:
        report_date = date.today()

    current_year = report_date.year
    date_from = date(current_year, 1, 1)
    date_to = date(current_year, 12, 31)
    filepaths = generate_all_vendedor_reports(db, date_from, date_to)

    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Consolidado"

    ws.cell(
        row=1, column=1, value="Reporte Consolidado de Vendedores"
    ).font = TITLE_FONT
    ws.cell(
        row=2,
        column=1,
        value=f"Fecha: {report_date.strftime('%d/%m/%Y')} - Año: {current_year}",
    ).font = SUBTITLE_FONT
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    headers = ["Vendedor", "RUT", "Clientes", "Venta Total Neto"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=4, column=i, value=h)
    _style_header(ws, 4, len(headers))

    vendedor_ids = (
        db.query(distinct(VentaHistorico.vendedor_id))
        .filter(
            VentaHistorico.vendedor_id != None,
            VentaHistorico.anulada != True,
            *date_range_filters(VentaHistorico.fecha, date_from, date_to),
        )
        .all()
    )
    vendedor_ids = [vid[0] for vid in vendedor_ids if vid[0]]

    row = 5
    total_general = 0
    for vid in vendedor_ids:
        empleado = db.query(Empleado).filter(Empleado.obuma_id == vid).first()
        if not empleado:
            continue

        stats = (
            db.query(
                func.count(distinct(VentaHistorico.cliente_id)),
                func.sum(
                    sql_case(
                        (
                            VentaHistorico.tipo_documento.in_(NC_DOC_TYPES),
                            -VentaHistorico.subtotal,
                        ),
                        else_=VentaHistorico.subtotal,
                    )
                ),
            )
            .filter(
                VentaHistorico.vendedor_id == vid,
                VentaHistorico.anulada != True,
                *date_range_filters(VentaHistorico.fecha, date_from, date_to),
                VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES),
            )
            .first()
        )

        num_clientes = stats[0] or 0
        total_neto = stats[1] or 0

        ws.cell(row=row, column=1, value=empleado.nombre or "").border = THIN_BORDER
        ws.cell(row=row, column=2, value=empleado.rut or "").border = THIN_BORDER
        ws.cell(row=row, column=3, value=num_clientes).border = THIN_BORDER
        cell = ws.cell(row=row, column=4, value=total_neto)
        cell.number_format = CURRENCY_FORMAT
        cell.border = THIN_BORDER

        total_general += total_neto
        row += 1

    row += 1
    ws.cell(row=row, column=3, value="TOTAL GENERAL").font = Font(bold=True)
    ws.cell(row=row, column=3).border = THIN_BORDER
    cell = ws.cell(row=row, column=4, value=total_general)
    cell.number_format = CURRENCY_FORMAT
    cell.font = Font(bold=True)
    cell.border = THIN_BORDER

    _auto_width(ws)

    os.makedirs("reports", exist_ok=True)
    filename = f"reporte_diario_{report_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join("reports", filename)
    wb.save(filepath)

    reporte = ReporteGenerado(
        nombre_archivo=filename,
        tipo="diario",
        fecha_reporte=report_date,
        ruta_archivo=filepath,
    )
    db.add(reporte)
    db.commit()

    logger.info(
        f"Daily consolidated report generated: {filepath} with {len(filepaths)} vendedor reports"
    )
    return filepath


# ============================================================================
# REPORTE CARTERA / COBRANZA (días de atraso desde fecha de emisión)
# ============================================================================

COBRANZA_HEADERS = [
    "DOCUMENTO", "FOLIO", "FECHA", "FECHA VCTO", "ESTADO", "FECHA HOY",
    "DÍAS ATRASO", "CLIENTE", "CLIENTE RUT", "VENDEDOR", "POR PAGAR",
]

COBRANZA_VERDE = PatternFill(start_color="FFC6EFCE", end_color="FFC6EFCE", fill_type="solid")
COBRANZA_NARANJA = PatternFill(start_color="FFFFD966", end_color="FFFFD966", fill_type="solid")
COBRANZA_ROJO = PatternFill(start_color="FFF4B7B7", end_color="FFF4B7B7", fill_type="solid")
COBRANZA_GRIS_SUBTOTAL = PatternFill(start_color="FFD9D9D9", end_color="FFD9D9D9", fill_type="solid")
COBRANZA_GRIS_TOTAL = PatternFill(start_color="FFA6A6A6", end_color="FFA6A6A6", fill_type="solid")

RESUMEN_TITLE_FILL = PatternFill(start_color="FF1F4E79", end_color="FF1F4E79", fill_type="solid")
RESUMEN_TITLE_FONT = Font(name="Calibri", bold=True, size=12, color="FFFFFFFF")
RESUMEN_LABEL_FONT = Font(name="Calibri", size=11, color="FF404040")
RESUMEN_VALUE_FONT = Font(name="Calibri", bold=True, size=11, color="FF1F4E79")
RESUMEN_TOTAL_FONT = Font(name="Calibri", bold=True, size=13, color="FF1F4E79")
RESUMEN_FILL = PatternFill(start_color="FFEAF1F8", end_color="FFEAF1F8", fill_type="solid")

NC_FONT = Font(name="Calibri", size=11, color="FFC00000", italic=True)
# ND (Nota de Debito): azul oscuro, sin italic, para distinguir de NC (rojo italic)
# y Facturas (negro). Suma positivo en POR PAGAR igual que las facturas.
ND_FONT = Font(name="Calibri", size=11, color="FF1F4E79", bold=True)
ESTADO_VENCIDO_FONT = Font(name="Calibri", size=11, color="FFC00000", bold=True)
ESTADO_PORVENCER_FONT = Font(name="Calibri", size=11, color="FF548235", bold=True)
ESTADO_SINVCTO_FONT = Font(name="Calibri", size=11, color="FF7F7F7F", italic=True)
SIN_VCTO_FONT = Font(name="Calibri", size=11, color="FF7F7F7F", italic=True)
SUBTOTAL_FONT = Font(name="Calibri", size=11, bold=True)
TOTAL_GENERAL_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFFFF")


def _semaforo_cobranza_fill(dias_atraso):
    """Retorna PatternFill segun dias de atraso desde fecha de emision.
    < 30: None | 30-45: verde | 46-60: naranja | >= 61: rojo"""
    if dias_atraso is None or dias_atraso < 30:
        return None
    if dias_atraso <= 45:
        return COBRANZA_VERDE
    if dias_atraso <= 60:
        return COBRANZA_NARANJA
    return COBRANZA_ROJO


def _estado_documento(fecha_vto, report_date):
    """Estado del documento segun fecha de vencimiento vs fecha del reporte.
    Retorna 'Vencido', 'Por vencer' o 'Sin vencimiento'."""
    if fecha_vto is None:
        return "Sin vencimiento"
    if fecha_vto < report_date:
        return "Vencido"
    return "Por vencer"


def _build_cobranza_rows(db, vendedor_obuma_id, report_date):
    """Consulta y arma lista de dicts con datos por documento pendiente.
    Filtros: vendedor_id == vendedor_obuma_id, total_por_pagar > 0,
             tipo_documento in VALID_DOC_TYPES, no anuladas.
    NCs: POR PAGAR positivo (igual que Obuma — NC pendiente es saldo a favor del cliente).
    Orden: cliente asc, dias_atraso desc."""
    ventas = (
        db.query(VentaHistorico, ClienteFinal, Empleado)
        .outerjoin(ClienteFinal, VentaHistorico.cliente_id == ClienteFinal.id)
        .outerjoin(Empleado, Empleado.obuma_id == VentaHistorico.vendedor_id)
        .filter(
            VentaHistorico.vendedor_id == str(vendedor_obuma_id),
            VentaHistorico.tipo_documento.in_(VALID_DOC_TYPES),
            VentaHistorico.total_por_pagar > 0,
            func.coalesce(VentaHistorico.anulada, False) == False,
        )
        .all()
    )

    rows = []
    for venta, cliente, empleado in ventas:
        detalle = _parse_detalle_json(venta)
        fecha_emi = _get_emision_date(venta, detalle)
        fecha_vto = _get_vencimiento_date(venta, detalle)
        dias = (report_date - fecha_emi).days if fecha_emi else None

        es_nc = venta.tipo_documento in NC_DOC_TYPES
        es_nd = venta.tipo_documento in ND_DOC_TYPES
        por_pagar_real = float(venta.total_por_pagar or 0)
        # NC: se mantiene POSITIVO (igual que Obuma "Facturas por Cobrar").
        # ND: se mantiene POSITIVO (cargo adicional al cliente, igual que Factura).
        # Distincion visual: NC -> font italic rojo, ND -> font azul oscuro bold.

        estado = _estado_documento(fecha_vto, report_date)
        # Dias hasta/desde vencimiento (positivo = vencido, negativo = por vencer)
        dias_vto = (report_date - fecha_vto).days if fecha_vto else None

        rows.append({
            "documento": venta.tipo_documento or "",
            "folio": venta.folio or "",
            "fecha_emi": fecha_emi,
            "fecha_vto": fecha_vto,
            "dias_atraso": dias,
            "dias_vto": dias_vto,
            "estado": estado,
            "cliente_nombre": (cliente.nombre if cliente else None) or "(Sin cliente)",
            "cliente_rut": (cliente.rut if cliente else "") or "",
            "cliente_id": venta.cliente_id,
            "vendedor_nombre": (empleado.nombre if empleado else f"Vendedor {vendedor_obuma_id}"),
            "por_pagar": por_pagar_real,
            "es_nc": es_nc,
            "es_nd": es_nd,
        })

    # Orden: cliente asc, luego dias_atraso desc (None al final dentro del cliente)
    rows.sort(key=lambda r: (
        r["cliente_nombre"].lower(),
        -(r["dias_atraso"] if r["dias_atraso"] is not None else -10**9),
    ))
    return rows


# Definicion de rangos para el resumen (label, min_dias_vto, max_dias_vto)
# dias_vto = report_date - fecha_vto
# Vencido si dias_vto > 0, Por vencer si dias_vto <= 0
RANGOS_RESUMEN = [
    # Vencidos
    ("Vencido > 90 días",       91,    None),
    ("Vencido 61-90 días",      61,    90),
    ("Vencido 31-60 días",      31,    60),
    ("Vencido 1-30 días",       1,     30),
    # Por vencer (dias_vto <= 0 -> abs)
    ("Vence hoy",               0,     0),
    ("Por vencer 1-30 días",    -30,   -1),
    ("Por vencer 31-60 días",   -60,   -31),
    ("Por vencer 61-90 días",   -90,   -61),
    ("Por vencer > 90 días",    None,  -91),
]


def _build_cobranza_summary(rows, report_date):
    """Calcula resumen de cartera estilo Obuma.
    Retorna dict con totales, vencido/no vencido, y distribución por rangos."""
    total_a_cobrar = sum(r["por_pagar"] for r in rows)
    total_vencido = sum(r["por_pagar"] for r in rows if r["estado"] == "Vencido")
    total_por_vencer = sum(r["por_pagar"] for r in rows if r["estado"] == "Por vencer")
    total_sin_vcto = sum(r["por_pagar"] for r in rows if r["estado"] == "Sin vencimiento")

    pct_vencido = (total_vencido / total_a_cobrar * 100) if total_a_cobrar > 0 else 0
    pct_por_vencer = (total_por_vencer / total_a_cobrar * 100) if total_a_cobrar > 0 else 0
    pct_sin_vcto = (total_sin_vcto / total_a_cobrar * 100) if total_a_cobrar > 0 else 0

    # Distribución por rango de dias_vto
    distribucion = []
    for label, dmin, dmax in RANGOS_RESUMEN:
        monto = 0.0
        cant = 0
        for r in rows:
            d = r["dias_vto"]
            if d is None:
                continue
            ok_min = (dmin is None) or (d >= dmin)
            ok_max = (dmax is None) or (d <= dmax)
            if ok_min and ok_max:
                monto += r["por_pagar"]
                cant += 1
        pct = (monto / total_a_cobrar * 100) if total_a_cobrar > 0 else 0
        distribucion.append({"label": label, "monto": monto, "pct": pct, "cant": cant})

    # Sin vencimiento (ultimo bloque)
    cant_sin_vcto = sum(1 for r in rows if r["estado"] == "Sin vencimiento")
    distribucion.append({
        "label": "Sin vencimiento",
        "monto": total_sin_vcto,
        "pct": pct_sin_vcto,
        "cant": cant_sin_vcto,
    })

    return {
        "total_a_cobrar": total_a_cobrar,
        "total_vencido": total_vencido,
        "total_por_vencer": total_por_vencer,
        "total_sin_vcto": total_sin_vcto,
        "pct_vencido": pct_vencido,
        "pct_por_vencer": pct_por_vencer,
        "pct_sin_vcto": pct_sin_vcto,
        "distribucion": distribucion,
        "cant_docs": len(rows),
    }


TRACKED_VENDEDOR_IDS_RECON = [
    ("28856", "Gabriel"),
    ("28886", "Jhonatan"),
    ("28887", "Ernesto"),
    ("28891", "Pablo"),
    ("28892", "Jesus"),
]


def sync_for_report(db, scope: str = "report") -> dict:
    """Sync inmediato compartido para TODOS los flujos de envio de reportes.

    Ejecuta secuencialmente:
        1) sync_clientes
        2) sync_ventas
        3) sync_ventas_items_incremental(YYYY-01-01..hoy)
        4) sync_ventas_cobros (para que total_pagado/total_por_pagar
           reflejen los pagos del dia, importante para cartera y KPIs)

    Retorna dict {endpoint -> resultado}. Levanta RuntimeError si falla
    CUALQUIERA de los 4 syncs (todos son bloqueantes). Los cobros se incluyen
    como bloqueante porque sin ellos los saldos POR PAGAR de cartera no
    reflejan los pagos del dia y los KPIs de cobranza serian erroneos.
    Los flujos llamadores DEBEN abortar el envio si esta funcion lanza,
    para no mandar correos con datos parciales o desactualizados.

    Args:
        db: SQLAlchemy session.
        scope: etiqueta para logs (p.ej. "Cartera/Cobranza", "Reporte Diario",
               "Reporte Semanal", "Reporte Programado").
    """
    import asyncio
    import time
    from datetime import date as _date
    from src.etl.sync_service import SyncService

    today = _date.today()
    fecha_desde = _date(today.year, 1, 1).strftime("%Y-%m-%d")
    fecha_hasta = today.strftime("%Y-%m-%d")

    service = SyncService(db)
    # Creamos un loop dedicado pero NO lo seteamos como default del thread:
    # asi evitamos dejar un loop cerrado como event_loop por defecto que
    # rompa llamadas posteriores a asyncio.get_event_loop() en el mismo
    # thread (p.ej. dentro de _generate_and_send_individual_reports tras
    # este sync).
    loop = asyncio.new_event_loop()
    results = {}
    t0 = time.time()
    try:
        logger.info(f"{scope}: iniciando sync inmediato (clientes + ventas + items + cobros)...")

        t = time.time()
        results["clientes"] = loop.run_until_complete(service.sync_clientes())
        if isinstance(results["clientes"], dict) and "error" in results["clientes"]:
            raise RuntimeError(f"sync_clientes fallo: {results['clientes'].get('error')}")
        logger.info(f"{scope} sync clientes: {results['clientes'].get('synced', '?')} regs en {time.time()-t:.1f}s")

        t = time.time()
        results["ventas"] = loop.run_until_complete(service.sync_ventas())
        if isinstance(results["ventas"], dict) and "error" in results["ventas"]:
            raise RuntimeError(f"sync_ventas fallo: {results['ventas'].get('error')}")
        logger.info(f"{scope} sync ventas: {results['ventas'].get('synced', '?')} regs en {time.time()-t:.1f}s")

        t = time.time()
        results["ventas_items"] = loop.run_until_complete(
            service.sync_ventas_items_incremental(fecha_desde, fecha_hasta)
        )
        if isinstance(results["ventas_items"], dict) and "error" in results["ventas_items"]:
            raise RuntimeError(
                f"sync_ventas_items_incremental fallo: {results['ventas_items'].get('error')}"
            )
        logger.info(
            f"{scope} sync ventas_items {fecha_desde}..{fecha_hasta}: "
            f"{results['ventas_items'].get('synced', '?')} items en {time.time()-t:.1f}s"
        )

        t = time.time()
        results["ventas_cobros"] = loop.run_until_complete(service.sync_ventas_cobros())
        if isinstance(results["ventas_cobros"], dict) and "error" in results["ventas_cobros"]:
            # cobros es BLOQUEANTE: sin pagos del dia los saldos POR PAGAR de
            # cartera no reflejan la realidad y los KPIs de cobranza serian
            # erroneos. Levantamos RuntimeError para que el flujo aborte el
            # envio (mismo criterio que clientes/ventas/items).
            raise RuntimeError(
                f"sync_ventas_cobros fallo: {results['ventas_cobros'].get('error')}"
            )
        logger.info(
            f"{scope} sync ventas_cobros: "
            f"{results['ventas_cobros'].get('synced', '?')} regs en {time.time()-t:.1f}s"
        )

        logger.info(f"{scope} sync inmediato OK en {time.time()-t0:.1f}s totales")
        return results
    finally:
        try:
            loop.close()
        except Exception:
            pass


def log_reconciliation_per_vendor(db, today=None, scope: str = "Reporte") -> None:
    """Loguea bloque RECONCILIACION con totales de cartera por vendedor
    trackeado, comparando contra OBUMA_REFERENCE_TOTALS. Reutilizable por
    todos los flujos de reportes (cartera, ventas, semanal).

    No levanta excepciones — fallos se loguean pero no abortan el flujo.
    """
    from datetime import date as _date
    if today is None:
        today = _date.today()
    try:
        logger.info(
            f"{scope} RECONCILIACION (post-sync vs pantalla 'Facturas por Cobrar' de Obuma):"
        )
        ok_count = 0
        warn_count = 0
        for vid, name in TRACKED_VENDEDOR_IDS_RECON:
            _rows = _build_cobranza_rows(db, vid, today)
            _summary = _build_cobranza_summary(_rows, today)
            excel_total = _summary["total_a_cobrar"]
            obuma_ref = OBUMA_REFERENCE_TOTALS.get(vid)
            if obuma_ref is None or obuma_ref == 0:
                logger.info(
                    f"  [{vid}] {name}: docs={_summary['cant_docs']} "
                    f"excel_total=${excel_total:,.0f} "
                    f"vencido=${_summary['total_vencido']:,.0f} "
                    f"({_summary['pct_vencido']:.1f}%) "
                    f"obuma_ref=N/A (sin referencia configurada)"
                )
            else:
                diff = excel_total - obuma_ref
                pct_diff = (diff / obuma_ref) * 100 if obuma_ref else 0.0
                estado = "OK" if abs(pct_diff) < 0.5 else "REVISAR"
                if estado == "OK":
                    ok_count += 1
                else:
                    warn_count += 1
                logger.info(
                    f"  [{vid}] {name}: docs={_summary['cant_docs']} "
                    f"excel_total=${excel_total:,.0f} "
                    f"obuma_ref=${obuma_ref:,.0f} "
                    f"diff=${diff:+,.0f} ({pct_diff:+.2f}%) "
                    f"vencido=${_summary['total_vencido']:,.0f} "
                    f"({_summary['pct_vencido']:.1f}%) "
                    f"=> {estado}"
                )
        if ok_count + warn_count > 0:
            logger.info(
                f"{scope} RECONCILIACION resumen: "
                f"{ok_count} vendedor(es) cuadran (<0.5%), "
                f"{warn_count} requieren revision."
            )
    except Exception as _e:
        logger.warning(f"No se pudo loguear reconciliacion por vendedor: {_e}")


def _sync_for_cartera_report(db) -> dict:
    """Compat wrapper: hace sync inmediato + reconciliacion para Cartera.
    Mantener para no romper callers existentes."""
    from datetime import date as _date
    results = sync_for_report(db, scope="Cartera/Cobranza")
    log_reconciliation_per_vendor(db, _date.today(), scope="Cartera/Cobranza")
    return results


def _check_cartera_data_freshness(db, max_hours: float = 2.0) -> None:
    """Logea WARNING si la ultima sincronizacion de Ventas tiene mas de
    max_hours horas. Cubre el caso en que alguien llame
    generate_cartera_cobranza_report directamente bypaseando el batch
    wrapper que hace sync inmediato.
    """
    try:
        from datetime import datetime, timedelta
        from src.models.models import ObumaApiEndpoint
        ep = (
            db.query(ObumaApiEndpoint)
            .filter(ObumaApiEndpoint.nombre.ilike("%Ventas%"))
            .filter(~ObumaApiEndpoint.nombre.ilike("%Cobros%"))
            .filter(~ObumaApiEndpoint.nombre.ilike("%Cotizaciones%"))
            .filter(~ObumaApiEndpoint.nombre.ilike("%>%"))
            .filter(ObumaApiEndpoint.ultima_sync.isnot(None))
            .order_by(ObumaApiEndpoint.ultima_sync.desc())
            .first()
        )
        if ep is None or ep.ultima_sync is None:
            logger.warning(
                "Cartera: no hay registro de ultima sincronizacion de Ventas. "
                "El reporte puede contener datos desactualizados."
            )
            return
        # ultima_sync se almacena con datetime.now() (hora local del server,
        # configurada como America/Santiago). Comparamos con la misma convencion
        # para evitar offset por timezone.
        age = datetime.now() - ep.ultima_sync
        if age > timedelta(hours=max_hours):
            logger.warning(
                f"Cartera: la ultima sincronizacion de Ventas tiene "
                f"{age.total_seconds()/3600:.1f}h (>{max_hours}h). "
                f"El reporte puede no incluir facturas emitidas hoy. "
                f"Recomendado: generar via generate_all_cartera_cobranza_reports(do_sync=True)."
            )
    except Exception as _e:
        logger.debug(f"No se pudo verificar staleness de cartera: {_e}")


def generate_cartera_cobranza_report(db: Session, vendedor_obuma_id, report_date: date = None):
    """Genera Excel de cartera/cobranza para UN vendedor.

    Retorna filepath del Excel generado, o None si el vendedor no tiene
    documentos pendientes (en cuyo caso no se debe enviar email).

    Estructura:
    - Bloque RESUMEN al inicio: total a cobrar, vencido, no vencido,
      sin vencimiento y distribución por rangos de días desde vencimiento.
    - Detalle: columnas DOCUMENTO, FOLIO, FECHA, FECHA VCTO, ESTADO,
      FECHA HOY, DÍAS ATRASO, CLIENTE, CLIENTE RUT, VENDEDOR, POR PAGAR.
    - Agrupado por cliente con subtotal en gris + TOTAL GENERAL al final.
    - Semaforo por dias desde emision: 30-45 verde, 46-60 naranja, 61+ rojo.
    - NCs: POR PAGAR positivo (igual que Obuma), font italic rojo como indicador.
    - Sin vencimiento: texto en gris cursiva.
    - Total = suma directa de POR PAGAR (sin restar adicional las NCs).
    """
    if report_date is None:
        report_date = date.today()

    # Si esta funcion se llama directamente (sin pasar por
    # generate_all_cartera_cobranza_reports que hace sync inmediato),
    # avisar si los datos pueden estar desactualizados.
    _check_cartera_data_freshness(db)

    vendedor_obuma_id = str(vendedor_obuma_id)
    rows = _build_cobranza_rows(db, vendedor_obuma_id, report_date)

    if not rows:
        logger.info(
            f"Cartera/Cobranza: vendedor {vendedor_obuma_id} sin documentos pendientes en {report_date}. Se omite generacion."
        )
        return None

    empleado = (
        db.query(Empleado).filter(Empleado.obuma_id == vendedor_obuma_id).first()
    )
    vendedor_nombre = (empleado.nombre if empleado else rows[0]["vendedor_nombre"]) or f"Vendedor {vendedor_obuma_id}"

    summary = _build_cobranza_summary(rows, report_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera Cobranza"

    n_cols = len(COBRANZA_HEADERS)  # 11

    # Titulo
    ws.cell(row=1, column=1, value=f"Reporte Cartera Cobranza - {vendedor_nombre}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"Fecha reporte: {report_date.strftime('%d/%m/%Y')}").font = SUBTITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    # ── BLOQUE RESUMEN ──
    def _section_header(row, text):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)
        c = ws.cell(row=row, column=1, value=text)
        c.font = RESUMEN_TITLE_FONT
        c.fill = RESUMEN_TITLE_FILL
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row].height = 22

    def _kpi_row(row, label, monto, pct=None, total_font=False, bg_fill=None):
        # Layout: cols 1-4 label, cols 5-7 monto, cols 8-11 pct
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        ws.merge_cells(start_row=row, start_column=8, end_row=row, end_column=n_cols)
        c_label = ws.cell(row=row, column=1, value=label)
        c_monto = ws.cell(row=row, column=5, value=monto)
        c_pct = ws.cell(row=row, column=8, value=(f"{pct:.1f}%" if pct is not None else ""))
        c_label.font = RESUMEN_TOTAL_FONT if total_font else RESUMEN_LABEL_FONT
        c_monto.font = RESUMEN_TOTAL_FONT if total_font else RESUMEN_VALUE_FONT
        c_pct.font = RESUMEN_TOTAL_FONT if total_font else RESUMEN_VALUE_FONT
        c_monto.number_format = CURRENCY_FORMAT
        c_label.alignment = Alignment(horizontal="left", indent=1)
        c_monto.alignment = Alignment(horizontal="right")
        c_pct.alignment = Alignment(horizontal="right", indent=1)
        if bg_fill is not None:
            for col in range(1, n_cols + 1):
                ws.cell(row=row, column=col).fill = bg_fill

    rcur = 4
    _section_header(rcur, "RESUMEN CARTERA")
    rcur += 1
    _kpi_row(rcur, f"Total ventas a cobrar ({summary['cant_docs']} docs)",
             summary["total_a_cobrar"], None, total_font=True, bg_fill=RESUMEN_FILL)
    rcur += 1
    _kpi_row(rcur, "  Total vencido", summary["total_vencido"], summary["pct_vencido"])
    rcur += 1
    _kpi_row(rcur, "  Total no vencido (por vencer)", summary["total_por_vencer"], summary["pct_por_vencer"])
    rcur += 1
    _kpi_row(rcur, "  Sin fecha de vencimiento", summary["total_sin_vcto"], summary["pct_sin_vcto"])
    rcur += 2

    # ── DISTRIBUCIÓN POR RANGOS ──
    _section_header(rcur, "DISTRIBUCIÓN POR DÍAS DE VENCIMIENTO")
    rcur += 1
    # Header table: col1-4 Rango, col5 Docs, col6-8 Monto, col9-11 %
    ws.merge_cells(start_row=rcur, start_column=1, end_row=rcur, end_column=4)
    ws.merge_cells(start_row=rcur, start_column=6, end_row=rcur, end_column=8)
    ws.merge_cells(start_row=rcur, start_column=9, end_row=rcur, end_column=n_cols)
    headers_dist = [(1, "Rango"), (5, "Docs"), (6, "Monto"), (9, "% del Total")]
    # Aplicar fill, font y border a todas las celdas del header
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=rcur, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
    for col, txt in headers_dist:
        c = ws.cell(row=rcur, column=col, value=txt)
        c.alignment = HEADER_ALIGNMENT
    rcur += 1

    for d in summary["distribucion"]:
        ws.merge_cells(start_row=rcur, start_column=1, end_row=rcur, end_column=4)
        ws.merge_cells(start_row=rcur, start_column=6, end_row=rcur, end_column=8)
        ws.merge_cells(start_row=rcur, start_column=9, end_row=rcur, end_column=n_cols)
        c_lbl = ws.cell(row=rcur, column=1, value=d["label"])
        c_cnt = ws.cell(row=rcur, column=5, value=d["cant"])
        c_mon = ws.cell(row=rcur, column=6, value=d["monto"])
        c_pct = ws.cell(row=rcur, column=9, value=f"{d['pct']:.1f}%")
        c_lbl.alignment = Alignment(horizontal="left", indent=1)
        c_cnt.alignment = Alignment(horizontal="center")
        c_mon.alignment = Alignment(horizontal="right")
        c_pct.alignment = Alignment(horizontal="right", indent=1)
        c_mon.number_format = CURRENCY_FORMAT
        # Coloreado segun tipo
        lbl = d["label"]
        if "Vencido" in lbl and "Por vencer" not in lbl:
            for col in range(1, n_cols + 1):
                ws.cell(row=rcur, column=col).fill = COBRANZA_ROJO if "> 90" in lbl or "61-90" in lbl else (
                    COBRANZA_NARANJA if "31-60" in lbl else COBRANZA_VERDE
                )
        elif lbl == "Vence hoy":
            for col in range(1, n_cols + 1):
                ws.cell(row=rcur, column=col).fill = COBRANZA_NARANJA
        elif lbl == "Sin vencimiento":
            for col in range(1, n_cols + 1):
                ws.cell(row=rcur, column=col).fill = COBRANZA_GRIS_SUBTOTAL
        for col in range(1, n_cols + 1):
            ws.cell(row=rcur, column=col).border = THIN_BORDER
        rcur += 1

    rcur += 2

    # ── DETALLE POR DOCUMENTO ──
    _section_header(rcur, "DETALLE POR DOCUMENTO")
    rcur += 1
    header_row = rcur
    for i, h in enumerate(COBRANZA_HEADERS, 1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header(ws, header_row, n_cols)

    current_row = header_row + 1
    current_cliente = None
    cliente_subtotal = 0.0
    grand_total = 0.0

    # Indices de columna (1-based) tras agregar ESTADO en posición 5
    COL_FECHA_EMI = 3
    COL_FECHA_VTO = 4
    COL_ESTADO = 5
    COL_FECHA_HOY = 6
    COL_DIAS_ATRASO = 7
    COL_VENDEDOR = 10
    COL_POR_PAGAR = 11

    def _flush_cliente_subtotal():
        """Escribe fila de subtotal para el cliente recien terminado."""
        nonlocal current_row
        if current_cliente is None:
            return
        for col in range(1, n_cols + 1):
            cell = ws.cell(row=current_row, column=col)
            cell.fill = COBRANZA_GRIS_SUBTOTAL
            cell.border = THIN_BORDER
            cell.font = SUBTOTAL_FONT
        ws.cell(row=current_row, column=COL_VENDEDOR, value=f"Subtotal {current_cliente}").alignment = Alignment(horizontal="right")
        cell_sub = ws.cell(row=current_row, column=COL_POR_PAGAR, value=cliente_subtotal)
        cell_sub.number_format = CURRENCY_FORMAT
        current_row += 1

    for r in rows:
        # Cambio de cliente -> escribir subtotal del anterior
        if current_cliente is not None and r["cliente_nombre"] != current_cliente:
            _flush_cliente_subtotal()
            cliente_subtotal = 0.0

        if current_cliente != r["cliente_nombre"]:
            current_cliente = r["cliente_nombre"]

        # Fila de datos
        fila_fill = _semaforo_cobranza_fill(r["dias_atraso"])
        values = [
            r["documento"],
            r["folio"],
            r["fecha_emi"].strftime("%d-%m-%Y") if r["fecha_emi"] else "",
            r["fecha_vto"].strftime("%d-%m-%Y") if r["fecha_vto"] else "Sin vencimiento",
            r["estado"],
            report_date.strftime("%d-%m-%Y"),
            r["dias_atraso"] if r["dias_atraso"] is not None else "",
            r["cliente_nombre"],
            r["cliente_rut"],
            r["vendedor_nombre"],
            r["por_pagar"],
        ]
        for i, val in enumerate(values, 1):
            cell = ws.cell(row=current_row, column=i, value=val)
            cell.border = THIN_BORDER
            if fila_fill is not None:
                cell.fill = fila_fill
            # FECHA VCTO "Sin vencimiento" estilizado
            if i == COL_FECHA_VTO and val == "Sin vencimiento":
                cell.font = SIN_VCTO_FONT
            # ESTADO con coloreado segun valor
            if i == COL_ESTADO:
                cell.alignment = Alignment(horizontal="center")
                if val == "Vencido":
                    cell.font = ESTADO_VENCIDO_FONT
                elif val == "Por vencer":
                    cell.font = ESTADO_PORVENCER_FONT
                else:
                    cell.font = ESTADO_SINVCTO_FONT
            # POR PAGAR formato moneda
            if i == COL_POR_PAGAR:
                cell.number_format = CURRENCY_FORMAT
                if r["es_nc"]:
                    cell.font = NC_FONT
                elif r.get("es_nd"):
                    cell.font = ND_FONT
            # Centrar columnas de fechas y dias
            if i in (COL_FECHA_EMI, COL_FECHA_VTO, COL_FECHA_HOY, COL_DIAS_ATRASO):
                cell.alignment = Alignment(horizontal="center")

        cliente_subtotal += r["por_pagar"]
        grand_total += r["por_pagar"]
        current_row += 1

    # Subtotal del ultimo cliente
    _flush_cliente_subtotal()

    # Fila TOTAL GENERAL
    current_row += 1
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=current_row, column=col)
        cell.fill = COBRANZA_GRIS_TOTAL
        cell.border = THIN_BORDER
        cell.font = TOTAL_GENERAL_FONT
    ws.cell(row=current_row, column=COL_VENDEDOR, value="TOTAL GENERAL").alignment = Alignment(horizontal="right")
    cell_total = ws.cell(row=current_row, column=COL_POR_PAGAR, value=grand_total)
    cell_total.number_format = CURRENCY_FORMAT
    cell_total.font = TOTAL_GENERAL_FONT
    cell_total.fill = COBRANZA_GRIS_TOTAL

    # Anchos de columna (11 columnas)
    col_widths = [16, 10, 12, 14, 14, 12, 12, 38, 14, 24, 16]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze panes (fija desde fin de header de detalle)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate

    os.makedirs("reports", exist_ok=True)
    safe_name = (vendedor_nombre or f"vendedor_{vendedor_obuma_id}").replace(" ", "_").replace("/", "_")
    filename = f"cartera_cobranza_{safe_name}_{report_date.strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join("reports", filename)
    wb.save(filepath)

    try:
        reporte = ReporteGenerado(
            nombre_archivo=filename,
            tipo="cartera_cobranza",
            fecha_reporte=report_date,
            ruta_archivo=filepath,
        )
        db.add(reporte)
        db.commit()
    except Exception as e:
        logger.warning(f"No se pudo registrar ReporteGenerado: {e}")
        db.rollback()

    logger.info(
        f"Cartera/Cobranza generado: {filepath} ({len(rows)} docs, total ${grand_total:,.0f})"
    )
    return filepath


def generate_all_cartera_cobranza_reports(db: Session, report_date: date = None, do_sync: bool = True):
    """Genera reportes de cartera/cobranza para TODOS los vendedores trackeados.

    Args:
        db: Session DB.
        report_date: fecha del reporte (default hoy).
        do_sync: si True (default), ejecuta sync inmediato (ventas + ventas_items
                 incremental año actual + clientes) antes de generar el Excel,
                 para que las facturas emitidas hoy aparezcan. Si el sync falla,
                 ABORTA la generación (no se envían reportes con datos viejos).

    Vendedores sin saldo pendiente son omitidos (no se genera Excel).
    Retorna lista de tuplas (vendedor_obuma_id, filepath_o_None).
    """
    if report_date is None:
        report_date = date.today()

    if do_sync:
        try:
            _sync_for_cartera_report(db)
        except Exception as e:
            logger.error(
                f"Cartera/Cobranza ABORTADO: sync inmediato falló ({e}). "
                f"No se generan reportes para evitar enviar datos desactualizados.",
                exc_info=True,
            )
            return []

    tracked_ids = ['28856', '28886', '28887', '28891', '28892']
    results = []
    for vid in tracked_ids:
        try:
            fp = generate_cartera_cobranza_report(db, vid, report_date)
            results.append((vid, fp))
        except Exception as e:
            logger.error(f"Error generando cartera/cobranza vendedor {vid}: {e}", exc_info=True)
            results.append((vid, None))

    generados = sum(1 for _, fp in results if fp)
    logger.info(
        f"Cartera/Cobranza batch completo: {generados}/{len(tracked_ids)} vendedores con docs pendientes"
    )
    return results
